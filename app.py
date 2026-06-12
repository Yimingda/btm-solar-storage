"""
 BTM v3.0
Professional BTM PV+BESS Financial Modelling System
South Africa C&I Megaflex / Miniflex Tariff 2025/26 | 8760-Hour Physical Dispatch Engine
Section 12B Accelerated Depreciation | Huawei SA Digital Power
"""

import streamlit as st
import pandas as pd
import numpy as np
import requests
import io
import re
import warnings

# ── Auto-scale unit helper ────────────────────────────────────────────────────
_MW_UP = {"kWp": "MWp", "kWh": "MWh", "kW": "MW", "kWh/yr": "MWh/yr"}

def _fmw(v: float, unit: str = "kWp") -> str:
    """Return formatted string: auto-upgrades to MW/MWh/MWp when v ≥ 1 000."""
    if v >= 1000:
        return f"{v / 1000:,.1f} {_MW_UP.get(unit, unit)}"
    return f"{v:,.0f} {unit}"
from datetime import date as _date, timedelta
warnings.filterwarnings("ignore")

# ── Auth / User system ──────────────────────────────────────────────────────
# Imported here so auth helpers are available throughout app.py
from auth import (is_logged_in, render_auth_gate, get_current_user,
                  get_tier, is_pro, is_admin, logout, flush_token_to_storage)
from snapshots import render_snapshot_panel, render_project_bar
from scenario_select import render_scenario_selector, render_wheeling_placeholder
from admin import render_admin_panel

# SA 2025 Public Holidays (off-peak all day like weekends)
SA_PUBLIC_HOLIDAYS_2025 = {
    _date(2025, 1, 1),   # New Year's Day
    _date(2025, 3, 21),  # Human Rights Day
    _date(2025, 4, 18),  # Good Friday
    _date(2025, 4, 21),  # Family Day
    _date(2025, 4, 27),  # Freedom Day
    _date(2025, 5, 1),   # Workers' Day
    _date(2025, 6, 16),  # Youth Day
    _date(2025, 8, 9),   # National Women's Day
    _date(2025, 9, 24),  # Heritage Day
    _date(2025, 12, 16), # Day of Reconciliation
    _date(2025, 12, 25), # Christmas Day
    _date(2025, 12, 26), # Day of Goodwill
}

# eThekwini 2025/26 public holidays — treated as "saturday" or "sunday" TOU level
# (Eskom treats all holidays as Sunday; eThekwini distinguishes between the two)
# Source: eThekwini Tariff Booklet 25/26 p21
ETHEKWINI_HOLIDAYS_2025_26: dict[_date, str] = {
    # ── Sunday-level (no peaks at all) ──
    _date(2025, 4, 18): "sunday",   # Good Friday
    _date(2025, 4, 21): "sunday",   # Family Day
    _date(2025, 4, 27): "sunday",   # Freedom Day (falls on Sunday anyway)
    _date(2025, 12, 25): "sunday",  # Christmas Day
    _date(2025, 12, 26): "sunday",  # Day of Goodwill
    _date(2026, 1, 1):  "sunday",   # New Year's Day
    _date(2026, 4, 3):  "sunday",   # Good Friday
    _date(2026, 4, 6):  "sunday",   # Family Day
    _date(2026, 4, 27): "sunday",   # Freedom Day
    # ── Saturday-level (standard during some windows, no peak) ──
    _date(2025, 4, 28): "saturday", # Public Holiday (day after Freedom Day)
    _date(2025, 5, 1):  "saturday", # Workers' Day
    _date(2025, 6, 16): "saturday", # Youth Day
    _date(2025, 8, 9):  "saturday", # National Women's Day
    _date(2025, 9, 24): "saturday", # Heritage Day
    _date(2025, 12, 16):"saturday", # Day of Reconciliation
    _date(2026, 3, 21): "saturday", # Human Rights Day
    _date(2026, 5, 1):  "saturday", # Worker's Day
    _date(2026, 6, 16): "saturday", # Youth Day
}

try:
    import folium
    from streamlit_folium import st_folium
    MAP_AVAILABLE = True
except ImportError:
    MAP_AVAILABLE = False

# ─────────────────────────────────────────────────────────────
# Page Configuration
# ─────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="BTM PV+BESS Financial Modelling System",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ─────────────────────────────────────────────────────────────
# Global Styles
# ─────────────────────────────────────────────────────────────
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=IBM+Plex+Sans:wght@300;400;600;700&display=swap');

    /* ── Design tokens ─────────────────────────────────────────── */
    :root {
        --primary:    #00E5A0;
        --secondary:  #FF6B35;
        --accent:     #4ECDC4;
        --bg-dark:    #0A0E1A;
        --bg-card:    #111827;
        --bg-input:   #1C2333;
        --text-main:  #E8ECF0;
        --text-dim:   #8B95A3;
        --border:     #2D3748;
        --border-hi:  #3D4F6A;
        --warning:    #F6C90E;
        --danger:     #FF4444;
        --radius:     6px;
    }

    /* ── Base ──────────────────────────────────────────────────── */
    html, body, .stApp {
        background-color: var(--bg-dark) !important;
        color: var(--text-main) !important;
        font-family: 'IBM Plex Sans', sans-serif !important;
    }
    .block-container {
        padding-top: 0 !important;
        padding-bottom: 1rem !important;
        max-width: 100% !important;
    }
    /* Hide ALL Streamlit chrome — header bar, toolbar, deploy btn, viewer badge */
    #MainMenu,
    footer,
    header[data-testid="stHeader"],
    [data-testid="stToolbar"],
    [data-testid="stDecoration"],
    [data-testid="stStatusWidget"],
    [data-testid="collapsedControl"],
    button[data-testid="baseButton-headerNoPadding"] { display: none !important; }

    /* ── App header row ────────────────────────────────────────── */
    div[data-testid="stHorizontalBlock"]:first-of-type {
        background: linear-gradient(180deg, #0D1220 0%, #0A0E1A 100%);
        border-bottom: 1px solid var(--primary);
        padding: 0.3rem 0;
        margin-bottom: 0.6rem;
    }
    div[data-testid="stHorizontalBlock"]:first-of-type div[data-testid="stColumn"] {
        padding-top: 0.2rem !important;
        padding-bottom: 0.2rem !important;
    }

    /* ── Custom components ─────────────────────────────────────── */
    .main-header {
        display: flex;
        align-items: center;
        gap: 0.75rem;
        padding: 0.4rem 0;
    }
    .main-title {
        font-family: 'IBM Plex Mono', monospace;
        font-size: 1.25rem;
        font-weight: 600;
        color: var(--primary);
        letter-spacing: 0.04em;
        margin: 0;
        line-height: 1.2;
    }
    .sub-title {
        font-size: 0.65rem;
        color: var(--text-dim);
        font-family: 'IBM Plex Mono', monospace;
        letter-spacing: 0.07em;
        margin-top: 2px;
    }

    /* ── Metric cards ──────────────────────────────────────────── */
    .metric-card {
        background: var(--bg-card);
        border: 1px solid var(--border);
        border-left: 3px solid var(--primary);
        border-radius: var(--radius);
        padding: 0.65rem 0.9rem;
        margin-bottom: 0.5rem;
    }
    .metric-label {
        font-size: 0.65rem;
        color: var(--text-dim);
        text-transform: uppercase;
        letter-spacing: 0.1em;
        font-family: 'IBM Plex Mono', monospace;
    }
    .metric-value {
        font-size: 1.35rem;
        font-weight: 700;
        color: var(--primary);
        font-family: 'IBM Plex Mono', monospace;
        line-height: 1.2;
    }
    .metric-unit {
        font-size: 0.65rem;
        color: var(--text-dim);
        font-family: 'IBM Plex Mono', monospace;
    }

    /* ── Info / warning / success boxes ───────────────────────── */
    .info-box {
        background: rgba(78,205,196,0.07);
        border: 1px solid rgba(78,205,196,0.4);
        border-left: 3px solid var(--accent);
        border-radius: var(--radius);
        padding: 0.55rem 0.9rem;
        font-size: 0.76rem;
        color: var(--accent);
        font-family: 'IBM Plex Mono', monospace;
        line-height: 1.55;
    }
    .warning-box {
        background: rgba(246,201,14,0.08);
        border: 1px solid rgba(246,201,14,0.45);
        border-left: 3px solid var(--warning);
        border-radius: var(--radius);
        padding: 0.55rem 0.9rem;
        font-size: 0.76rem;
        color: var(--warning);
        font-family: 'IBM Plex Mono', monospace;
        line-height: 1.55;
    }
    .success-box {
        background: rgba(0,229,160,0.07);
        border: 1px solid rgba(0,229,160,0.4);
        border-left: 3px solid var(--primary);
        border-radius: var(--radius);
        padding: 0.55rem 0.9rem;
        font-size: 0.76rem;
        color: var(--primary);
        font-family: 'IBM Plex Mono', monospace;
        line-height: 1.55;
    }

    /* ── Section header divider ────────────────────────────────── */
    .section-header {
        font-family: 'IBM Plex Mono', monospace;
        font-size: 0.72rem;
        color: var(--primary);
        text-transform: uppercase;
        letter-spacing: 0.14em;
        border-bottom: 1px solid var(--border);
        padding-bottom: 0.35rem;
        margin: 0.9rem 0 0.5rem 0;
    }

    /* ── Derived / read-only value display ─────────────────────── */
    .derived-value {
        background: rgba(0,229,160,0.05);
        border: 1px solid rgba(0,229,160,0.25);
        border-radius: 4px;
        padding: 0.28rem 0.55rem;
        font-family: 'IBM Plex Mono', monospace;
        font-size: 0.8rem;
        color: var(--primary);
        text-align: center;
        margin-bottom: 0.35rem;
    }
    .param-label {
        font-size: 0.7rem;
        color: var(--text-dim);
        font-family: 'IBM Plex Mono', monospace;
        margin-bottom: 0.1rem;
    }

    /* ── Tariff season markers ─────────────────────────────────── */
    .tariff-winter { border-left: 3px solid var(--accent);   padding-left: 0.5rem; margin-bottom: 0.3rem; }
    .tariff-summer { border-left: 3px solid var(--warning);  padding-left: 0.5rem; margin-bottom: 0.3rem; }

    /* ── EoL badge ─────────────────────────────────────────────── */
    .eol-badge {
        display: inline-block;
        background: rgba(255,107,53,0.12);
        border: 1px solid var(--secondary);
        color: var(--secondary);
        font-family: 'IBM Plex Mono', monospace;
        font-size: 0.65rem;
        padding: 0.12rem 0.45rem;
        border-radius: 20px;
        vertical-align: middle;
    }

    /* ── Tabs ──────────────────────────────────────────────────── */
    .stTabs [data-baseweb="tab-list"] {
        gap: 2px;
        background: transparent;
        border-bottom: 1px solid var(--border);
        padding-bottom: 0;
    }
    .stTabs [data-baseweb="tab"] {
        font-family: 'IBM Plex Mono', monospace;
        font-size: 0.76rem;
        color: var(--text-dim);
        background: transparent;
        border-radius: 4px 4px 0 0;
        padding: 0.4rem 0.75rem;
        border: none !important;
        transition: color 0.15s;
    }
    .stTabs [data-baseweb="tab"]:hover { color: var(--text-main); }
    .stTabs [data-baseweb="tab"][aria-selected="true"] {
        color: var(--primary) !important;
        background: rgba(0,229,160,0.08) !important;
        border-bottom: 2px solid var(--primary) !important;
    }
    .stTabs [data-baseweb="tab-highlight"] { display: none !important; }
    .stTabs [data-baseweb="tab-border"]    { display: none !important; }

    /* ── Buttons (supports both legacy .stButton and new data-testid) ── */
    .stButton > button,
    button[data-testid="stBaseButton-secondary"] {
        background:      transparent !important;
        border:          1px solid var(--border) !important;
        color:           var(--text-dim) !important;
        font-family:     'IBM Plex Mono', monospace !important;
        font-size:       0.73rem !important;
        letter-spacing:  0.05em !important;
        border-radius:   4px !important;
        transition:      background 0.15s, color 0.15s, box-shadow 0.15s !important;
        padding:         0.28rem 0.75rem !important;
    }
    .stButton > button:hover,
    button[data-testid="stBaseButton-secondary"]:hover {
        background:   rgba(255,255,255,0.06) !important;
        border-color: var(--primary) !important;
        color:        var(--text-main) !important;
        box-shadow:   none !important;
    }
    /* Primary buttons — green fill */
    .stButton > button[kind="primary"],
    button[data-testid="stBaseButton-primary"] {
        background:  var(--primary) !important;
        color:       var(--bg-dark) !important;
        border:      none !important;
        font-weight: 600 !important;
    }
    .stButton > button[kind="primary"]:hover,
    button[data-testid="stBaseButton-primary"]:hover {
        box-shadow: 0 0 12px rgba(0,229,160,0.5) !important;
        filter:     brightness(1.08) !important;
    }
    /* Disabled buttons */
    button[data-testid="stBaseButton-secondary"]:disabled {
        opacity:    0.4 !important;
        cursor:     not-allowed !important;
    }
    .stDownloadButton > button {
        background: transparent;
        border: 1px solid var(--accent);
        color: var(--accent);
        font-family: 'IBM Plex Mono', monospace;
        font-size: 0.73rem;
        border-radius: 4px;
        transition: background 0.15s, color 0.15s;
    }
    .stDownloadButton > button:hover {
        background: var(--accent);
        color: var(--bg-dark);
    }

    /* ── Number inputs ─────────────────────────────────────────── */
    div[data-testid="stNumberInput"] input,
    div[data-testid="stTextInput"]   input {
        font-family: 'IBM Plex Mono', monospace !important;
        font-size: 0.82rem !important;
        background: var(--bg-input) !important;
        color: var(--text-main) !important;
        border: 1px solid var(--border) !important;
        border-radius: 4px !important;
    }
    div[data-testid="stNumberInput"] input:focus,
    div[data-testid="stTextInput"]   input:focus {
        border-color: var(--primary) !important;
        box-shadow: 0 0 0 2px rgba(0,229,160,0.15) !important;
        outline: none !important;
    }
    div[data-testid="stNumberInput"] label,
    div[data-testid="stTextInput"]   label,
    div[data-testid="stSelectbox"]   label {
        font-family: 'IBM Plex Mono', monospace !important;
        font-size: 0.71rem !important;
        color: var(--text-dim) !important;
        letter-spacing: 0.04em;
    }

    /* ── Selectbox ─────────────────────────────────────────────── */
    div[data-testid="stSelectbox"] > div > div {
        background: var(--bg-input) !important;
        border: 1px solid var(--border) !important;
        border-radius: 4px !important;
        color: var(--text-main) !important;
        font-family: 'IBM Plex Mono', monospace !important;
        font-size: 0.8rem !important;
        transition: border-color 0.15s;
    }
    div[data-testid="stSelectbox"] > div > div:focus-within {
        border-color: var(--primary) !important;
        box-shadow: 0 0 0 2px rgba(0,229,160,0.15) !important;
    }

    /* ── Expanders ─────────────────────────────────────────────── */
    details[data-testid="stExpander"] > summary {
        font-family: 'IBM Plex Mono', monospace !important;
        font-size: 0.77rem !important;
        color: var(--text-main) !important;
        background: var(--bg-card) !important;
        border: 1px solid var(--border) !important;
        border-radius: 4px !important;
        padding: 0.4rem 0.75rem !important;
        transition: border-color 0.15s, background 0.15s;
    }
    details[data-testid="stExpander"][open] > summary {
        border-bottom-left-radius: 0 !important;
        border-bottom-right-radius: 0 !important;
        border-color: var(--border-hi) !important;
        background: #141E2E !important;
    }
    details[data-testid="stExpander"] > div:last-child {
        background: rgba(12,17,30,0.6) !important;
        border: 1px solid var(--border) !important;
        border-top: none !important;
        border-radius: 0 0 4px 4px !important;
        padding: 0.5rem 0.6rem !important;
    }

    /* ── Scrollable param panel ────────────────────────────────── */
    div[data-testid="stVerticalBlockBorderWrapper"] {
        scrollbar-width: thin;
        scrollbar-color: var(--accent) #1A2035;
    }
    div[data-testid="stVerticalBlockBorderWrapper"]::-webkit-scrollbar       { width: 4px; }
    div[data-testid="stVerticalBlockBorderWrapper"]::-webkit-scrollbar-track { background: #1A2035; border-radius: 2px; }
    div[data-testid="stVerticalBlockBorderWrapper"]::-webkit-scrollbar-thumb { background: var(--accent); border-radius: 2px; }
    div[data-testid="stVerticalBlockBorderWrapper"]::-webkit-scrollbar-thumb:hover { background: #3ab5ad; }

    /* ── Dataframe / table ─────────────────────────────────────── */
    div[data-testid="stDataFrame"] {
        border: 1px solid var(--border) !important;
        border-radius: var(--radius) !important;
    }

    /* ── Plotly chart containers ───────────────────────────────── */
    div[data-testid="stPlotlyChart"] {
        border: 1px solid var(--border);
        border-radius: var(--radius);
        overflow: hidden;
    }

    /* ── Horizontal rule ───────────────────────────────────────── */
    hr { border-color: var(--border) !important; margin: 0.6rem 0 !important; }

    /* ── Checkbox / radio ──────────────────────────────────────── */
    div[data-testid="stCheckbox"]  label span,
    div[data-testid="stRadio"]     label span {
        font-family: 'IBM Plex Mono', monospace !important;
        font-size: 0.78rem !important;
        color: var(--text-main) !important;
    }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────
# Theme helpers
# ─────────────────────────────────────────────────────────────

def _plt():
    """Return (paper_bg, plot_bg, font_col, grid_col, legend_bg) for the
    current light/dark theme.  Call once at render time; reads session_state."""
    if st.session_state.get("_light_mode", False):
        return "#F0F4F8", "#FFFFFF", "#1A202C", "#E2E8F0", "#F7FAFC"
    return "#0A0E1A", "#111827", "#E8ECF0", "#2D3748", "#111827"


# ─────────────────────────────────────────────────────────────
# Constants
# ─────────────────────────────────────────────────────────────

# Eskom Megaflex
# SA Megaflex: High season = June, July, August ONLY (May is low season!)
WINTER_MONTHS = {6, 7, 8}

# ─────────────────────────────────────────────────────────────
# Eskom 2025/26 Tariff Database
# Tuple: (w_peak, w_standard, w_off_peak, s_peak, s_standard, s_off_peak) ZAR/kWh incl VAT
# = active energy charge + legacy charge (both incl VAT, from Tariff Booklet 2025/26)
# TOU applies to Megaflex & Miniflex. Nightsave = flat rate (all periods same price).
# ─────────────────────────────────────────────────────────────
TARIFF_DB = {
    # ── Megaflex Non-local Authority (p13) ────────────────────────────────────
    "Megaflex ≤300km <500V":        (8.1348, 2.2302, 1.5740, 3.5294, 2.0990, 1.5740),
    "Megaflex ≤300km 500V-66kV":    (7.9249, 2.1727, 1.5335, 3.4383, 2.0449, 1.5335),
    "Megaflex ≤300km 66-132kV":     (7.3544, 2.0162, 1.4231, 3.1908, 1.8977, 1.4231),
    "Megaflex ≤300km >132kV":       (6.8579, 1.8801, 1.3271, 2.9754, 1.7696, 1.3271),
    "Megaflex 300-600km <500V":     (8.2134, 2.2499, 1.5871, 3.5620, 2.1174, 1.5871),
    "Megaflex 300-600km 500V-66kV": (8.0017, 2.1919, 1.5464, 3.4701, 2.0628, 1.5464),
    "Megaflex 300-600km 66-132kV":  (7.4256, 2.0341, 1.4350, 3.2203, 1.9143, 1.4350),
    "Megaflex 300-600km >132kV":    (6.9243, 1.8968, 1.3381, 3.0029, 1.7850, 1.3381),
    "Megaflex 600-900km <500V":     (8.2922, 2.2696, 1.6003, 3.5947, 2.1357, 1.6003),
    "Megaflex 600-900km 500V-66kV": (8.0784, 2.2111, 1.5591, 3.5020, 2.0807, 1.5591),
    "Megaflex 600-900km 66-132kV":  (7.4967, 2.0518, 1.4468, 3.2499, 1.9310, 1.4468),
    "Megaflex 600-900km >132kV":    (6.9906, 1.9134, 1.3492, 3.0305, 1.8005, 1.3492),
    "Megaflex >900km <500V":        (8.3710, 2.2892, 1.6135, 3.6274, 2.1541, 1.6135),
    "Megaflex >900km 500V-66kV":    (8.1551, 2.2303, 1.5719, 3.5338, 2.0985, 1.5719),
    "Megaflex >900km 66-132kV":     (7.5679, 2.0697, 1.4588, 3.2793, 1.9475, 1.4588),
    "Megaflex >900km >132kV":       (7.0570, 1.9299, 1.3601, 3.0580, 1.8160, 1.3601),
    # ── Miniflex Non-local Authority (p20) — same energy rates as Megaflex ───
    "Miniflex ≤300km <500V":        (8.1348, 2.2302, 1.5740, 3.5294, 2.0990, 1.5740),
    "Miniflex ≤300km 500V-66kV":    (7.9249, 2.1727, 1.5335, 3.4383, 2.0449, 1.5335),
    "Miniflex ≤300km 66-132kV":     (7.3544, 2.0162, 1.4231, 3.1908, 1.8977, 1.4231),
    "Miniflex ≤300km >132kV":       (6.8579, 1.8801, 1.3271, 2.9754, 1.7696, 1.3271),
    "Miniflex 300-600km <500V":     (8.2134, 2.2499, 1.5871, 3.5620, 2.1174, 1.5871),
    "Miniflex 300-600km 500V-66kV": (8.0017, 2.1919, 1.5464, 3.4701, 2.0628, 1.5464),
    "Miniflex 300-600km 66-132kV":  (7.4256, 2.0341, 1.4350, 3.2203, 1.9143, 1.4350),
    "Miniflex 300-600km >132kV":    (6.9243, 1.8968, 1.3381, 3.0029, 1.7850, 1.3381),
    "Miniflex 600-900km <500V":     (8.2922, 2.2696, 1.6003, 3.5947, 2.1357, 1.6003),
    "Miniflex 600-900km 500V-66kV": (8.0784, 2.2111, 1.5591, 3.5020, 2.0807, 1.5591),
    "Miniflex 600-900km 66-132kV":  (7.4967, 2.0518, 1.4468, 3.2499, 1.9310, 1.4468),
    "Miniflex 600-900km >132kV":    (6.9906, 1.9134, 1.3492, 3.0305, 1.8005, 1.3492),
    "Miniflex >900km <500V":        (8.3710, 2.2892, 1.6135, 3.6274, 2.1541, 1.6135),
    "Miniflex >900km 500V-66kV":    (8.1551, 2.2303, 1.5719, 3.5338, 2.0985, 1.5719),
    "Miniflex >900km 66-132kV":     (7.5679, 2.0697, 1.4588, 3.2793, 1.9475, 1.4588),
    "Miniflex >900km >132kV":       (7.0570, 1.9299, 1.3601, 3.0580, 1.8160, 1.3601),
    # ── MunicFlex Local Authority (p15) ——————————————————————————————————————————
    # Effective 1 July 2025. Replaces Megaflex/Miniflex/Nightsave for local authorities.
    # Values = (active energy incl VAT + legacy charge incl VAT) / 100 → ZAR/kWh
    # Tuple: (w_peak, w_std, w_off, s_peak, s_std, s_off)
    "MunicFlex ≤300km <500V":        (8.2878, 2.2719, 1.6040, 3.5963, 2.1386, 1.6046),
    "MunicFlex ≤300km 500V-66kV":    (8.0493, 2.2067, 1.5577, 3.4922, 2.0769, 1.5577),
    "MunicFlex ≤300km 66-132kV":     (7.4693, 2.0478, 1.4453, 3.2407, 1.9275, 1.4453),
    "MunicFlex ≤300km >132kV":       (6.9649, 1.9096, 1.3478, 3.0219, 1.7973, 1.3478),
    "MunicFlex 300-600km <500V":     (8.3980, 2.3006, 1.6232, 3.6422, 2.1653, 1.6232),
    "MunicFlex 300-600km 500V-66kV": (8.1390, 2.2298, 1.5732, 3.5297, 2.0984, 1.5732),
    "MunicFlex 300-600km 66-132kV":  (7.5416, 2.0660, 1.4575, 3.2706, 1.9443, 1.4575),
    "MunicFlex 300-600km >132kV":    (7.0324, 1.9264, 1.3590, 3.0497, 1.8130, 1.3590),
    "MunicFlex 600-900km <500V":     (8.4743, 2.3193, 1.6358, 3.6741, 2.1828, 1.6360),
    "MunicFlex 600-900km 500V-66kV": (8.2200, 2.2500, 1.5868, 3.5635, 2.1172, 1.5868),
    "MunicFlex 600-900km 66-132kV":  (7.6138, 2.0840, 1.4696, 3.3007, 1.9611, 1.4696),
    "MunicFlex 600-900km >132kV":    (7.0998, 1.9433, 1.3703, 3.0778, 1.8288, 1.3703),
    "MunicFlex >900km <500V":        (8.5663, 2.3427, 1.6512, 3.7118, 2.2045, 1.6512),
    "MunicFlex >900km 500V-66kV":    (8.2955, 2.2686, 1.5992, 3.5946, 2.1348, 1.5992),
    "MunicFlex >900km 66-132kV":     (7.6861, 2.1021, 1.4816, 3.3305, 1.9780, 1.4816),
    "MunicFlex >900km >132kV":       (7.1672, 1.9602, 1.3814, 3.1057, 1.8443, 1.3814),
    # ── Nightsave Urban Non-local Authority (p22) — flat energy, no TOU arbitrage
    # All 6 period prices identical per season; BESS peak-shaving via demand charge not modelled
    "Nightsave ≤300km <500V":       (2.1991, 2.1991, 2.1991, 2.1227, 2.1227, 2.1227),
    "Nightsave ≤300km 500V-66kV":   (2.1423, 2.1423, 2.1423, 2.0679, 2.0679, 2.0679),
    "Nightsave ≤300km 66-132kV":    (1.9881, 1.9881, 1.9881, 1.9191, 1.9191, 1.9191),
    "Nightsave ≤300km >132kV":      (1.8540, 1.8540, 1.8540, 1.7895, 1.7895, 1.7895),
    # ── eThekwini Metropolitan Municipality 2025/26 TOU ──────────────────────
    # Source: eThekwini Energy Management Directorate Tariff Booklet 25/26 p16,23
    # Prices in c/kWh (excl. VAT) → ×1.15/100 → ZAR/kWh incl. 15% VAT
    # Tuple: (w_peak, w_std, w_off, s_peak, s_std, s_off)
    # High demand season = Jun–Aug (WINTER_MONTHS); Low demand = Sep–May
    # CTOU: NMD ≤ 100 kVA | ITOU: NMD > 100 kVA
    "eThekwini CTOU (≤100kVA)": (7.3847, 3.6950, 1.8000, 3.6435, 2.9310, 1.7050),
    "eThekwini ITOU (>100kVA)": (7.0941, 1.9881, 1.4207, 3.1114, 1.8747, 1.4207),
    # ── Ekurhuleni Metropolitan Municipality — Tariff C (Large Power Users) ──────
    # Source: Schedule 2, CoE Electricity Tariffs 2025/26, p13 (energy charges only)
    # VAT EXCLUDED in source → values below are ×1.15 incl 15% VAT
    # Tariff C = TOU large power user; also has Basic (R3 701/mth) + Demand (R299/kVA/mth)
    # Only energy component is modelled here (marginal rate for solar/BESS dispatch)
    # PV export credit: Winter R1.6162/kWh; Summer R1.1053/kWh (net-consumer only)
    # Tuple: (w_peak, w_std, w_off, s_peak, s_std, s_off) — VAT incl ZAR/kWh
    "Ekurhuleni C LV 230/400V":        (5.7712, 4.5493, 4.4135, 3.2299, 2.9339, 2.8252),
    "Ekurhuleni C LV-Sub (direct)":    (5.6586, 4.4366, 4.3009, 3.1488, 2.8529, 2.7442),
    "Ekurhuleni C MV ≤11kV":           (5.5704, 4.3484, 4.2127, 3.0915, 2.7956, 2.6870),
    # ── City of Cape Town (CoCT) 2025/26 Business TOU ───────────────────────
    # Source: CoCT Tariff Book 2025/26 — capetown.gov.za/en/budget
    # TOU hours (weekday): Morning peak 07:00-09:00; Evening peak 17:00-19:00
    # High demand season: Jun-Aug; Low: Sep-May. Values = energy charge incl 15% VAT.
    # Excl. basic charge (~R3 500/mth) and demand charge (~R180-250/kVA/mth).
    # Tuple: (w_peak, w_std, w_off, s_peak, s_std, s_off)
    "Cape Town BTOU >100kVA (LV)":    (8.4900, 3.0500, 1.9400, 3.3900, 2.3200, 1.6700),
    "Cape Town BTOU ≤100kVA":         (9.0800, 3.2800, 2.0900, 3.6500, 2.4900, 1.7900),
    # ── CityPower (Johannesburg) 2025/26 TOU ────────────────────────────────
    # Source: CityPower Tariff Schedule 2025/26 — citypower.co.za
    # TOU hours (weekday): Morning peak 07:00-10:00; Evening peak 18:00-20:00
    # Note: CityPower retains OLD Eskom 3h morning / 2h evening peak schedule.
    # High demand season: Jun-Aug (aligned with Eskom). Values incl 15% VAT.
    "CityPower Jhb LPU (>500kVA)":    (6.8500, 2.5200, 1.5100, 2.9900, 1.9000, 1.3300),
    "CityPower Jhb MPU (100-500kVA)": (7.3700, 2.7100, 1.6200, 3.2200, 2.0400, 1.4300),
    # ── City of Tshwane (Pretoria) 2025/26 ──────────────────────────────────
    # Source: City of Tshwane Tariff Schedule 2025/26 — tshwane.gov.za
    # Commercial Schedule C (large power users). Energy charge incl 15% VAT.
    # TOU hours: Morning peak 07:00-10:00; Evening peak 18:00-20:00 (old-style).
    "Tshwane Business TOU (LV)":      (8.9400, 3.2300, 2.1200, 3.8400, 2.4500, 1.8600),
    # ── Nelson Mandela Bay (Gqeberha/Port Elizabeth) 2025/26 ────────────────
    # Source: NMB Metro Tariff Booklet 2025/26 — nelsonmandelabay.gov.za
    # Business TOU (large commercial, >100kVA). Energy charge incl 15% VAT.
    # TOU hours: Morning peak 07:00-09:00; Evening peak 17:00-20:00 weekdays.
    "NMB Business TOU >100kVA":       (8.8200, 3.1000, 2.0100, 3.6500, 2.3500, 1.7500),
    # ── Dis-Trans (Distribution-Transmission) TOU ────────────────────────────
    # Source: user-provided tariff table (R/MWh ÷ 1000 = R/kWh)
    # High Season (winter) / Low Season (summer); morning & evening peak same rate.
    # TOU hours follow standard Eskom Megaflex windows.
    "Dis_Trans":                      (5.7194, 1.4299, 0.9532, 2.3735, 1.3345, 0.9532),
    # ── PPA (Power Purchase Agreement) — flat rate, user-defined ─────────────
    # UI PPA
    # All periods same price; user sets a single PPA rate in the UI
    "PPA Custom (flat rate)":          None,   # sentinel: UI shows PPA rate input
    # ── Custom — manual input, no auto-populate ──────────────────────────────
    "Custom (manual)":               None,
}

# ─────────────────────────────────────────────────────────────
# Huawei LUNA2000-2236-1S SOH (1C, DOD100%, ≤40°C)
# : Utility Smart String Grid Forming ESS 2.0 Performance Guide p10
# key = / annual equivalent full cycles
# value = [SOH year0, year1, ..., yearN]  (1.0 = 100%)
# EoL threshold = 60% (Huawei official spec)
# ─────────────────────────────────────────────────────────────
BESS_EOL_SOH = 0.60 # / EoL threshold

# ── Legacy table: LUNA2000-2236-1S (1C reference, kept for backward compat) ──
HUAWEI_SOH_TABLE: dict[int, list[float]] = {
    292: [1.0000,0.9552,0.9323,0.9113,0.8914,0.8724,0.8538,0.8355,0.8175,
          0.7999,0.7826,0.7657,0.7490,0.7326,0.7166,0.7008,0.6853,0.6701,
          0.6552,0.6405,0.6262],
    365: [1.0000,0.9509,0.9243,0.8996,0.8762,0.8536,0.8315,0.8098,0.7887,
          0.7681,0.7479,0.7281,0.7089,0.6900,0.6716,0.6536,0.6361,0.6189,
          0.6021],
    475: [1.0000,0.9446,0.9122,0.8822,0.8535,0.8255,0.7985,0.7723,0.7468,
          0.7221,0.6980,0.6748,0.6521,0.6302,0.6089],
    548: [1.0000,0.9403,0.9043,0.8706,0.8384,0.8072,0.7771,0.7479,0.7198,
          0.6927,0.6665,0.6411,0.6166],
    657: [1.0000,0.9340,0.8923,0.8533,0.8160,0.7801,0.7457,0.7127,0.6810,
          0.6507,0.6215],
    730: [1.0000,0.9298,0.8845,0.8419,0.8013,0.7625,0.7254,0.6900,0.6563,
          0.6240],
}

# ── LUNA2000-5015-2S SOH ────────────────────────────────────────
# : Utility Smart String ESS Performance Guide LUNA2000-5015-2S 20250630
# pp. 16 (0.25C), 17 (0.33C), 18 (0.5C)
# : DOD 100%, SOC ≤50%, ≤40°C
# key = C (0.25 / 0.33 / 0.50)
# key = (annual cycles)
# = [SOH year0, year1, ..., yearN] (1.0 = 100%)
# SOH EoL ≥60%
# ─────────────────────────────────────────────────────────────────────────────
LUNA2000_5015_SOH: dict[float, dict[int, list[float]]] = {
    # ── 0.25C (4h discharge interval) — page 16 ──────────────────────────────
    0.25: {
        292: [1.0000,0.9776,0.9596,0.9437,0.9291,0.9153,0.9020,0.8892,0.8765,
              0.8641,0.8518,0.8396,0.8276,0.8158,0.8042,0.7926,0.7812,0.7699,
              0.7588,0.7479,0.7370],
        365: [1.0000,0.9746,0.9543,0.9362,0.9193,0.9032,0.8877,0.8726,0.8576,
              0.8429,0.8284,0.8141,0.8001,0.7862,0.7726,0.7592,0.7460,0.7330,
              0.7202,0.7077,0.6952],
        475: [1.0000,0.9703,0.9464,0.9248,0.9046,0.8851,0.8661,0.8475,0.8291,
              0.8113,0.7937,0.7766,0.7596,0.7431,0.7269,0.7110,0.6954,0.6802,
              0.6653,0.6506,0.6362],
        548: [1.0000,0.9675,0.9412,0.9173,0.8948,0.8730,0.8517,0.8308,0.8104,
              0.7905,0.7711,0.7521,0.7335,0.7153,0.6976,0.6803,0.6633,0.6468,
              0.6306,0.6148],
        657: [1.0000,0.9634,0.9335,0.9060,0.8800,0.8547,0.8301,0.8061,0.7828,
              0.7602,0.7380,0.7166,0.6956,0.6754,0.6556,0.6363,0.6176],
        730: [1.0000,0.9605,0.9283,0.8985,0.8701,0.8426,0.8159,0.7899,0.7647,
              0.7404,0.7167,0.6937,0.6714,0.6498,0.6288,0.6085],
    },
    # ── 0.33C (3h discharge interval) — page 17 ──────────────────────────────
    0.33: {
        292: [1.0000,0.9763,0.9575,0.9408,0.9254,0.9108,0.8967,0.8831,0.8696,
              0.8563,0.8432,0.8303,0.8176,0.8051,0.7927,0.7805,0.7685,0.7567,
              0.7451,0.7336,0.7222],
        365: [1.0000,0.9730,0.9517,0.9326,0.9146,0.8976,0.8810,0.8648,0.8488,
              0.8331,0.8177,0.8025,0.7877,0.7730,0.7586,0.7446,0.7307,0.7170,
              0.7037,0.6904,0.6775],
        475: [1.0000,0.9685,0.9432,0.9202,0.8985,0.8778,0.8574,0.8374,0.8179,
              0.7989,0.7802,0.7620,0.7442,0.7267,0.7096,0.6930,0.6767,0.6607,
              0.6451,0.6298,0.6149],
        548: [1.0000,0.9654,0.9375,0.9119,0.8878,0.8644,0.8416,0.8193,0.7976,
              0.7765,0.7559,0.7357,0.7161,0.6970,0.6784,0.6603,0.6426,0.6253,
              0.6085],
        657: [1.0000,0.9608,0.9290,0.8996,0.8716,0.8445,0.8181,0.7926,0.7678,
              0.7437,0.7205,0.6978,0.6759,0.6545,0.6339,0.6138],
        730: [1.0000,0.9579,0.9233,0.8914,0.8608,0.8314,0.8028,0.7751,0.7484,
              0.7226,0.6977,0.6734,0.6501,0.6275,0.6057],
    },
    # ── 0.50C (2h discharge interval) — page 18 ──────────────────────────────
    0.50: {
        292: [1.0000,0.9750,0.9543,0.9357,0.9184,0.9020,0.8861,0.8705,0.8552,
              0.8403,0.8255,0.8110,0.7966,0.7827,0.7689,0.7554,0.7421,0.7290,
              0.7161,0.7035,0.6912],
        365: [1.0000,0.9714,0.9476,0.9261,0.9058,0.8864,0.8675,0.8490,0.8309,
              0.8131,0.7958,0.7787,0.7621,0.7458,0.7298,0.7142,0.6989,0.6839,
              0.6692,0.6549,0.6408],
        475: [1.0000,0.9660,0.9376,0.9114,0.8868,0.8629,0.8395,0.8169,0.7948,
              0.7734,0.7525,0.7321,0.7123,0.6929,0.6741,0.6559,0.6381,0.6207,
              0.6038],
        548: [1.0000,0.9625,0.9309,0.9017,0.8741,0.8473,0.8211,0.7958,0.7714,
              0.7476,0.7246,0.7022,0.6806,0.6595,0.6391,0.6194,0.6003],
        657: [1.0000,0.9571,0.9208,0.8871,0.8549,0.8239,0.7940,0.7650,0.7372,
              0.7103,0.6845,0.6595,0.6353,0.6121],
        730: [1.0000,0.9537,0.9141,0.8775,0.8424,0.8087,0.7763,0.7451,0.7151,
              0.6864,0.6589,0.6323,0.6068],
    },
}


def get_soh_by_year(annual_cycles: float, c_rate: float | None = None) -> list[float]:
    """
     year 0~25 SOH
    EoL SOH 0
    c_rate: 0.25 / 0.33 / 0.50 → LUNA2000-5015-2S C ;
            None
    Interpolate official SOH table for given annual_cycles → list[26] (year 0-25).
    """
    # Choose lookup table
    if c_rate is not None and c_rate in LUNA2000_5015_SOH:
        tbl = LUNA2000_5015_SOH[c_rate]
    else:
        tbl = HUAWEI_SOH_TABLE

    keys = sorted(tbl.keys())
    ac = float(max(min(annual_cycles, keys[-1]), keys[0]))

    if ac <= keys[0]:
        base = list(tbl[keys[0]])
    elif ac >= keys[-1]:
        base = list(tbl[keys[-1]])
    else:
        lo = max(k for k in keys if k <= ac)
        hi = min(k for k in keys if k >= ac)
        if lo == hi:
            base = list(tbl[lo])
        else:
            frac = (ac - lo) / (hi - lo)
            lo_a = tbl[lo]
            hi_a = tbl[hi]
            n = min(len(lo_a), len(hi_a))
            base = [lo_a[i] * (1 - frac) + hi_a[i] * frac for i in range(n)]

    # Pad to 21 entries (year 0-20); years beyond table = 0.0 (past EoL)
    result: list[float] = list(base)
    while len(result) < 21:
        result.append(0.0)

    # Zero out any year after EoL (SOH drops below threshold)
    eol_hit = False
    for i in range(1, len(result)):
        if eol_hit or result[i] < BESS_EOL_SOH:
            eol_hit = True
            result[i] = 0.0
    return result


def compute_bess_eol(throughput_yr1: float, bess_kwh: float,
                     cycles: int, dod: float) -> tuple[float, float]:
    """
     + EoL SOH EoL = SOH < 60%
    Returns (eol_years, annual_cycles)
    """
    usable = bess_kwh * dod / 100.0
    if usable <= 0:
        return 25.0, 0.0
    annual_cycles = throughput_yr1 / (2.0 * usable)
    if annual_cycles <= 0:
        return 25.0, 0.0

    soh_arr = get_soh_by_year(annual_cycles)
    eol = 25.0
    for yr in range(1, len(soh_arr)):
        if soh_arr[yr] < BESS_EOL_SOH:
            prev = soh_arr[yr - 1]
            curr = soh_arr[yr]
            frac = (prev - BESS_EOL_SOH) / max(prev - curr, 1e-9)
            eol = round(yr - 1 + frac, 1)
            break
    return eol, round(annual_cycles, 2)

# Section 12B: SA accelerated depreciation — 50%/30%/20% over 3yr (PV >1MW)
# Year 1: 50%  Year 2: 30%  Year 3: 20%
SECTION_12B = {1: 0.50, 2: 0.30, 3: 0.20}

# Pure BESS depreciation (non-12B): straight-line 20% × 5 years
SECTION_BESS_ONLY = {1: 0.20, 2: 0.20, 3: 0.20, 4: 0.20, 5: 0.20}

ANALYSIS_YEARS = 20
FOREX_FALLBACK = 18.5 # USD/ZAR / fallback rate

# C-rate options
C_RATE_OPTIONS = {"0.25C (4h)": 0.25, "0.33C (3h)": 0.33, "0.5C (2h)": 0.50, "1C (1h)": 1.00}

# ─────────────────────────────────────────────────────────────
# UI string constants
# ─────────────────────────────────────────────────────────────


# Column display map: internal fin_df names → display names per language
_COL_ENGLISH = {
    "Year":              "Year",
    "Total Saving (ZAR)":      "Total Saving (ZAR)",
    "PV Saving (ZAR)":                 "PV Saving (ZAR)",
    "BESS Saving (ZAR)":               "BESS Saving (ZAR)",
    "Depreciation (ZAR)":             "Depreciation (ZAR)",
    "Assess. Loss C/F (ZAR)":     "Assess. Loss C/F (ZAR)",
    "Net Profit (ZAR)":           "Net Profit (ZAR)",
    "Net Cash Flow NCF (ZAR)":     "Net Cash Flow NCF (ZAR)",
    "Discounted CF PV (ZAR)":        "Discounted CF PV (ZAR)",
    "Cumulative CF (ZAR)":           "Cumulative CF (ZAR)",
    "SOH%":                   "SOH%",
    "Status":                   "Status",
}







# Default parameters
DEFAULT_PARAMS = {
    "lat": -26.1,
    "lon": 28.0,
    "pv_kwp": 4000.0,              # 4 MWp
    "bess_kwh": 10000.0,           # 10 MWh
    "c_rate_label": "0.25C (4h)",
    "load_peak_kw": 2500.0, # kW
    "load_std_kw": 2500.0, # kW
    "load_offpeak_kw": 2500.0, # kW
    "pv_loss": 14.0,
    "tilt": 26.0,                  # abs(lat) for SA
    "azimuth": 180.0, # = 180° in PVGIS convention
    "rte": 88.2,
    "bess_cycles": 10000,
    "dod": 100.0,
    "forex_usd_zar": FOREX_FALLBACK,
    "pv_usd_per_w": 0.75,
    "bess_usd_per_wh": 0.20,
    "pv_opex_per_kwp": 125.0,
    "bess_opex_per_kwh": 10.0,
    "tariff_escalation": 8.76,   # NERSA FY2026/27 approved average increase
    "discount_rate": 10.0,
    "tax_rate": 27.0,
    "pv_degradation": 0.5,
    # Project commissioning lead times (months from PO)
    "bess_lead_months": 6,
    "pv_lead_months": 12,
    # Tariff mode (see TARIFF_DB)
    "tariff_mode": "Megaflex ≤300km <500V",
    # Jun-Aug: Winter HIGH season (Eskom Megaflex 2025/26 ≤300km <500V)
    "w_morning_peak": 8.1348,
    "w_evening_peak": 8.1348,
    "w_standard":     2.2302,
    "w_off_peak":     1.5740,
    # Sep-May: Summer LOW season
    "s_morning_peak": 3.5294,
    "s_evening_peak": 3.5294,
    "s_standard":     2.0990,
    "s_off_peak":     1.5740,
}

# ─────────────────────────────────────────────────────────────
# Auth Gate (runs before any UI is rendered)
# ─────────────────────────────────────────────────────────────
if not is_logged_in():
    render_auth_gate()
    st.stop()

# Flush any deferred localStorage op (token save / clear) that was queued
# during login or logout.  Must run here — in the first *stable* logged-in
# render — so the component iframe has time to execute before any st.rerun().
flush_token_to_storage()

# ── Light / dark theme CSS injection ─────────────────────────────────────────
if st.session_state.get("_light_mode", False):
    st.markdown("""<style>
/* ═══════════════════════════════════════════════════════════════
   BTM LIGHT MODE — comprehensive overrides
   ═══════════════════════════════════════════════════════════════ */

/* ── 1. Root surfaces ──────────────────────────────────────── */
html, body,
.stApp,
[data-testid="stAppViewContainer"],
[data-testid="stMain"],
[data-testid="stMainBlockContainer"],
[data-testid="stVerticalBlockBorderWrapper"],
[data-testid="stVerticalBlock"],
section[data-testid="stSidebar"] {
    background-color: #F0F4F8 !important;
    color:            #1A202C !important;
}

/* ── 2. All text nodes ─────────────────────────────────────── */
p, span, li, h1, h2, h3, h4, h5, h6, label, div,
[data-testid="stMarkdownContainer"] *,
[data-testid="stCaptionContainer"] *,
[data-testid="stText"] * {
    color: #1A202C !important;
}
/* Muted / secondary text */
small, em, .stCaption,
[data-testid="stCaptionContainer"] p {
    color: #4A5568 !important;
}

/* ── 3. Metric tiles ───────────────────────────────────────── */
[data-testid="stMetric"] {
    background:    #FFFFFF !important;
    border:        1px solid #CBD5E0 !important;
    border-radius: 8px !important;
    padding:       10px 14px !important;
    box-shadow:    0 1px 4px rgba(0,0,0,0.06) !important;
}
[data-testid="stMetric"] label,
[data-testid="stMetric"] [data-testid="stMetricLabel"] *,
[data-testid="stMetricLabel"] * { color: #4A5568 !important; }
[data-testid="stMetric"] [data-testid="stMetricValue"] *,
[data-testid="stMetricValue"] * { color: #1A202C !important; font-weight: 700 !important; }
[data-testid="stMetric"] [data-testid="stMetricDelta"] *,
[data-testid="stMetricDelta"] * { color: #059669 !important; }

/* ── 4. Number inputs ──────────────────────────────────────── */
[data-testid="stNumberInputContainer"],
[data-testid="stNumberInput"] > div {
    background:    #FFFFFF !important;
    border:        1px solid #CBD5E0 !important;
    border-radius: 6px !important;
}
[data-testid="stNumberInputContainer"] input,
[data-testid="stNumberInput"] input {
    background: #FFFFFF !important;
    color:      #1A202C !important;
}
[data-testid="stNumberInputContainer"] button,
[data-testid="stNumberInput"] button {
    background:   #F0F4F8 !important;
    border-color: #CBD5E0 !important;
    color:        #374151 !important;
}

/* ── 5. Text inputs & text areas ───────────────────────────── */
[data-testid="stTextInputRootElement"],
[data-testid="stTextInput"] > div {
    background:    #FFFFFF !important;
    border:        1px solid #CBD5E0 !important;
    border-radius: 6px !important;
}
[data-testid="stTextInput"] input,
[data-testid="stTextArea"] textarea {
    background: #FFFFFF !important;
    color:      #1A202C !important;
}

/* ── 6. Selectbox ──────────────────────────────────────────── */
[data-testid="stSelectbox"] > div > div {
    background:    #FFFFFF !important;
    border:        1px solid #CBD5E0 !important;
    border-radius: 6px !important;
    color:         #1A202C !important;
}
[data-testid="stSelectboxVirtualDropdown"],
[data-testid="stSelectboxDropdown"] {
    background: #FFFFFF !important;
    border:     1px solid #CBD5E0 !important;
    color:      #1A202C !important;
}
[data-testid="stSelectboxDropdown"] li,
[data-testid="stSelectboxVirtualDropdown"] li {
    color: #1A202C !important;
}
[data-testid="stSelectboxDropdown"] li:hover,
[data-testid="stSelectboxVirtualDropdown"] li:hover {
    background: #EDF2F7 !important;
}

/* ── 7. Sliders ────────────────────────────────────────────── */
[data-testid="stSlider"] [data-baseweb="slider"] [data-testid="stSliderTrackFill"] {
    background: #00C48C !important;
}
[data-testid="stSlider"] [role="slider"] {
    background:   #00C48C !important;
    border-color: #FFFFFF !important;
    box-shadow:   0 0 0 2px #00C48C !important;
}
[data-testid="stSlider"] > div > div > div > div > div {
    background: #CBD5E0 !important;
}

/* ── 8. Checkboxes & radio ─────────────────────────────────── */
[data-testid="stCheckbox"] label,
[data-testid="stRadio"] label {
    color: #1A202C !important;
}
[data-testid="stCheckbox"] [data-baseweb="checkbox"] > div {
    background:   #FFFFFF !important;
    border-color: #A0AEC0 !important;
}

/* ── 9. Buttons ────────────────────────────────────────────── */
/* Primary */
[data-testid="stBaseButton-primary"] {
    background: #00C48C !important;
    color:      #FFFFFF !important;
    border:     none !important;
}
[data-testid="stBaseButton-primary"]:hover {
    background: #00A870 !important;
}
/* Secondary */
[data-testid="stBaseButton-secondary"] {
    background:   #FFFFFF !important;
    color:        #374151 !important;
    border-color: #CBD5E0 !important;
}
[data-testid="stBaseButton-secondary"]:hover {
    background: #EDF2F7 !important;
    color:      #111827 !important;
}

/* ── 10. Expanders ─────────────────────────────────────────── */
[data-testid="stExpander"] {
    background:    #FFFFFF !important;
    border:        1px solid #CBD5E0 !important;
    border-radius: 8px !important;
}
[data-testid="stExpander"] summary,
[data-testid="stExpander"] details > summary {
    background: #FFFFFF !important;
    color:      #1A202C !important;
}
[data-testid="stExpander"] summary:hover {
    background: #F7FAFC !important;
}
[data-testid="stExpander"] summary svg { color: #718096 !important; }
[data-testid="stExpanderDetails"] {
    background: #FFFFFF !important;
    border-top:  1px solid #E2E8F0 !important;
}

/* ── 11. Tabs ──────────────────────────────────────────────── */
[data-testid="stTabs"] [role="tablist"] {
    border-bottom: 2px solid #E2E8F0 !important;
}
[data-testid="stTabs"] [role="tab"] {
    color:            #718096 !important;
    background:       transparent !important;
    border-bottom:    2px solid transparent !important;
}
[data-testid="stTabs"] [role="tab"][aria-selected="true"] {
    color:            #00C48C !important;
    border-bottom:    2px solid #00C48C !important;
    font-weight:      600 !important;
}
[data-testid="stTabContent"] {
    background: #F0F4F8 !important;
    padding-top: 8px !important;
}

/* ── 12. Columns & blocks ──────────────────────────────────── */
[data-testid="stHorizontalBlock"],
[data-testid="column"] {
    background: transparent !important;
}

/* ── 13. Info / warning / success / error boxes ────────────── */
[data-testid="stAlert"] {
    border-radius: 6px !important;
}
[data-testid="stAlert"][data-baseweb="notification"][kind="info"] {
    background: #EBF8FF !important; border-color: #90CDF4 !important;
}
[data-testid="stAlert"][data-baseweb="notification"][kind="success"] {
    background: #F0FFF4 !important; border-color: #9AE6B4 !important;
}
[data-testid="stAlert"][data-baseweb="notification"][kind="warning"] {
    background: #FFFBEB !important; border-color: #FAD075 !important;
}
[data-testid="stAlert"][data-baseweb="notification"][kind="error"] {
    background: #FFF5F5 !important; border-color: #FEB2B2 !important;
}

/* ── 14. Dataframes / tables ───────────────────────────────── */
[data-testid="stDataFrame"],
[data-testid="stTable"] {
    background: #FFFFFF !important;
}
[data-testid="stDataFrame"] th,
[data-testid="stTable"] th {
    background: #EDF2F7 !important;
    color:      #2D3748 !important;
}
[data-testid="stDataFrame"] td,
[data-testid="stTable"] td {
    color:           #1A202C !important;
    border-color:    #E2E8F0 !important;
}

/* ── 15. Plot containers (chart backgrounds) ───────────────── */
[data-testid="stPlotlyChart"] > div,
[data-testid="stVegaLiteChart"] > div,
[data-testid="stArrowVegaLiteChart"] > div {
    background: #FFFFFF !important;
    border-radius: 8px !important;
}

/* ── 16. Toast notifications ───────────────────────────────── */
[data-testid="stToast"] {
    background: #FFFFFF !important;
    color:      #1A202C !important;
    border:     1px solid #CBD5E0 !important;
    box-shadow: 0 4px 12px rgba(0,0,0,0.1) !important;
}

/* ── 17. Dividers ──────────────────────────────────────────── */
hr { border-color: #CBD5E0 !important; }

/* ── 18. Scrollbars ────────────────────────────────────────── */
::-webkit-scrollbar { width: 6px !important; height: 6px !important; }
::-webkit-scrollbar-track { background: #EDF2F7 !important; }
::-webkit-scrollbar-thumb { background: #A0AEC0 !important; border-radius: 3px !important; }
::-webkit-scrollbar-thumb:hover { background: #718096 !important; }

/* ── 19. Header bar area ───────────────────────────────────── */
header[data-testid="stHeader"] {
    background: #F0F4F8 !important;
    border-bottom: 1px solid #E2E8F0 !important;
}

/* ── 20. Popover override (handled by _card_css but reinforce) */
[data-testid="stPopoverBody"] {
    background: #FFFFFF !important;
    border:     1px solid #E2E8F0 !important;
    box-shadow: 0 4px 20px rgba(0,0,0,0.12) !important;
}

</style>""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────
# Scenario Gate
# Shows the landing page until the user picks a scenario.
# ─────────────────────────────────────────────────────────────
_scenario = st.session_state.get("_scenario")
if not _scenario:
    render_scenario_selector()
    st.stop()
elif _scenario == "wheeling":
    render_wheeling_placeholder()
    st.stop()
# _scenario == "btm"  →  fall through to the BTM app below

# ─────────────────────────────────────────────────────────────
# Session State Initialization
# ─────────────────────────────────────────────────────────────
def init_session_state():
    if "initialized" not in st.session_state:
        st.session_state.initialized = True
        for k, v in DEFAULT_PARAMS.items():
            st.session_state[k] = v
        st.session_state.stashed_params = None
        st.session_state.results = None
        st.session_state.hourly_df = None
        st.session_state.fin_df = None
        st.session_state.pvgis_status = "Pending"
        st.session_state.annual_pv_kwh = None
        st.session_state.pvgis_key = "" # / detect location change
        # Load profile CSV upload state (not persisted in snapshots)
        st.session_state["load_profile_8760"] = None
        st.session_state["load_profile_name"] = ""

init_session_state()

# ─────────────────────────────────────────────────────────────
# Utility Functions
# ─────────────────────────────────────────────────────────────

def fetch_forex_rate() -> float:
    """ USD/ZAR / Live USD/ZAR rate with 3-tier fallback"""
    apis = [
        ("https://open.er-api.com/v6/latest/USD",         lambda r: r.json()["rates"]["ZAR"]),
        ("https://api.exchangerate-api.com/v4/latest/USD", lambda r: r.json()["rates"]["ZAR"]),
        ("https://api.frankfurter.app/latest?from=USD&to=ZAR",
         lambda r: r.json()["rates"]["ZAR"]),
    ]
    for url, extractor in apis:
        try:
            r = requests.get(url, timeout=6)
            r.raise_for_status()
            return float(extractor(r))
        except Exception:
            continue
    return FOREX_FALLBACK


@st.cache_data(ttl=3600)
def _fetch_forex_cached() -> float:
    """ 1 / Auto-fetch on startup, cached 1h"""
    return fetch_forex_rate()


# Auto-fetch live forex rate on startup — functions are now defined, safe to call
if "forex_auto_fetched" not in st.session_state:
    st.session_state.forex_usd_zar = _fetch_forex_cached()
    st.session_state.forex_auto_fetched = True


def get_capex_zar() -> tuple[float, float]:
    """USDZAR / Compute ZAR unit costs from USD and forex"""
    rate = st.session_state.forex_usd_zar
    pv_per_kwp   = st.session_state.pv_usd_per_w   * 1000.0 * rate  # $/W→ZAR/kWp
    bess_per_kwh = st.session_state.bess_usd_per_wh * 1000.0 * rate  # $/Wh→ZAR/kWh
    return pv_per_kwp, bess_per_kwh


def get_tariff_for_hour(hour: int, month: int, day_type: str = "weekday") -> tuple[float, str]:
    """
    Unified TOU tariff lookup — routes to the correct schedule per tariff_mode.

    Eskom 2025/26 (Megaflex/Miniflex/MunicFlex/Nightsave):
      Weekday: Morning peak 07:00-09:00, Evening peak 17:00-20:00
      Saturday: off-peak all day
      Sunday/holidays: off-peak + standard 18:00-20:00
      High season: June-August

    CityPower / Tshwane / NMB — "old-style" TOU (pre-2025/26 update):
      Morning peak 07:00-10:00 (3h), Evening peak 18:00-20:00 (2h)

    Cape Town BTOU:
      Morning peak 07:00-09:00 (2h), Evening peak 17:00-19:00 (2h) — shorter evening

    day_type: "weekday" | "saturday" | "sunday"
    """
    is_winter  = month in WINTER_MONTHS
    pfx        = "w_" if is_winter else "s_"
    off_peak_p = st.session_state[f"{pfx}off_peak"]
    morning_p  = st.session_state[f"{pfx}morning_peak"]
    evening_p  = st.session_state[f"{pfx}evening_peak"]
    standard_p = st.session_state[f"{pfx}standard"]

    if day_type == "sunday":
        if 18 <= hour < 20:
            return standard_p, "standard"
        return off_peak_p, "off_peak"

    if day_type == "saturday":
        return off_peak_p, "off_peak"

    # ── weekday: choose peak window by tariff ──────────────────────────────
    _mode = st.session_state.get("tariff_mode", "")

    if _mode.startswith("Cape Town"):
        # CoCT BTOU: Morning 07-09 (2h), Evening 17-19 (2h shorter)
        if hour < 6 or hour >= 22:      return off_peak_p, "off_peak"
        elif 7 <= hour < 9:             return morning_p, "morning_peak"
        elif 17 <= hour < 19:           return evening_p, "evening_peak"
        else:                           return standard_p, "standard"

    if _mode.startswith("CityPower") or _mode.startswith("Tshwane") or _mode.startswith("NMB"):
        # Old-style Eskom: Morning 07-10 (3h), Evening 18-20 (2h)
        if hour < 6 or hour >= 22:      return off_peak_p, "off_peak"
        elif 7 <= hour < 10:            return morning_p, "morning_peak"
        elif 18 <= hour < 20:           return evening_p, "evening_peak"
        else:                           return standard_p, "standard"

    # ── Eskom 2025/26 default (Megaflex / Miniflex / MunicFlex / Nightsave) ──
    if hour < 6 or hour >= 22:          return off_peak_p, "off_peak"
    elif 7 <= hour < 9:                 return morning_p, "morning_peak"
    elif 17 <= hour < 20:               return evening_p, "evening_peak"
    else:                               return standard_p, "standard"


def get_ethekwini_tou(hour: int, month: int, day_type: str) -> tuple[float, str]:
    """
    eThekwini CTOU / ITOU 2025/26 TOU schedule.
    Source: eThekwini Tariff Booklet 25/26 p22 (time period tables).
    Returns (tariff_price_ZAR, period_string)
    period_string: "peak" | "standard" | "off_peak"

    High demand season (Jun–Aug):
      Weekday peaks: 06:00–08:00 and 17:00–20:00
      Saturday:      standard 07:00–12:00 and 17:00–19:00; rest off-peak
      Sunday:        standard 17:00–19:00; rest off-peak

    Low demand season (Sep–May):
      Weekday peaks: 07:00–09:00 and 18:00–21:00
      Saturday:      standard 07:00–12:00 and 18:00–20:00; rest off-peak
      Sunday:        standard 18:00–20:00; rest off-peak

    Note: eThekwini has NO peak on weekends/holidays — at most Standard.
    """
    is_high = month in WINTER_MONTHS
    pfx     = "w_" if is_high else "s_"
    peak_p  = st.session_state[f"{pfx}morning_peak"]   # one peak rate (morning_peak = evening_peak)
    std_p   = st.session_state[f"{pfx}standard"]
    off_p   = st.session_state[f"{pfx}off_peak"]

    if day_type == "sunday":
        if is_high:
            return (std_p, "standard") if 17 <= hour < 19 else (off_p, "off_peak")
        else:
            return (std_p, "standard") if 18 <= hour < 20 else (off_p, "off_peak")

    if day_type == "saturday":
        if is_high:
            if (7 <= hour < 12) or (17 <= hour < 19):
                return std_p, "standard"
        else:
            if (7 <= hour < 12) or (18 <= hour < 20):
                return std_p, "standard"
        return off_p, "off_peak"

    # weekday
    if is_high:                                   # HIGH demand Jun–Aug
        if (6 <= hour < 8) or (17 <= hour < 20):
            return peak_p, "peak"
        elif (8 <= hour < 17) or (20 <= hour < 22):
            return std_p, "standard"
    else:                                         # LOW demand Sep–May
        if (7 <= hour < 9) or (18 <= hour < 21):
            return peak_p, "peak"
        elif (6 <= hour < 7) or (9 <= hour < 18) or (21 <= hour < 22):
            return std_p, "standard"

    return off_p, "off_peak"                      # 22:00–06:00


def get_pvgis_data(lat, lon, kwp, loss, tilt, azimuth) -> dict:
    """
     PVGIS API / Call EU PVGIS API
     1650h / Auto fallback to 1650h empirical on failure
    """
    try:
        url = "https://re.jrc.ec.europa.eu/api/v5_2/PVcalc"
        params = {
            "lat": lat, "lon": lon,
            "peakpower": kwp, "loss": loss,
            "angle": tilt, "aspect": azimuth,
            "outputformat": "json",
            "pvtechchoice": "crystSi",
            "mountingplace": "free",
        }
        resp = requests.get(url, params=params, timeout=15)
        resp.raise_for_status()
        data = resp.json()
        monthly = data["outputs"]["monthly"]["fixed"]
        annual_kwh = sum(m["E_m"] for m in monthly)
        return {
            "status": ("API Success ✓"),
            "annual_kwh": annual_kwh,
            "winter_daily_kwh": monthly[6]["E_d"],   # July
            "summer_daily_kwh": monthly[0]["E_d"],   # January
            "monthly_kwh": [m["E_m"] for m in monthly],
        }
    except Exception as e:
        loss_f = 1 - loss / 100
        annual_kwh = 1650 * kwp * loss_f
        avg_d = annual_kwh / 365
        return {
            "status": (f"Offline / {str(e)[:35]}"),
            "annual_kwh": annual_kwh,
            "winter_daily_kwh": avg_d * 0.70,
            "summer_daily_kwh": avg_d * 1.30,
            "monthly_kwh": [annual_kwh / 12] * 12,
        }


def build_hourly_pv_profile(pvgis_data: dict, pv_kwp: float) -> np.ndarray:
    """
     → 876006:00-18:00
    Monthly generation → 8760h profile using Gaussian daytime weights
    """
    hours = np.arange(24)
    w = np.exp(-0.5 * ((hours - 12) / 2.5) ** 2)
    w[hours < 6] = 0.0
    w[hours > 18] = 0.0
    w /= w.sum()

    days_in_month = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    hourly_pv = np.zeros(8760)
    idx = 0
    for m_i, n_days in enumerate(days_in_month):
        daily = pvgis_data["monthly_kwh"][m_i] / n_days
        for _ in range(n_days):
            for h in range(24):
                hourly_pv[idx] = daily * w[h]
                idx += 1
    return hourly_pv


def parse_load_csv(uploaded_file) -> "tuple[np.ndarray | None, str]":
    """
    Parse an uploaded load-profile file (CSV / Excel) → 8760-float array (kW).

    Accepted formats
    ────────────────
    • 8760 rows  — 1-hour resolution  (passed through)
    • 17520 rows — 30-min resolution  (averaged to hourly pairs)
    • 8784 rows  — leap-year 1-h      (last 24 h trimmed)
    • 1 or 2+ columns; first purely-numeric column is used
    • Optional header row (auto-detected)
    • Delimiters: comma · semicolon · tab · pipe
    Returns (array, info_msg) on success; (None, error_msg) on failure.
    """
    import io as _io

    fname = uploaded_file.name.lower()
    try:
        if fname.endswith((".xlsx", ".xls")):
            df_raw = pd.read_excel(uploaded_file, header=None, dtype=str)
        else:
            raw_bytes = uploaded_file.read()
            raw = raw_bytes.decode("utf-8-sig", errors="replace")
            first_line = raw.split("\n")[0]
            sep = ","
            for _s in (";", "\t", "|"):
                if _s in first_line:
                    sep = _s
                    break
            df_raw = pd.read_csv(_io.StringIO(raw), header=None, sep=sep, dtype=str)

        # ── Detect numeric column ─────────────────────────────────────
        def _first_numeric_col(df):
            for col in df.columns:
                try:
                    vals = pd.to_numeric(df[col], errors="raise")
                    return vals.astype(float).values
                except Exception:
                    pass
            return None

        series = _first_numeric_col(df_raw)
        if series is None:
            # First row might be a header — drop it and retry
            series = _first_numeric_col(df_raw.iloc[1:].reset_index(drop=True))
        if series is None:
            return None, "❌ No numeric column found — ensure the file contains load values in kW"

        series = series[~np.isnan(series)]  # drop NaN
        n = len(series)

        resample_note = ""
        if n == 17520:
            series = series.reshape(-1, 2).mean(axis=1)
            n = 8760
            resample_note = " (resampled from 30-min → 1-h)"
        elif n == 8784:
            series = series[:8760]
            n = 8760
            resample_note = " (leap-year trimmed to 8760 h)"
        elif n != 8760:
            return None, (f"❌ Got {n} rows — need exactly 8760 (1-h) or 17520 (30-min). "
                          "Check the file has one value per interval and no extra blank rows.")

        if series.min() < 0:
            return None, "❌ Negative values found — file should contain positive kW load values"
        if series.max() > 500_000:
            return None, "❌ Max value > 500,000 kW — check units (file must be in kW, not W)"

        annual_mwh = series.sum() / 1000
        msg = (f"✅ {n} hourly values{resample_note} · "
               f"Min {series.min():.0f} kW · Max {series.max():.0f} kW · "
               f"Avg {series.mean():.0f} kW · Annual {annual_mwh:.1f} MWh")
        return series.astype(np.float64), msg

    except Exception as exc:
        return None, f"❌ Parse error: {exc}"


def run_8760_dispatch(
    pv_kwp: float,
    bess_kwh: float,
    bess_kw: float,
    load_peak_kw: float,       # kW during peak tariff hours
    load_std_kw: float,        # kW during standard hours
    load_offpeak_kw: float,    # kW during off-peak hours
    rte: float,
    dod: float,
    pvgis_data: dict,
    pv_degradation_pct: float = 0.0,
    load_profile_8760: "np.ndarray | None" = None,  # 8760-element kW array overrides load_map
) -> dict:
    """
    ████████ 8760 / 8760-Hour Physical Dispatch Engine ████████

     Charging Priority (strict order):
      1. PV excess →
      2. Off-peak grid → ×8h22-06h
      3. Daytime std grid → 09-16hpeak ≥ (std/RTE)×1.2
                                          SOC = 3h×min(bess_kw, load_peak)
                                         ×09h→8h16h→1h
      4. Evening std 20-21h → ≤2h
      5. Peak grid →

     Discharge Priority:
       →

     Min-Power Principle:
       = min(bess_kw, need / remaining_hours / RTE)
      → need / op_hrs_left8h
      → need / (17-h)09h8h
      → bess_kw C

     Revenue-First Logic:
      (off_peak) < (standard) < ()
      20-21h ≤2h

    :
      peak_price ≥ (standard_price / RTE) × 1.2
      → Megaflex/Miniflex TOU
      → Nightsave / PPA

    :
      - SOC = × RTE RTE
      - SOC = SOC
    """
    rte_dec   = rte / 100.0
    dod_dec   = dod / 100.0
    usable    = bess_kwh * dod_dec

    load_map = {
        "morning_peak": load_peak_kw,
        "evening_peak": load_peak_kw,
        "peak":         load_peak_kw,   # eThekwini unified peak period
        "standard":     load_std_kw,
        "off_peak":     load_offpeak_kw,
    }

    # ── If a real 8760-h profile was provided, derive better period averages ──
    _use_profile = (load_profile_8760 is not None and len(load_profile_8760) == 8760)
    if _use_profile:
        _arr = load_profile_8760
        _eve_peak_avg = float(np.mean([_arr[g] for g in range(8760) if 17 <= (g % 24) < 20]))
        _eff_load_peak = max(_eve_peak_avg, 1.0)
    else:
        _eff_load_peak = load_peak_kw

    hourly_pv = build_hourly_pv_profile(pvgis_data, pv_kwp)
    hourly_pv *= (1 - pv_degradation_pct / 100.0)

    soc = usable * 0.5
    tot_throughput = 0.0
    records = []

    # Pre-compute standard-period charge target = SOC needed for full evening-peak discharge
    # Evening target: exactly enough to drain SOC to 0 over 3 evening-peak hours.
    # Formula: 3h × min(C-rate power, peak load), capped at usable capacity.
    _EVE_HRS = 3   # Eskom 2025/26: evening peak 17:00-19:59
    std_charge_target = min(usable, _EVE_HRS * min(bess_kw, _eff_load_peak))

    # ── Municipal vs Eskom routing ──────────────────
    _tariff_mode   = st.session_state.get("tariff_mode", "Megaflex ≤300km <500V")
    is_ethekwini   = _tariff_mode.startswith("eThekwini")
    # Use eThekwini holiday table (sat/sun levels) or Eskom set (all → sunday)
    _holiday_db    = ETHEKWINI_HOLIDAYS_2025_26 if is_ethekwini else SA_PUBLIC_HOLIDAYS_2025

    days_in_month = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    g = 0  # global hour index

    for m_i, n_days in enumerate(days_in_month):
        month = m_i + 1
        for d in range(n_days):
            # Determine day type using 2025 calendar
            cal_date = _date(2025, month, d + 1)
            wday = cal_date.weekday()  # 0=Mon, 5=Sat, 6=Sun
            if cal_date in _holiday_db:
                # Eskom: all holidays → "sunday"
                # eThekwini: holidays map to "saturday" or "sunday" level
                day_type = _holiday_db[cal_date] if is_ethekwini else "sunday"
            elif wday == 5:
                day_type = "saturday"
            elif wday == 6:
                day_type = "sunday"
            else:
                day_type = "weekday"

            for h in range(24):
                if is_ethekwini:
                    tariff_price, period = get_ethekwini_tou(h, month, day_type)
                else:
                    tariff_price, period = get_tariff_for_hour(h, month, day_type)
                pv_gen  = hourly_pv[g]
                load    = float(load_profile_8760[g]) if _use_profile else load_map[period]

                pv_to_load   = min(pv_gen, load)
                pv_excess    = pv_gen - pv_to_load
                net_load     = load - pv_to_load

                charge_pv    = 0.0
                charge_grid  = 0.0
                discharge    = 0.0

                # ── Season-aware peak & off-peak prices
                pfx = "w_" if month in WINTER_MONTHS else "s_"
                season_peak  = max(
                    st.session_state[f"{pfx}morning_peak"],
                    st.session_state[f"{pfx}evening_peak"],
                )
                off_peak_p   = st.session_state[f"{pfx}off_peak"]
                standard_p   = st.session_state[f"{pfx}standard"]

                # ── Economic guards ─────────────────────────────
                # : peak > off_peak
                # Night grid-charge only if peak tariff beats off-peak
                grid_charge_viable = season_peak > off_peak_p

                # : tariff > off_peak
                # Peak discharge only if this hour's rate exceeds off-peak
                # (Economic break-even including RTE: tariff_price * rte_dec > off_peak_p)
                discharge_viable = tariff_price > off_peak_p

                if period in ("morning_peak", "evening_peak", "peak"):
                    # "peak" = eThekwini unified peak period (no morning/evening split)
                    if discharge_viable:
                        dis = min(bess_kw, soc, net_load)
                        discharge = max(0.0, dis)
                        soc      -= discharge
                        net_load -= discharge

                    # Absorb surplus PV during peak (charge remaining headroom)
                    if pv_excess > 0 and soc < usable:
                        space = usable - soc
                        c = min(pv_excess, space / rte_dec, bess_kw)
                        charge_pv = c
                        soc += c * rte_dec
                        tot_throughput += c * rte_dec

                elif period == "off_peak":
                    # 1. Absorb surplus PV into BESS
                    if pv_excess > 0 and soc < usable:
                        space = usable - soc
                        c = min(pv_excess, space / rte_dec, bess_kw)
                        charge_pv = c
                        soc += c * rte_dec
                        tot_throughput += c * rte_dec

                    # 2. Grid charge during off-peak (pre-fill before next-day peak)
                    # c = min(bess_kw, need / op_hrs_left / RTE)
                    # Grid charge window: 22:00-06:00 (8h window)
                    # Grid charge ONLY when the next morning has weekday peaks.
                    # Fri/Sat night → Sat (no peaks) → skip.
                    # Sat/Sun night → Sun (no peaks) → skip.
                    # Sun/Mon night → Mon (weekday) → charge. ✓
                    if grid_charge_viable and soc < usable:
                        # Look ahead: determine whether tomorrow is a peak day
                        _next_dt = cal_date + timedelta(days=1) if h >= 22 else cal_date
                        _nwd     = _next_dt.weekday()
                        if _next_dt in _holiday_db:
                            # eThekwini: both "saturday" and "sunday" holidays have no peak
                            _next_type = _holiday_db[_next_dt] if is_ethekwini else "sunday"
                        elif _nwd == 5:
                            _next_type = "saturday"
                        elif _nwd == 6:
                            _next_type = "sunday"
                        else:
                            _next_type = "weekday"
                        next_day_has_peak = (_next_type == "weekday")

                        if next_day_has_peak:
                            # 22:00-06:00
                            # max(1,...) h=6
                            op_hrs_left = max(1, ((24 - h) + 6) if h >= 22 else (6 - h))
                            need = usable - soc
                            c = min(bess_kw, need / op_hrs_left / rte_dec)
                            if c > 0.01:
                                charge_grid = c
                                soc += c * rte_dec
                                tot_throughput += c * rte_dec

                else: # standard
                    # 1. Absorb surplus PV into BESS
                    if pv_excess > 0 and soc < usable:
                        space = usable - soc
                        c = min(pv_excess, space / rte_dec, bess_kw)
                        charge_pv = c
                        soc += c * rte_dec
                        tot_throughput += c * rte_dec

                    # ── 2. 09:00-16:59 +
                    # 20-21h
                    # ≤2h 22:00 ~29%
                    # → 20-21h 22h
                    # 09-16h
                    # peak ≥ (std/RTE) × 1.2Megaflex
                    # min(bess_kw, need / hrs_remaining / RTE)
                    # → 17:00 bess_kw
                    # Grid charge ONLY during daytime standard (09:00-16:59).
                    # Evening standard (20-21h): NEVER charge — off-peak starts ≤2h later
                    # at ~29% lower cost. Revenue-maximising rule: always wait for cheapest.
                    # Power = min(bess_kw, need / hrs_remaining / RTE) — min-power spread.
                    # Daytime standard charging window:
                    #   Eskom:           09:00–17:00 (morning peak 07-09, evening peak 17-20)
                    #   eThekwini high:  08:00–17:00 (morning peak 06-08, evening peak 17-20)
                    #   eThekwini low:   09:00–18:00 (morning peak 07-09, evening peak 18-21)
                    is_high_season = month in WINTER_MONTHS
                    if is_ethekwini and is_high_season:
                        _std_start, _std_end = 8, 17
                    elif is_ethekwini and not is_high_season:
                        _std_start, _std_end = 9, 18
                    else:
                        _std_start, _std_end = 9, 17
                    std_effective_cost = tariff_price / rte_dec
                    if (grid_charge_viable
                            and season_peak >= std_effective_cost * 1.2
                            and soc < std_charge_target #
                            and _std_start <= h < _std_end):
                        hrs_left = max(1, _std_end - h)
                        # SOC = std_charge_target usable
                        # Target = evening-peak target, not full usable capacity
                        need = max(0.0, std_charge_target - soc)
                        c = min(bess_kw, need / hrs_left / rte_dec)
                        if c > 0.01:
                            charge_grid = c
                            soc += c * rte_dec
                            tot_throughput += c * rte_dec

                    # Preserve SOC for upcoming peaks
                    # discharge remains 0

                if discharge > 0:
                    tot_throughput += discharge

                soc = np.clip(soc, 0.0, usable)

                grid_buy = max(0.0, net_load)

                # tariff_price
                pv_saving   = pv_to_load * tariff_price
                dis_saving  = discharge * tariff_price

                # × tariff_price
                if period == "off_peak":
                    grid_charge_cost = charge_grid * st.session_state[f"{pfx}off_peak"]
                else:
                    grid_charge_cost = charge_grid * tariff_price

                net_saving = pv_saving + dis_saving - grid_charge_cost

                records.append({
                    "hour": g, "month": month, "day": d + 1, "hour_of_day": h,
                    "period": period,
                    "tariff_ZAR_kWh": round(tariff_price, 4),
                    "load_kWh": round(load, 4),
                    "pv_gen_kWh": round(pv_gen, 4),
                    "pv_to_load_kWh": round(pv_to_load, 4),
                    "pv_excess_kWh": round(pv_excess, 4),
                    "charge_pv_kWh": round(charge_pv, 4),
                    "charge_grid_kWh": round(charge_grid, 4),
                    "discharge_kWh": round(discharge, 4),
                    "soc_kWh": round(soc, 4),
                    "soc_pct": round(soc / usable * 100, 1) if usable > 0 else 0.0,
                    "grid_buy_kWh": round(grid_buy, 4),
                    "pv_saving_ZAR": round(pv_saving, 4),
                    "bess_net_saving_ZAR": round(dis_saving - grid_charge_cost, 4),
                    "net_saving_ZAR": round(net_saving, 4),
                })
                g += 1

    df = pd.DataFrame(records)
    return {
        "annual_saving_ZAR":      df["net_saving_ZAR"].sum(),
        "annual_pv_saving_ZAR":   df["pv_saving_ZAR"].sum(),
        "annual_bess_saving_ZAR": df["bess_net_saving_ZAR"].sum(),
        "annual_pv_gen_kWh":      df["pv_gen_kWh"].sum(),
        "annual_load_kWh":        df["load_kWh"].sum(),
        "tot_throughput_kWh":     tot_throughput,
        "annual_grid_charge_kWh": df["charge_grid_kWh"].sum(),
        "annual_pv_charge_kWh":   df["charge_pv_kWh"].sum(),
        "annual_discharge_kWh":   df["discharge_kWh"].sum(),
        "annual_grid_buy_kWh":    df["grid_buy_kWh"].sum(),
        "hourly_df": df,
    }


def compute_bess_eol(throughput_yr1: float, bess_kwh: float,
                     cycles: int, dod: float) -> tuple[float, float]:
    """ BESS / Compute BESS EoL and annual cycles
    EoL capped at 20 years (beyond the 25-yr model's meaningful SOH data range).
    """
    usable = bess_kwh * dod / 100.0
    if usable <= 0:
        return 20.0, 0.0
    annual_cycles = throughput_yr1 / (2.0 * usable)
    if annual_cycles <= 0:
        return 20.0, 0.0
    return min(20.0, round(cycles / annual_cycles, 2)), round(annual_cycles, 2)


def run_20yr_financial_model(
    dispatch_yr1: dict,
    pv_kwp: float,
    bess_kwh: float,
    eol_years: float,
    params: dict,
    annual_cycles: float = 365.0,
    c_rate: float = 0.25,
    precomm_bess_months: int = 0,
) -> pd.DataFrame:
    """
    20 Section 12BSOH EoL
    SOH LUNA2000-5015-2S c_rate 0.25/0.33/0.50C
    PV BESS SOH SOH < 60%

    precomm_bess_months: BESS-only period before PV goes live (between BESS go-live and PV go-live).
    During this period BESS earns savings that partially offset the initial CAPEX outflow.
    """
    pv_capex   = pv_kwp   * params["pv_capex_per_kwp"]
    bess_capex = bess_kwh * params["bess_capex_per_kwh"]
    total_capex = pv_capex + bess_capex

    # Depreciation schedule:
    # - PVSection 12B (>1MW) 50/30/20
    # - BESS12B 20%×5
    pure_bess  = (pv_kwp == 0)
    _depr_sched = SECTION_BESS_ONLY if pure_bess else SECTION_12B

    esc  = params["tariff_escalation"] / 100.0
    disc = params["discount_rate"] / 100.0
    tax  = params["tax_rate"] / 100.0
    deg  = params["pv_degradation"] / 100.0

    base_pv_save   = dispatch_yr1["annual_pv_saving_ZAR"]
    base_bess_save = dispatch_yr1["annual_bess_saving_ZAR"]

    # Official SOH curve at actual cycle rate
    soh_arr = get_soh_by_year(annual_cycles, c_rate=c_rate)

    # ── Pre-commissioning BESS-only period ───────────────────────────────────
    # Between BESS go-live and PV go-live (months), BESS earns while PV is still being built.
    # Treated as a Year-0 cash inflow that partially offsets the initial CAPEX.
    # After-tax: BTM company deducts BESS O&M against gross saving; no depreciation yet.
    _precomm_bess_gross = base_bess_save * precomm_bess_months / 12.0
    _precomm_bess_opex  = bess_kwh * params["bess_opex_per_kwh"] * precomm_bess_months / 12.0
    _precomm_ebitda     = _precomm_bess_gross - _precomm_bess_opex
    # Immediate BTM tax realization (same logic as main loop): cash_tax = ebitda × tax
    precomm_net_cf      = _precomm_ebitda * (1 - tax)

    rows = []
    cum_cf = -total_capex + precomm_net_cf   # Year-0 starts partially recovered by BESS precomm
    assessed_loss_cf = 0.0   # SA assessed loss carry-forward (accumulated unabsorbed losses)

    for yr in range(1, ANALYSIS_YEARS + 1):
        esc_mult  = (1 + esc) ** (yr - 1)
        deg_mult  = (1 - deg) ** (yr - 1)

        # SOH from official table (0.0 after EoL, i.e. SOH < 60%)
        soh = soh_arr[yr] if yr < len(soh_arr) else 0.0
        bess_alive = soh >= BESS_EOL_SOH

        # PV SOH
        pv_save   = base_pv_save   * esc_mult * deg_mult
        bess_save = base_bess_save * esc_mult * soh # EoLsoh=0
        saving    = pv_save + bess_save

        # O&M costs scale with 50% of general escalation
        pv_opex   = pv_kwp   * params["pv_opex_per_kwp"]   * (1 + esc * 0.5) ** (yr - 1)
        bess_opex = (bess_kwh * params["bess_opex_per_kwh"] * (1 + esc * 0.5) ** (yr - 1)
                     if bess_alive else 0.0)
        total_opex = pv_opex + bess_opex

        ebitda = saving - total_opex

        # Depreciation (12B 50/30/20 for PV; 20%×5yr straight-line for pure BESS)
        # Only the equipment portion of CAPEX qualifies for Section 12B / 11E.
        # Services (installation, commissioning, civil works) ≈ 40% of total CAPEX
        # are NOT depreciable under these sections.
        _svc_frac    = params.get("service_fraction", 0.40)   # default 40% services
        _equip_capex = total_capex * (1.0 - _svc_frac)        # depreciable equipment only
        depreciation = _equip_capex * _depr_sched.get(yr, 0.0)
        # Tax shield = dep × tax rate — realized immediately (BTM at profitable company)
        tax_shield = depreciation * tax

        ebit = ebitda - depreciation

        # ── BTM Project Tax Treatment ─────────────────────────────────────
        # The host company already has significant taxable income from operations.
        # Section 12B / straight-line depreciation is deducted against THAT income
        # immediately — there is no "standalone entity" loss to carry forward.
        # Net tax impact on the company = EBIT × tax_rate
        #   positive → company pays more tax (savings exceed deduction)
        #   negative → company pays less tax (deduction exceeds project income)
        #              this negative amount is a REAL cash inflow (smaller SARS cheque)
        # NCF = EBITDA − cash_tax = EBITDA×(1−tax) + depreciation×tax
        cash_tax = ebit * tax          # can be negative in 12B years → net_cf > ebitda
        assessed_loss_cf = 0.0         # N/A: absorbed against company's main income
        # ─────────────────────────────────────────────────────────────────

        net_profit = ebit - cash_tax   # = ebit × (1 − tax)
        net_cf   = ebitda - cash_tax   # = EBITDA×(1−tax) + dep×tax
        pv_cf    = net_cf / (1 + disc) ** yr

        cum_cf += net_cf

        rows.append({
            "Year": yr,
            "Status": "✓" if bess_alive else "✗ EoL",
            "SOH%": round(soh * 100, 1),
            "Total Saving (ZAR)": round(saving, 0),
            "PV Saving (ZAR)": round(pv_save, 0),
            "BESS Saving (ZAR)": round(bess_save, 0),
            "PV O&M (ZAR)": round(pv_opex, 0),
            "BESS O&M (ZAR)": round(bess_opex, 0),
            "EBITDA (ZAR)": round(ebitda, 0),
            "Depreciation (ZAR)": round(depreciation, 0),
            "Tax Shield (ZAR)": round(tax_shield, 0),
            "Assess. Loss C/F (ZAR)": round(assessed_loss_cf, 0),
            "Net Profit (ZAR)": round(net_profit, 0),
            "Net Cash Flow NCF (ZAR)": round(net_cf, 0),
            "Discounted CF PV (ZAR)": round(pv_cf, 0),
            "Cumulative CF (ZAR)": round(cum_cf, 0),
        })

    return pd.DataFrame(rows), round(precomm_net_cf, 0)


def compute_npv_irr(fin_df: pd.DataFrame, total_capex: float,
                    discount_rate: float,
                    year0_extra_cf: float = 0.0) -> tuple[float, float]:
    """ NPV IRR/ Compute NPV and IRR via bisection
    year0_extra_cf: pre-commissioning BESS net CF added to Year-0 (reduces effective initial outflow)
    """
    cfs = [-total_capex + year0_extra_cf] + fin_df["Net Cash Flow NCF (ZAR)"].tolist()
    dr = discount_rate / 100.0
    npv = sum(cf / (1 + dr) ** t for t, cf in enumerate(cfs))

    try:
        lo, hi = -0.5, 5.0
        irr = 0.0
        for _ in range(120):
            mid = (lo + hi) / 2
            pv = sum(cf / (1 + mid) ** t for t, cf in enumerate(cfs))
            if abs(pv) < 0.5:
                irr = mid
                break
            if pv > 0:
                lo = mid
            else:
                hi = mid
        irr = mid * 100
    except Exception:
        irr = 0.0

    return round(npv, 0), round(irr, 2)


def compute_lcoe(
    pv_kwp: float,
    bess_kwh: float,
    params: dict,
    dispatch_yr1: dict,
    annual_cycles: float = 0.0,   # unused — kept for call-site compatibility
    c_rate: float = 0.25,         # unused — kept for call-site compatibility
) -> dict:
    """
    Returns separate LCOE and LCOS metrics, each based on Year-1 data only.

    LCOE (PV component only)  — present when pv_kwp > 0
      Cost  = PV CAPEX/life + PV Year-1 O&M   (BESS costs excluded)
      Denom = annual_pv_gen_kWh  (all PV generation avoids grid in BTM)
      → ZAR per kWh of solar generated

    LCOS (BESS component only) — present when bess_kwh > 0
      Cost  = BESS CAPEX/life + BESS Year-1 O&M  (PV costs excluded)
      Denom = annual_discharge_kWh  (kWh the battery actually delivered)
      → ZAR per kWh cycled through the BESS
      BESS arbitrage shifts WHEN you buy, not HOW MUCH, so net avoided kWh
      ≈ 0; discharge kWh is the only meaningful cost denominator.

    Returns:
      {"lcoe": {...} | None, "lcos": {...} | None}
      Each sub-dict: lcoe/lcos_zar_kwh, annual_cost_zar, kwh_yr1, total_mwh
    """
    has_pv   = pv_kwp   > 0
    has_bess = bess_kwh > 0
    result   = {"lcoe": None, "lcos": None}

    if has_pv:
        # ── LCOE: PV costs only ──────────────────────────────────────────
        pv_capex      = pv_kwp * params["pv_capex_per_kwp"]
        pv_opex_yr1   = pv_kwp * params["pv_opex_per_kwp"]
        pv_annual_cost = pv_capex / ANALYSIS_YEARS + pv_opex_yr1
        pv_gen_kwh    = max(0.0, dispatch_yr1.get("annual_pv_gen_kWh", 0))
        lcoe_val      = pv_annual_cost / pv_gen_kwh if pv_gen_kwh > 0 else 0.0
        result["lcoe"] = {
            "lcoe_zar_kwh":      round(lcoe_val, 4),
            "annual_cost_zar":   round(pv_annual_cost, 0),
            "avoided_kwh_yr1":   round(pv_gen_kwh, 0),
            "total_avoided_mwh": round(pv_gen_kwh * ANALYSIS_YEARS / 1_000, 1),
        }

    if has_bess:
        # ── LCOS: BESS costs only ────────────────────────────────────────
        bess_capex       = bess_kwh * params["bess_capex_per_kwh"]
        bess_opex_yr1    = bess_kwh * params["bess_opex_per_kwh"]
        bess_annual_cost = bess_capex / ANALYSIS_YEARS + bess_opex_yr1
        discharge_kwh    = max(0.0, dispatch_yr1.get("annual_discharge_kWh", 0))
        lcos_val         = bess_annual_cost / discharge_kwh if discharge_kwh > 0 else 0.0
        result["lcos"] = {
            "lcos_zar_kwh":        round(lcos_val, 4),
            "annual_cost_zar":     round(bess_annual_cost, 0),
            "discharge_kwh_yr1":   round(discharge_kwh, 0),
            "total_discharge_mwh": round(discharge_kwh * ANALYSIS_YEARS / 1_000, 1),
        }

    return result


# ─────────────────────────────────────────────────────────────
# Auto-PVGIS on location/capacity change
# ─────────────────────────────────────────────────────────────
def check_auto_pvgis():
    """ PVGIS / Auto-fetch PVGIS when location or PV changes"""
    if st.session_state.pv_kwp <= 0:
        return
    # Use location+capacity+angles as cache key
    key = (f"{st.session_state.lat:.3f}_{st.session_state.lon:.3f}"
           f"_{st.session_state.pv_kwp:.0f}_{st.session_state.pv_loss:.1f}"
           f"_{st.session_state.tilt:.1f}_{st.session_state.azimuth:.1f}")
    if st.session_state.pvgis_key != key:
        result = get_pvgis_data(
            st.session_state.lat, st.session_state.lon,
            st.session_state.pv_kwp, st.session_state.pv_loss,
            st.session_state.tilt, st.session_state.azimuth,
        )
        st.session_state.pvgis_data  = result
        st.session_state.pvgis_status = result["status"]
        st.session_state.annual_pv_kwh = result["annual_kwh"]
        st.session_state.pvgis_key   = key


# ─────────────────────────────────────────────────────────────
# Professional Excel Report Generator
# 6 sheets: Cover · Parameters · 20Y Model · Monthly · Charts · Raw Data
# ─────────────────────────────────────────────────────────────
def generate_excel_report() -> bytes:
    """
     6-sheet Excel session_state
    Sheet 3 $B$5 Excel
    Returns raw bytes for st.download_button.
    """
    import datetime as _dt
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, LineChart, Reference

    # ── Session-state snapshot ────────────────────────────────
    ss      = st.session_state
    res     = ss.results
    fin_df  = ss.fin_df
    h_df    = ss.hourly_df
    d1      = res["dispatch_yr1"]
    pvg     = res["pvgis_data"]
    soh_arr = get_soh_by_year(res["annual_cycles"], c_rate=res.get("c_rate", 0.25))
    pv_zar, bess_zar = get_capex_zar()
    bess_zero = (ss.bess_kwh == 0)   # flag: hide BESS charts/columns when no BESS
    pv_zero   = (ss.pv_kwp   == 0)   # flag: pure BESS → non-12B depreciation

    # ── Colour palette ────────────────────────────────────────
    C_NAVY   = "1F3864"
    C_DARK   = "2E4057"
    C_MID    = "2E5F8A"
    C_LTBLUE = "D6E4F0"
    C_LOCKED = "FFF9C4"
    C_ALT    = "F2F6FC"
    C_GREEN  = "196F3D"
    C_RED    = "922B21"

    _thin = Side(style="thin", color="BDBDBD")

    def _fill(hex_c):
        return PatternFill("solid", fgColor=hex_c)

    def _font(bold=False, sz=10, color="1A1A2E", italic=False):
        return Font(name="Calibri", size=sz, bold=bold, color=color, italic=italic)

    def _align(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    def _bdr():
        return Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

    # ── Language helper: strip Chinese from bilingual labels ──────────────────
    # In English mode: remove CJK characters so only English text appears in Excel
    def _L(s: str) -> str:
        """Pass-through — all labels are already English-only."""
        return s

    wb = Workbook()

    # ════════════════════════════════════════════════════════════
    # SHEET 1 — Cover
    # ════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Cover"
    ws1.sheet_view.showGridLines = False
    for ci, w in enumerate([2, 28, 20, 20, 20, 20, 2], 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w

    # Title banner
    ws1.merge_cells("B2:G2")
    c = ws1["B2"]
    c.value = "BTM PV+BESS Financial Feasibility Report"
    c.font = Font("Calibri", size=16, bold=True, color="FFFFFF")
    c.fill = _fill(C_NAVY); c.alignment = _align("center")
    ws1.row_dimensions[2].height = 42

    ws1.merge_cells("B3:G3")
    c = ws1["B3"]
    c.value = "Powered by Huawei SA Digital Power  ·  South Africa C&I BTM Solar + Storage"
    c.font = Font("Calibri", size=9, color="D6E4F0")
    c.fill = _fill(C_NAVY); c.alignment = _align("center")
    ws1.row_dimensions[3].height = 16

    # Project overview header
    ws1.merge_cells("B5:G5")
    c = ws1["B5"]
    _en_mode = (True)
    c.value = (" Project Overview" if _en_mode else " / Project Overview")
    c.font = Font("Calibri", size=10, bold=True, color="FFFFFF")
    c.fill = _fill(C_DARK); c.alignment = _align("left")
    ws1.row_dimensions[5].height = 20

    payback_str = f"{res['payback']:.2f} yr" if res['payback'] else "20yr+"
    if _en_mode:
        info_rows = [
            ("Location",             f"Lat {ss.lat:.3f}°  /  Lon {ss.lon:.3f}°"),
            ("System Size",          f"PV  {_fmw(ss.pv_kwp,'kWp')}  +  BESS  {_fmw(ss.bess_kwh,'kWh')}"),
            ("Tariff Mode",          ss.get("tariff_mode", "—")),
            ("Forex Rate",           f"1 USD = {ss.forex_usd_zar:.2f} ZAR"),
            ("Tariff Escalation",    f"{ss.tariff_escalation:.1f}% / yr"),
            ("Discount Rate",        f"{ss.discount_rate:.1f}%"),
            ("PVGIS Annual PV Gen",  _fmw(pvg.get('annual_kwh', 0), "kWh/yr")),
            ("Export Time",          _dt.datetime.now().strftime("%Y-%m-%d  %H:%M")),
        ]
    else:
        info_rows = [
            (" Location", f"Lat {ss.lat:.3f}° / Lon {ss.lon:.3f}°"),
            (" System Size", f"PV {_fmw(ss.pv_kwp,'kWp')} + BESS {_fmw(ss.bess_kwh,'kWh')}"),
            ("Tariff Mode",    ss.get("tariff_mode", "—")),
            (" Forex Rate", f"1 USD = {ss.forex_usd_zar:.2f} ZAR"),
            (" Tariff Escalation", f"{ss.tariff_escalation:.1f}% / yr"),
            (" Discount Rate", f"{ss.discount_rate:.1f}%"),
            ("PVGIS PV Gen", _fmw(pvg.get('annual_kwh', 0), "kWh/yr")),
            (" Export Time", _dt.datetime.now().strftime("%Y-%m-%d %H:%M")),
        ]
    for i, (lbl, val) in enumerate(info_rows):
        r = 6 + i
        ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        cl = ws1.cell(row=r, column=2, value=lbl)
        cl.font = _font(bold=True, sz=10); cl.fill = _fill(C_LTBLUE)
        cl.alignment = _align(); cl.border = _bdr()
        ws1.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
        cv = ws1.cell(row=r, column=4, value=val)
        cv.font = _font(sz=10); cv.fill = _fill("FFFFFF")
        cv.alignment = _align(); cv.border = _bdr()
        ws1.row_dimensions[r].height = 18

    # Timeline row (row 14) — between info block and KPI header
    _tl_res = res  # res is the results dict available in generate_excel_report
    _bess_lead_xl = _tl_res.get("bess_lead_months", ss.get("bess_lead_months", 6))
    _pv_lead_xl   = _tl_res.get("pv_lead_months",   ss.get("pv_lead_months",  12))
    _po_xl        = _tl_res.get("po_date",        _dt.date.today().isoformat())
    _bess_gol_xl  = _tl_res.get("bess_golive",    "—")
    _pv_gol_xl    = _tl_res.get("pv_golive",      "—")
    _end_xl       = _tl_res.get("model_end",      "—")
    _pcomm_xl     = _tl_res.get("precomm_months", 0) or 0
    _pcomm_ncf_xl = _tl_res.get("precomm_ncf",   0.0) or 0.0

    if _en_mode:
        _tl_str = (
            f"Project Timeline:  PO: {_po_xl}  →  BESS: {_bess_gol_xl} (+{_bess_lead_xl}mo)"
            f"  →  PV: {_pv_gol_xl} (+{_pv_lead_xl}mo)  →  End: {_end_xl} (PV+{ANALYSIS_YEARS}yr)"
            + (f"  |  BESS pre-comm: {_pcomm_xl}mo  →  Year-0 BESS net: R {_pcomm_ncf_xl:,.0f}"
               if _pcomm_xl > 0 else "")
        )
    else:
        _tl_str = (
            f"PO: {_po_xl} → BESS: {_bess_gol_xl} (+{_bess_lead_xl})"
            f" → PV: {_pv_gol_xl} (+{_pv_lead_xl}) → : {_end_xl} (PV{ANALYSIS_YEARS})"
            + (f" | BESS: {_pcomm_xl} → Year-0: R {_pcomm_ncf_xl:,.0f}"
               if _pcomm_xl > 0 else "")
        )
    ws1.merge_cells("B14:G14")
    c14 = ws1.cell(row=14, column=2, value=_tl_str)
    c14.font = Font("Calibri", size=9, italic=True, color=C_NAVY)
    c14.fill = _fill("EAF0FB"); c14.alignment = _align(wrap=False); c14.border = _bdr()
    ws1.row_dimensions[14].height = 16

    # KPI section header
    ws1.merge_cells("B15:G15")
    c = ws1["B15"]
    c.value = (" Key Financial Metrics" if _en_mode else " / Key Financial Metrics")
    c.font = Font("Calibri", size=10, bold=True, color="FFFFFF")
    c.fill = _fill(C_DARK); c.alignment = _align("left")
    ws1.row_dimensions[15].height = 20

    if _en_mode:
        kpis = [
            ("Total CAPEX",         f"R {res['total_capex']/1e6:.2f} M",      "ZAR"),
            ("NPV (20yr)",          f"R {res['npv']/1e6:.2f} M",              "ZAR"),
            ("Project IRR",         f"{res['irr']:.2f}%",                     "IRR"),
            ("Simple Payback",      payback_str,                              ""),
            ("Yr1 Total Saving",    f"R {d1['annual_saving_ZAR']/1e6:.2f} M", "ZAR/yr"),
            ("BESS EoL",            f"Year {min(20.0, res['eol_years']):.1f}", "SOH < 60%"),
        ]
    else:
        kpis = [
            (" CAPEX", f"R {res['total_capex']/1e6:.2f} M", "ZAR"),
            (" NPV 20yr", f"R {res['npv']/1e6:.2f} M", "ZAR"),
            (" IRR", f"{res['irr']:.2f}%", "Project IRR"),
            (" Payback", payback_str, "Simple Payback"),
            ("1 Yr1 Save", f"R {d1['annual_saving_ZAR']/1e6:.2f} M", "ZAR/yr"),
            ("BESS EoL", f"Year {min(20.0, res['eol_years']):.1f}", "SOH < 60%"),
        ]
    # 3 cols × 2 rows layout
    kpi_layout = [
        ((16, 17), (2, 3)), ((16, 17), (4, 5)), ((16, 17), (6, 6)),
        ((19, 20), (2, 3)), ((19, 20), (4, 5)), ((19, 20), (6, 6)),
    ]
    for idx, ((r_top, r_bot), (c_l, c_r)) in enumerate(kpi_layout):
        if idx >= len(kpis):
            break
        lbl_k, val_k, _ = kpis[idx]
        for r_ in (r_top, r_bot):
            if c_l != c_r:
                ws1.merge_cells(start_row=r_, start_column=c_l,
                                end_row=r_, end_column=c_r)
            ws1.row_dimensions[r_].height = 22 if r_ == r_top else 26
        cl_k = ws1.cell(row=r_top, column=c_l, value=lbl_k)
        cl_k.font = _font(sz=8, bold=True, color="555555")
        cl_k.fill = _fill(C_LTBLUE); cl_k.alignment = _align("center"); cl_k.border = _bdr()
        cv_k = ws1.cell(row=r_bot, column=c_l, value=val_k)
        kpi_color = (C_GREEN if res["npv"] > 0 else C_RED) if idx == 1 else C_NAVY
        cv_k.font = Font("Calibri", size=13, bold=True, color=kpi_color)
        cv_k.fill = _fill(C_LTBLUE); cv_k.alignment = _align("center"); cv_k.border = _bdr()
    ws1.row_dimensions[18].height = 6  # gap row

    # Instructions
    ws1.merge_cells("B22:G22")
    c = ws1["B22"]
    c.value = (" Instructions" if _en_mode else " / Instructions")
    c.font = Font("Calibri", size=10, bold=True, color="FFFFFF")
    c.fill = _fill(C_MID); c.alignment = _align("left")
    ws1.row_dimensions[22].height = 20

    if _en_mode:
        notes_text = (
            "● Yellow cells are locked export snapshots (PVGIS / dispatch results) — do not edit\n"
            "● Green cells in Sheet 3 Row 4 are editable — all financial formulas recalculate automatically\n"
            "● Charts (Sheet 5) are linked to Sheet 3 & 4 — they update automatically when you change parameters\n"
            "● Sheet 6 is a static 8,760-hour dispatch snapshot — it does not update with parameter changes\n"
            "● Adjustable: tariff escalation, discount rate, tax rate, PV/BESS unit costs, O&M rates"
        )
    else:
        notes_text = (
            "● Yellow cells are locked export snapshots (PVGIS / dispatch results) — do not edit"
            "● Green cells in Sheet 3 Row 4 are editable — all financial formulas recalculate automatically / Sheet3 4\n"
            "● Charts (Sheet 5) are linked to Sheet 3 & 4 — they update automatically when you change parameters"
            "● Sheet 6 is a static 8,760-hour dispatch snapshot — it does not update with parameter changes / Sheet6 \n"
            "● Adjustable: tariff escalation, discount rate, tax rate, PV/BESS unit costs, O&M rates"
        )
    ws1.merge_cells("B23:G26")
    cn = ws1.cell(row=23, column=2, value=notes_text)
    cn.font = Font("Calibri", size=9, color="444444")
    cn.alignment = Alignment(wrap_text=True, vertical="top")
    cn.fill = _fill("F9F9F9"); cn.border = _bdr()
    for r_ in range(23, 27):
        ws1.row_dimensions[r_].height = 15

    # ════════════════════════════════════════════════════════════
    # SHEET 2 — Parameters
    # ════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Parameters")
    ws2.sheet_view.showGridLines = False
    for ci, w in enumerate([3, 34, 18, 14, 30], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    ws2.merge_cells("B1:E1")
    _ws2_title = "Parameters  ——  White = editable · Yellow = locked"
    c = ws2.cell(row=1, column=2, value=_ws2_title)
    c.font = Font("Calibri", size=12, bold=True, color="FFFFFF")
    c.fill = _fill(C_NAVY); c.alignment = _align("center")
    ws2.row_dimensions[1].height = 30

    def _p_sec(r, title, bg=C_DARK):
        ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        c_ = ws2.cell(row=r, column=1, value=_L(title))
        c_.font = Font("Calibri", size=10, bold=True, color="FFFFFF")
        c_.fill = _fill(bg); c_.alignment = _align("left")
        ws2.row_dimensions[r].height = 18

    def _p_row(r, label, value, unit="", locked=False, fmt=None, formula=None):
        lk = ws2.cell(row=r, column=1, value="🔒" if locked else "")
        lk.font = _font(sz=9); lk.alignment = _align("center")
        lb = ws2.cell(row=r, column=2, value=_L(label))
        lb.font = _font(sz=10); lb.fill = _fill(C_ALT)
        lb.alignment = _align(); lb.border = _bdr()
        vc = ws2.cell(row=r, column=3, value=(formula if formula else value))
        vc.fill = _fill(C_LOCKED if locked else "FFFFFF")
        vc.alignment = _align("center"); vc.border = _bdr()
        if fmt:
            vc.number_format = fmt
        un = ws2.cell(row=r, column=4, value=unit)
        un.font = _font(sz=9, italic=True, color="666666")
        un.fill = _fill(C_ALT); un.alignment = _align(); un.border = _bdr()
        ws2.row_dimensions[r].height = 17
        return vc

    # ── System Parameters (rows 3-12) ─────────────────────────
    _p_sec(3, "▌ / System Parameters")
    _p_row(4, " PV Capacity", ss.pv_kwp, "kWp", fmt="#,##0.0")
    _p_row(5, " BESS Capacity", ss.bess_kwh, "kWh", fmt="#,##0.0")
    _p_row(6,  "C-rate",                  ss.c_rate_label,     "—")
    _p_row(7, " RTE", ss.rte, "%", fmt="0.0")
    _p_row(8, " DoD", ss.dod, "%", fmt="0.0")
    _p_row(9, "PV PV Degradation", ss.pv_degradation, "%/yr", fmt="0.00")
    _p_row(10, " Tilt", ss.tilt, "°", fmt="0.0")
    _p_row(11, " Azimuth", ss.azimuth, "° (180=N)", fmt="0.0")

    # ── Load Parameters (rows 13-17) ──────────────────────────
    _p_sec(13, "▌ / Load Profile")
    _p_row(14, " Peak Load", ss.load_peak_kw, "kW", fmt="#,##0.0")
    _p_row(15, " Standard Load", ss.load_std_kw, "kW", fmt="#,##0.0")
    _p_row(16, " Off-peak Load", ss.load_offpeak_kw, "kW", fmt="#,##0.0")

    # ── CAPEX & Forex (rows 18-25) ─────────────────────────────
    _p_sec(18, "▌ / CAPEX & Forex")
    _p_row(19, " PV Cost", ss.pv_usd_per_w, "USD/W", fmt="0.000")
    _p_row(20, " BESS Cost", ss.bess_usd_per_wh, "USD/Wh", fmt="0.000")
    _p_row(21, " Forex Rate", ss.forex_usd_zar, "USD/ZAR", fmt="0.00")
    _p_row(22, "PV ZAR/kWp", pv_zar, "ZAR/kWp", fmt="#,##0.00",
           formula="=C19*1000*C21")
    _p_row(23, "BESS ZAR/kWh", bess_zar, "ZAR/kWh", fmt="#,##0.00",
           formula="=C20*1000*C21")
    _p_row(24, " Total CAPEX", res["total_capex"], "ZAR", fmt="#,##0",
           formula="=C4*C22+C5*C23")

    # ── OPEX (rows 26-29) ─────────────────────────────────────
    _p_sec(26, "▌ / OPEX")
    _p_row(27, "PV PV O&M", ss.pv_opex_per_kwp, "ZAR/kWp/yr", fmt="0.00")
    _p_row(28, "BESS BESS O&M", ss.bess_opex_per_kwh, "ZAR/kWh/yr", fmt="0.00")

    # ── Financial Assumptions (rows 30-34) ────────────────────
    _p_sec(30, "▌ / Financial Assumptions")
    _p_row(31, " Tariff Escalation", ss.tariff_escalation, "%/yr", fmt="0.0")
    _p_row(32, " Discount Rate", ss.discount_rate, "%", fmt="0.0")
    _p_row(33, " Corp Tax Rate", ss.tax_rate, "%", fmt="0.0")

    # ── Project Timeline (row 34) ──────────────────────────────
    _p_sec(34, "▌ / Project Timeline")
    _p_row(35, "PO PO Date", res.get("po_date", _dt.date.today().isoformat()), "—", locked=True)
    _p_row(36, "BESS BESS Lead", ss.get("bess_lead_months", 6), "months", fmt="0")
    _p_row(37, "PV PV Lead", ss.get("pv_lead_months", 12), "months", fmt="0")
    _p_row(38, "BESS BESS Go-Live", res.get("bess_golive", "—"), "—", locked=True)
    _p_row(39, "PV PV Go-Live", res.get("pv_golive", "—"), "—", locked=True)
    _p_row(40, " Model End", res.get("model_end", "—"), "—", locked=True)
    _p_row(41, "BESS Pre-comm (mo)", res.get("precomm_months", 0), "months", fmt="0", locked=True)
    _p_row(42, "Year-0 BESS Pre-comm NCF", res.get("precomm_ncf", 0.0), "ZAR", fmt="#,##0", locked=True)

    # ── Tariff Rates (rows 44-52) ─────────────────────────────
    _p_sec(44, "▌ / Tariff Rates (ZAR/kWh incl VAT)")
    _p_row(45, "Tariff Mode",             ss.get("tariff_mode", "—"), "—", locked=True)
    _p_row(46, " Winter Peak", ss.get("w_morning_peak", 0), "ZAR/kWh", fmt="0.0000")
    _p_row(47, " Winter Std", ss.get("w_standard", 0), "ZAR/kWh", fmt="0.0000")
    _p_row(48, " Winter Off-peak", ss.get("w_off_peak", 0), "ZAR/kWh", fmt="0.0000")
    _p_row(49, " Summer Peak", ss.get("s_morning_peak", 0),"ZAR/kWh", fmt="0.0000")
    _p_row(50, " Summer Std", ss.get("s_standard", 0), "ZAR/kWh", fmt="0.0000")
    _p_row(51, " Summer Off-peak", ss.get("s_off_peak", 0), "ZAR/kWh", fmt="0.0000")

    # ── PVGIS Data LOCKED (rows 53-69) ────────────────────────
    _p_sec(53, "▌ PVGIS LOCKED — ", bg=C_MID)
    monthly_kwh = pvg.get("monthly_kwh", [0] * 12)
    _p_row(54, " Latitude", ss.lat, "°", locked=True, fmt="0.000")
    _p_row(55, " Longitude", ss.lon, "°", locked=True, fmt="0.000")
    _p_row(56, " Annual PV", pvg.get("annual_kwh", 0), "kWh/yr", locked=True, fmt="#,##0")
    for mi, mn in enumerate(["Jan","Feb","Mar","Apr","May","Jun",
                              "Jul","Aug","Sep","Oct","Nov","Dec"]):
        val_m = monthly_kwh[mi] if mi < len(monthly_kwh) else 0
        _pv_gen_lbl = f"{mn} PV Generation"
        _p_row(57 + mi, _pv_gen_lbl, val_m, "kWh", locked=True, fmt="#,##0")

    # ── Dispatch Results LOCKED (rows 70-74) ──────────────────
    _p_sec(70, "▌ Yr1 Dispatch Results (Locked)", bg=C_MID)
    _p_row(71, "1 PV Yr1 PV Saving", d1["annual_pv_saving_ZAR"], "ZAR/yr", locked=True, fmt="#,##0")
    _p_row(72, "1 BESS Yr1 BESS Save", d1["annual_bess_saving_ZAR"], "ZAR/yr", locked=True, fmt="#,##0")
    _p_row(73, " Annual Cycles", res["annual_cycles"], "/yr", locked=True, fmt="0.00")
    _p_row(74, "BESS BESS EoL", min(20.0, res["eol_years"]), "years", locked=True, fmt="0.0")

    # ── SOH Table LOCKED (rows 76-97) ─────────────────────────
    _p_sec(76, "▌ BESS SOH Degradation Table — Huawei LUNA2000-2236-1S (Locked)", bg=C_MID)
    _soh_hdrs = ["Year", "SOH (%)", "Status"]
    for ci_h, hdr_h in enumerate(_soh_hdrs, start=2):
        hc = ws2.cell(row=77, column=ci_h, value=hdr_h)
        hc.font = _font(bold=True, sz=9, color="FFFFFF")
        hc.fill = _fill(C_DARK); hc.alignment = _align("center"); hc.border = _bdr()
    ws2.row_dimensions[77].height = 15
    for yr_i in range(1, ANALYSIS_YEARS + 1):  # years 1..20 → Parameters rows 78..97
        sr = 77 + yr_i
        sv = soh_arr[yr_i] if yr_i < len(soh_arr) else 0.0
        alive_s = sv >= BESS_EOL_SOH
        yc = ws2.cell(row=sr, column=2, value=yr_i)
        yc.font = _font(sz=9); yc.fill = _fill(C_LOCKED)
        yc.alignment = _align("center"); yc.border = _bdr()
        sc = ws2.cell(row=sr, column=3, value=round(sv * 100, 2))
        sc.font = _font(sz=9, color=C_GREEN if alive_s else C_RED)
        sc.fill = _fill(C_LOCKED); sc.alignment = _align("center")
        sc.number_format = "0.00"; sc.border = _bdr()
        stc = ws2.cell(row=sr, column=4, value="✓ Active" if alive_s else "✗ EoL")
        stc.font = _font(sz=9, bold=True, color=C_GREEN if alive_s else C_RED)
        stc.fill = _fill(C_LOCKED); stc.alignment = _align("center"); stc.border = _bdr()
        ws2.row_dimensions[sr].height = 15

    # ════════════════════════════════════════════════════════════
    # SHEET 3 — 20Y Model
    # $B$5 Excel
    # ════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("20Y Financial Model")
    ws3.sheet_view.showGridLines = False
    ws3.freeze_panes = "A8"    # freeze title + param section + header
    for ci, w in enumerate([6,14,10,14,14,14,12,12,14,13,12,12,12,14,13,14,12,14], 1):
        ws3.column_dimensions[get_column_letter(ci)].width = w

    # ── Title row 1 ───────────────────────────────────────────
    ws3.merge_cells("A1:R1")
    _p2_sheet = "Parameters"   # sheet name
    _ws3_title = f"20-Year Financial Model — Edit white cells in '{_p2_sheet}' sheet · Row 4 formulas auto-pull & recalculate"
    c = ws3.cell(row=1, column=1, value=_ws3_title)
    c.font = Font("Calibri", size=12, bold=True, color="FFFFFF")
    c.fill = _fill(C_NAVY); c.alignment = _align("center")
    ws3.row_dimensions[1].height = 28

    # ── Parameter Section header row 2 ───────────────────────
    ws3.merge_cells("A2:R2")
    _ws3_row2 = f"⚙️  Row 4 = cross-sheet links to '{_p2_sheet}' — 🟢 Edit white cells in Parameters sheet → auto-update · 🔒 Locked = re-run simulation"
    c = ws3.cell(row=2, column=1, value=_ws3_row2)
    c.font = Font("Calibri", size=9, bold=True, color="FFFFFF")
    c.fill = _fill(C_MID); c.alignment = _align("left")
    ws3.row_dimensions[2].height = 18

    # ── Parameter labels row 3 & values row 4 ────────────────
    # Parameters mapped to columns B-L (2-12)
    # B5: Base PV Save  C5: Base BESS Save  D5: Esc%  E5: Disc%  F5: Tax%
    # G5: PV Deg%  H5: Total CAPEX  I5: PV kWp  J5: BESS kWh
    # K5: PV O&M  L5: BESS O&M
    # Row 4 Sheet 2Sheet 2
    # D $X$4
    # Cross-sheet links: Row 4 cells reference ' Parameters' sheet.
    # Changing any white cell in Sheet 2 propagates to the full 20Y model instantly.
    _P2 = f"'Parameters'" # dynamic: matches actual sheet name
    param_meta = [
        # (col_idx, label, value_fallback, editable, fmt, sheet2_formula)
        (2, "Yr1 PV Save\nPV\n(ZAR) 🔒", d1["annual_pv_saving_ZAR"], False, "#,##0", f"={_P2}!C71"),
        (3, "Yr1 BESS Save\nBESS\n(ZAR) 🔒", d1["annual_bess_saving_ZAR"], False, "#,##0", f"={_P2}!C72"),
        (4, "Tariff Esc.\n\n(%/yr) 🟢 Sheet2", ss.tariff_escalation, True, "0.0", f"={_P2}!C31"),
        (5, "Discount Rate\n\n(%) 🟢 Sheet2", ss.discount_rate, True, "0.0", f"={_P2}!C32"),
        (6, "Tax Rate\n\n(%) 🟢 Sheet2", ss.tax_rate, True, "0.0", f"={_P2}!C33"),
        (7, "PV Degrad.\nPV\n(%/yr) 🟢 Sheet2", ss.pv_degradation, True, "0.00", f"={_P2}!C9"),
        (8, "Total CAPEX\n\n(ZAR) 🔗 Sheet2", res["total_capex"], False, "#,##0", f"={_P2}!C24"),
        (9, "PV Capacity\n\n(kWp) 🔗 Sheet2", ss.pv_kwp, False, "#,##0.0", f"={_P2}!C4"),
        (10, "BESS Capacity\n\n(kWh) 🔗 Sheet2", ss.bess_kwh, False, "#,##0.0", f"={_P2}!C5"),
        (11, "PV O&M\nPV\n(ZAR/kWp) 🟢 Sheet2", ss.pv_opex_per_kwp, True, "0.00", f"={_P2}!C27"),
        (12, "BESS O&M\nBESS\n(ZAR/kWh) 🟢 Sheet2", ss.bess_opex_per_kwh, True, "0.00", f"={_P2}!C28"),
    ]
    ws3.row_dimensions[3].height = 34
    ws3.row_dimensions[4].height = 20
    ws3.row_dimensions[5].height = 6   # thin divider
    for ci_p, lbl_p, val_p, edit_p, fmt_p, formula_ref in param_meta:
        # Label row 3
        lc3 = ws3.cell(row=3, column=ci_p, value=_L(lbl_p))
        lc3.font = _font(sz=8, bold=True)
        lc3.fill = _fill(C_LTBLUE)
        lc3.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        lc3.border = _bdr()
        # Value row 4 — cross-sheet formula linking to ' Parameters' sheet
        # 🟢 = edit in Sheet 2 (white cell) to drive this model
        # 🔗 = driven by Sheet 2 formula (CAPEX, capacity)
        # 🔒 = locked simulation result (Yr1 savings — re-run to update)
        vc4 = ws3.cell(row=4, column=ci_p, value=formula_ref)
        vc4.font = _font(sz=10, bold=True,
                         color="1B5E20" if edit_p else C_DARK)
        vc4.fill = _fill("E8F5E9" if edit_p else C_LOCKED)
        vc4.alignment = _align("center")
        vc4.border = _bdr()
        vc4.number_format = fmt_p

    # Label col A rows 3-4
    for r_ in (3, 4):
        ac = ws3.cell(row=r_, column=1,
                      value=("Parameter" if r_ == 3 else "Value"))
        ac.font = _font(sz=8, bold=True, color="FFFFFF")
        ac.fill = _fill(C_NAVY); ac.alignment = _align("center"); ac.border = _bdr()
    # Fill cols M-R rows 3-4 as spacers
    for r_ in (3, 4):
        for ci_sp in range(13, 19):
            ws3.cell(row=r_, column=ci_sp).fill = _fill(C_ALT)
            ws3.cell(row=r_, column=ci_sp).border = _bdr()

    # Divider row 5
    ws3.merge_cells("A5:R5")
    ws3.cell(row=5, column=1).fill = _fill(C_NAVY)

    # ── Column headers row 6 ─────────────────────────────────
    hdrs3 = ["Year", "BESS\nStatus", "SOH\n%",
             "PV Saving\nZAR", "BESS Saving\nZAR", "Total Saving\nZAR",
             "PV O&M\nZAR", "BESS O&M\nZAR", "EBITDA\nZAR",
             ("Depreciation\n(20%×5yr)\nZAR" if pv_zero else "Sec12B Depr.\n(50/30/20)\nZAR"), "Tax Shield\n(realized)", "EBIT\nZAR",
             "Net Tax\nZAR", "Net CF\nNCF ZAR", "Disc. CF\nPV ZAR",
             "Cum. CF\nZAR", "Net Profit\nZAR", "Assess.Loss\nB/F\n(N/A-BTM)"]
    for ci, hdr in enumerate(hdrs3, 1):
        hc = ws3.cell(row=6, column=ci, value=hdr)
        hc.font = Font("Calibri", size=8, bold=True, color="FFFFFF")
        hc.fill = _fill(C_NAVY)
        hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        hc.border = _bdr()
    ws3.row_dimensions[6].height = 32

    # ── Year 0 row (row 7) ────────────────────────────────────
    # $H$4 = Total CAPEX (parameter value cell)
    r0 = 7
    _precomm_ncf_xl2 = float(res.get("precomm_ncf", 0.0) or 0.0)
    _precomm_mo_xl2  = int(res.get("precomm_months", 0) or 0)
    y0 = ws3.cell(row=r0, column=1, value=0)
    y0.font = _font(bold=True, sz=10); y0.fill = _fill(C_ALT)
    y0.alignment = _align("center"); y0.border = _bdr()
    if _precomm_mo_xl2 > 0:
        _y0_label = f"Initial Investment  |  BESS pre-commission: {_precomm_mo_xl2} months"
    else:
        _y0_label = "Initial Investment"
    ws3.merge_cells(start_row=r0, start_column=2, end_row=r0, end_column=13)
    lc = ws3.cell(row=r0, column=2, value=_y0_label)
    lc.font = Font("Calibri", size=10, bold=True, color=C_RED)
    lc.fill = _fill(C_ALT); lc.alignment = _align("center"); lc.border = _bdr()
    # NCF yr0 = -CAPEX + precomm BESS net saving
    _y0_ncf_val = -res["total_capex"] + _precomm_ncf_xl2
    for col_y0, y0_value in [
        (14, _y0_ncf_val),           # NCF yr0
        (15, _y0_ncf_val),           # DiscCF yr0 (not discounted since yr=0)
        (16, _y0_ncf_val),           # CumCF yr0
    ]:
        cc = ws3.cell(row=r0, column=col_y0, value=round(y0_value, 0))
        cc.font = Font("Calibri", size=10, bold=True, color=C_RED if _y0_ncf_val < 0 else C_GREEN)
        cc.fill = _fill("FFF0F0"); cc.alignment = _align("center")
        cc.number_format = "#,##0"; cc.border = _bdr()
    ws3.row_dimensions[r0].height = 20

    # ── Years 1-20 (rows 8-27) ───────────────────────────────
    # Parameter absolute refs:  $B$4=BasePVSave  $C$4=BaseBESSSave
    #   $D$4=Esc%  $E$4=Disc%  $F$4=Tax%  $G$4=PVDeg%
    #   $H$4=CAPEX  $I$4=PVkWp  $J$4=BESSCkWh  $K$4=PVOM  $L$4=BESSOM

    for yr in range(1, ANALYSIS_YEARS + 1):
        r = r0 + yr          # rows 8..27
        sv    = soh_arr[yr] if yr < len(soh_arr) else 0.0
        alive = sv >= BESS_EOL_SOH
        alt_bg = C_ALT if yr % 2 == 0 else "FFFFFF"
        _xl_depr = SECTION_BESS_ONLY if pv_zero else SECTION_12B
        dep_pct  = _xl_depr.get(yr, 0.0)

        def _wc(col, val=None, formula=None, fmt="#,##0",
                bg=None, bold=False, color="1A1A2E"):
            cell = ws3.cell(row=r, column=col,
                            value=(formula if formula else val))
            cell.fill = _fill(bg or alt_bg)
            cell.font = Font("Calibri", size=9, bold=bold, color=color)
            cell.alignment = _align("center")
            cell.number_format = fmt
            cell.border = _bdr()
            return cell

        # Col A: Year number
        _wc(1, val=yr, fmt="0", bold=True, bg=C_LTBLUE)

        # Col B: BESS status
        if bess_zero:
            b_label, b_color = "— N/A", "888888"
        elif alive:
            b_label, b_color = "✓ Active", C_GREEN
        else:
            b_label, b_color = "✗ EoL", C_RED
        bs = ws3.cell(row=r, column=2, value=b_label)
        bs.font = Font("Calibri", size=9, bold=True, color=b_color)
        bs.fill = _fill(alt_bg); bs.alignment = _align("center"); bs.border = _bdr()

        # Col C: SOH% (static from computation, not formula)
        soh_pct = round(sv * 100, 2) if not bess_zero else 0.0
        soh_color = (C_GREEN if alive else C_RED) if not bess_zero else "888888"
        _wc(3, val=soh_pct, fmt="0.0", color=soh_color, bold=True)

        # Col D: PV Saving = $B$4 × (1+$D$4/100)^(yr-1) × (1-$G$4/100)^(yr-1)
        _wc(4, formula=f"=$B$4*(1+$D$4/100)^(A{r}-1)*(1-$G$4/100)^(A{r}-1)")

        # Col E: BESS Saving (0 when no BESS installed)
        if bess_zero:
            _wc(5, val=0)
        else:
            # IF(SOH% >= 60, BaseBESS × esc_factor × SOH/100, 0)
            _wc(5, formula=f"=IF(C{r}>=60,$C$4*(1+$D$4/100)^(A{r}-1)*C{r}/100,0)")

        # Col F: Total Saving = D + E
        _wc(6, formula=f"=D{r}+E{r}", bold=True)

        # Col G: PV O&M = PV_kWp × rate × (1 + esc×0.5)^(yr-1)
        _wc(7, formula=f"=$I$4*$K$4*(1+$D$4/200)^(A{r}-1)")

        # Col H: BESS O&M (0 if no BESS or past EoL)
        if bess_zero or not alive:
            _wc(8, val=0, bg=("FFEEEE" if (not bess_zero and not alive) else alt_bg))
        else:
            _wc(8, formula=f"=$J$4*$L$4*(1+$D$4/200)^(A{r}-1)")

        # Col I: EBITDA = F - G - H
        _wc(9, formula=f"=F{r}-G{r}-H{r}", bold=True)

        # Col J: Section 12B Depreciation (equipment portion only — 60% of CAPEX)
        # 40% services are excluded from 12B / 11E accelerated depreciation
        if dep_pct > 0:
            _wc(10, formula=f"=$H$4*0.60*{dep_pct}")
        else:
            _wc(10, val=0)

        # Col K: Tax Shield = Dep × Tax%  (realized immediately — BTM at profitable company)
        if dep_pct > 0:
            _wc(11, formula=f"=J{r}*$F$4/100")
        else:
            _wc(11, val=0)

        # Col L: EBIT = EBITDA - Depreciation
        _wc(12, formula=f"=I{r}-J{r}")

        # Col M: Net Tax = EBIT × Tax%  (negative = tax saving → positive cash benefit)
        # BTM model: company has taxable income; dep deducted against main business income.
        # No MAX(0,...) — negative tax means company writes a smaller cheque to SARS.
        _wc(13, formula=f"=L{r}*$F$4/100")

        # Col N: Net Cash Flow = EBITDA - Net Tax  (dep is non-cash; neg tax = real cash inflow)
        ncf_val = float(fin_df["Net Cash Flow NCF (ZAR)"].iloc[yr - 1])
        _wc(14, formula=f"=I{r}-M{r}", bold=True,
            color=C_GREEN if ncf_val > 0 else C_RED)

        # Col O: Discounted CF = NCF / (1 + disc%)^yr
        _wc(15, formula=f"=N{r}/(1+$E$4/100)^A{r}")

        # Col P: Cumulative CF
        cum_prev = f"P{r - 1}" if yr > 1 else f"P{r0}"
        _wc(16, formula=f"={cum_prev}+N{r}", bold=True)

        # Col Q: Net Profit = EBIT - Cash Tax
        _wc(17, formula=f"=L{r}-M{r}")

        # Col R: Assessed Loss B/F — always 0 for BTM model
        # Company absorbs 12B deduction against its main business income immediately.
        _wc(18, val=0, fmt="#,##0", color="888888")

        ws3.row_dimensions[r].height = 16

    # ── Totals row (row 28) ───────────────────────────────────
    r_tot = r0 + ANALYSIS_YEARS + 1   # = 28
    ws3.merge_cells(start_row=r_tot, start_column=1, end_row=r_tot, end_column=3)
    tc_lbl = ws3.cell(row=r_tot, column=1, value="20Y Total" )
    tc_lbl.font = Font("Calibri", size=10, bold=True, color="FFFFFF")
    tc_lbl.fill = _fill(C_NAVY); tc_lbl.alignment = _align("center"); tc_lbl.border = _bdr()
    for ci_s, cl_s in {4:"D",5:"E",6:"F",7:"G",8:"H",9:"I",
                        10:"J",11:"K",12:"L",13:"M",14:"N",15:"O",17:"Q"}.items():
        ts = ws3.cell(row=r_tot, column=ci_s,
                      value=f"=SUM({cl_s}8:{cl_s}{r_tot-1})")
        ts.font = Font("Calibri", size=10, bold=True, color="FFFFFF")
        ts.fill = _fill(C_NAVY); ts.alignment = _align("center")
        ts.number_format = "#,##0"; ts.border = _bdr()
    tc_cum = ws3.cell(row=r_tot, column=16, value=f"=P{r_tot-1}")
    tc_cum.font = Font("Calibri", size=10, bold=True, color="FFFFFF")
    tc_cum.fill = _fill(C_NAVY); tc_cum.alignment = _align("center")
    tc_cum.number_format = "#,##0"; tc_cum.border = _bdr()
    ws3.row_dimensions[r_tot].height = 20

    # ════════════════════════════════════════════════════════════
    # SHEET 4 — Monthly
    # ════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Monthly Summary")
    ws4.sheet_view.showGridLines = False
    for ci, w in enumerate([2, 16, 14, 14, 14, 14, 14, 14, 14], 1):
        ws4.column_dimensions[get_column_letter(ci)].width = w

    ws4.merge_cells("B1:I1")
    _sheet4_title = "Monthly Dispatch Summary — Year 1"
    c = ws4.cell(row=1, column=2, value=_sheet4_title)
    c.font = Font("Calibri", size=12, bold=True, color="FFFFFF")
    c.fill = _fill(C_NAVY); c.alignment = _align("center")
    ws4.row_dimensions[1].height = 26

    _hdrs4 = ["Month", "PV Gen\nkWh", "PV→Load\nkWh", "BESS Disch.\nkWh",
              "Grid Buy\nkWh", "PV Saving\nZAR", "BESS Saving\nZAR", "Total Saving\nZAR"]
    for ci, hdr in enumerate(_hdrs4, 2):
        hc = ws4.cell(row=2, column=ci, value=hdr)
        hc.font = Font("Calibri", size=9, bold=True, color="FFFFFF")
        hc.fill = _fill(C_DARK)
        hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        hc.border = _bdr()
    ws4.row_dimensions[2].height = 28

    mon_grp = h_df.groupby("month").agg(
        pv_gen=("pv_gen_kWh", "sum"),
        pv_to_load=("pv_to_load_kWh", "sum"),
        discharge=("discharge_kWh", "sum"),
        grid_buy=("grid_buy_kWh", "sum"),
        pv_save=("pv_saving_ZAR", "sum"),
        bess_save=("bess_net_saving_ZAR", "sum"),
        total_save=("net_saving_ZAR", "sum"),
    ).reset_index().sort_values("month")

    mon_names_full = {
        1:"January",2:"February",3:"March",4:"April",5:"May",6:"June",
        7:"July",8:"August",9:"September",10:"October",11:"November",12:"December"
    }
    for _, mrow in mon_grp.iterrows():
        mr = int(mrow["month"]) + 2      # rows 3-14
        alt_m = C_ALT if int(mrow["month"]) % 2 == 0 else "FFFFFF"
        mc = ws4.cell(row=mr, column=2, value=mon_names_full.get(int(mrow["month"]), ""))
        mc.font = _font(bold=True, sz=10); mc.fill = _fill(C_LTBLUE)
        mc.alignment = _align(); mc.border = _bdr()
        for ci, key in enumerate(["pv_gen","pv_to_load","discharge","grid_buy",
                                   "pv_save","bess_save","total_save"], 3):
            vc = ws4.cell(row=mr, column=ci, value=round(float(mrow[key]), 0))
            vc.fill = _fill(alt_m); vc.alignment = _align("center")
            vc.number_format = "#,##0"; vc.border = _bdr(); vc.font = _font(sz=10)
        ws4.row_dimensions[mr].height = 17

    # Totals row 15
    mtt = ws4.cell(row=15, column=2, value="Annual Total" )
    mtt.font = Font("Calibri", size=10, bold=True, color="FFFFFF")
    mtt.fill = _fill(C_NAVY); mtt.alignment = _align("center"); mtt.border = _bdr()
    for ci, cl in enumerate(["C","D","E","F","G","H","I"], 3):
        tc = ws4.cell(row=15, column=ci, value=f"=SUM({cl}3:{cl}14)")
        tc.font = Font("Calibri", size=10, bold=True, color="FFFFFF")
        tc.fill = _fill(C_NAVY); tc.alignment = _align("center")
        tc.number_format = "#,##0"; tc.border = _bdr()
    ws4.row_dimensions[15].height = 20

    # ════════════════════════════════════════════════════════════
    # SHEET 5 — Charts
    # Sheet3 data: header row 6, Year0 row 7, Years 1-20 rows 8-27, totals row 28
    # Chart data range: rows 6-27 (includes header for title), categories rows 7-27
    # BESS charts skipped when bess_zero=True
    # ════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("Charts")
    ws5.sheet_view.showGridLines = False
    ws5.column_dimensions["A"].width = 2

    ws5.merge_cells("A1:R1")
    bess_chart_note = "" if not bess_zero else "  (BESS=0 — BESS charts omitted)"
    _ws5_title = f"Charts — Linked to Sheet3 & Sheet4 · Auto-updates on parameter change{bess_chart_note}"
    c = ws5.cell(row=1, column=1, value=_ws5_title)
    c.font = Font("Calibri", size=11, bold=True, color="FFFFFF")
    c.fill = _fill(C_NAVY); c.alignment = _align("center")
    ws5.row_dimensions[1].height = 24

    # Shared category reference: col A rows 7-32 (year 0..25)
    _cat = Reference(ws3, min_col=1, max_col=1, min_row=7, max_row=r0 + ANALYSIS_YEARS)
    # Shared category reference for years 1-20 only
    _cat_yr1 = Reference(ws3, min_col=1, max_col=1, min_row=8, max_row=r0 + ANALYSIS_YEARS)

    # Chart 1: Cumulative CF line (col P=16, rows 6-27)
    ch1 = LineChart()
    ch1.title = "Cumulative Cash Flow — 20 Years (ZAR)"
    ch1.style = 10; ch1.width = 17; ch1.height = 11
    ch1.y_axis.title = "ZAR"; ch1.x_axis.title = "Year"
    ch1.add_data(Reference(ws3, min_col=16, max_col=16, min_row=6, max_row=r0 + ANALYSIS_YEARS),
                 titles_from_data=True)
    ch1.set_categories(_cat)
    ws5.add_chart(ch1, "B3")

    # Chart 2: Annual NCF bar (col N=14, rows 6-27)
    ch2 = BarChart()
    ch2.type = "col"
    ch2.title = "Annual Net Cash Flow (ZAR)"
    ch2.style = 10; ch2.width = 17; ch2.height = 11
    ch2.y_axis.title = "ZAR"; ch2.x_axis.title = "Year"
    ch2.add_data(Reference(ws3, min_col=14, max_col=14, min_row=6, max_row=r0 + ANALYSIS_YEARS),
                 titles_from_data=True)
    ch2.set_categories(_cat)
    ws5.add_chart(ch2, "L3")

    # Chart 3: BESS SOH degradation — only when BESS installed
    chart3_anchor = "B24"
    if not bess_zero:
        ch3 = LineChart()
        ch3.title = "Battery SOH Degradation (%)"
        ch3.style = 10; ch3.width = 17; ch3.height = 11
        ch3.y_axis.title = "SOH (%)"; ch3.x_axis.title = "Year"
        ch3.add_data(Reference(ws3, min_col=3, max_col=3, min_row=6, max_row=r0 + ANALYSIS_YEARS),
                     titles_from_data=True)
        ch3.set_categories(_cat_yr1)
        ws5.add_chart(ch3, chart3_anchor)

    # Chart 4: Annual savings breakdown
    ch4 = BarChart()
    ch4.type = "col"
    if not bess_zero:
        ch4.grouping = "stacked"
        ch4.title = "Annual Savings Breakdown (ZAR)"
        # D=PV savings, E=BESS savings — both series
        ch4.add_data(Reference(ws3, min_col=4, max_col=5, min_row=6, max_row=r0 + ANALYSIS_YEARS),
                     titles_from_data=True)
    else:
        ch4.title = "Annual PV Savings (ZAR)"
        # Only D=PV savings when no BESS
        ch4.add_data(Reference(ws3, min_col=4, max_col=4, min_row=6, max_row=r0 + ANALYSIS_YEARS),
                     titles_from_data=True)
    ch4.style = 10; ch4.width = 17; ch4.height = 11
    ch4.y_axis.title = "ZAR"; ch4.x_axis.title = "Year"
    ch4.set_categories(_cat_yr1)
    ws5.add_chart(ch4, "L24")

    # Chart 5: Monthly PV generation bar (ws4 col C=3, rows 2-14)
    ch5 = BarChart()
    ch5.type = "col"
    ch5.title = "Monthly PV Generation (kWh)"
    ch5.style = 10; ch5.width = 17; ch5.height = 11
    ch5.y_axis.title = "kWh"; ch5.x_axis.title = "Month"
    ch5.add_data(Reference(ws4, min_col=3, max_col=3, min_row=2, max_row=14),
                 titles_from_data=True)
    ch5.set_categories(Reference(ws4, min_col=2, max_col=2, min_row=3, max_row=14))
    chart5_anchor = "B45" if not bess_zero else "B24"
    ws5.add_chart(ch5, chart5_anchor)

    # ════════════════════════════════════════════════════════════
    # SHEET 6 — Raw Data (8760h)
    # ════════════════════════════════════════════════════════════
    ws6 = wb.create_sheet("Raw Data")
    ws6.freeze_panes = "A2"
    h_cols = list(h_df.columns)
    ws6.auto_filter.ref = f"A1:{get_column_letter(len(h_cols))}1"
    for ci, hdr in enumerate(h_cols, 1):
        hc = ws6.cell(row=1, column=ci, value=hdr)
        hc.font = Font("Calibri", size=9, bold=True, color="FFFFFF")
        hc.fill = _fill(C_NAVY); hc.alignment = _align("center"); hc.border = _bdr()
        ws6.column_dimensions[get_column_letter(ci)].width = max(10, len(str(hdr)) + 2)
    ws6.row_dimensions[1].height = 18
    for record in h_df.round(3).itertuples(index=False):
        ws6.append(list(record))

    # ── Serialize ─────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────
# App header — back | title | user info
# ─────────────────────────────────────────────────────────────
_main_title = "BTM PV+BESS Financial Modelling System"
_back_col, _hdr_col, _user_col = st.columns([1, 5, 4], gap="small")

with _back_col:
    if st.button("◀", key="btm_back_to_scenarios",
                 help="Back to project list",
                 use_container_width=True):
        st.session_state.pop("_scenario", None)
        st.rerun()

with _hdr_col:
    st.markdown(f"""
<div class="main-header">
    <div style="font-size:1.8rem;line-height:1">⚡</div>
    <div>
        <div class="main-title">{_main_title}</div>
        <div class="sub-title">MEGAFLEX / MINIFLEX 2025/26 &nbsp;·&nbsp; 8760H DISPATCH &nbsp;·&nbsp; SECTION 12B &nbsp;·&nbsp; SA BTM FINANCIAL MODEL</div>
    </div>
</div>
""", unsafe_allow_html=True)

with _user_col:
    _u_hdr = get_current_user()
    if _u_hdr:
        _tbadge = {"free": "🆓", "pro": "🔵", "admin": "🔴"}.get(
            _u_hdr.get("tier", "free"), "🆓")
        _uname  = _u_hdr.get("full_name") or _u_hdr.get("email", "User")
        _uname_short = (_uname[:16] + "…") if len(_uname) > 18 else _uname

        # ── Business-mode badge ──────────────────────────────────────
        _scen_now = st.session_state.get("_scenario", "btm")
        _lm_badge = st.session_state.get("_light_mode", False)
        _scen_cfg = {
            "btm":      ("⚡", "BTM",
                         "#059669" if _lm_badge else "#00E5A0",
                         "rgba(0,168,112,0.12)" if _lm_badge else "#0d2b1e"),
            "wheeling": ("🔄", "Wheeling",
                         "#0E7490" if _lm_badge else "#4ECDC4",
                         "rgba(78,205,196,0.12)" if _lm_badge else "#0d2424"),
        }
        _s_icon, _s_name, _s_clr, _s_bg = _scen_cfg.get(
            _scen_now, ("⚡", "BTM",
                        "#059669" if _lm_badge else "#00E5A0",
                        "rgba(0,168,112,0.12)" if _lm_badge else "#0d2b1e"))

        # ── Light / dark mode toggle ─────────────────────────────────
        _light_mode = st.session_state.get("_light_mode", False)
        _theme_icon = "☀️" if _light_mode else "🌙"
        _theme_tip  = "Switch to Dark mode" if _light_mode else "Switch to Light mode"

        # ── Layout: [🌙] [badge + name  →right-aligned] [⏻ 退出]
        # Logout gets 2 units so it's visually part of the same row
        _tc, _mc, _lc = st.columns([1, 7, 2], gap="small")

        with _tc:
            if st.button(_theme_icon, key="hdr_theme_btn",
                         help=_theme_tip, use_container_width=True):
                st.session_state["_light_mode"] = not _light_mode
                st.rerun()

        with _mc:
            # justify-content:flex-end pushes badge+name to the right edge
            st.markdown(
                f"<div style='padding-top:4px;display:flex;"
                f"align-items:center;justify-content:flex-end;"
                f"gap:10px;overflow:hidden'>"
                f"<span style='background:{_s_bg};color:{_s_clr};"
                f"border:1px solid {_s_clr}44;border-radius:4px;"
                f"padding:2px 8px;font-size:0.72rem;"
                f"font-family:IBM Plex Mono,monospace;letter-spacing:0.05em;"
                f"white-space:nowrap;flex-shrink:0'>{_s_icon} {_s_name}</span>"
                f"<span style='font-size:0.80rem;color:var(--text-dim);"
                f"white-space:nowrap;overflow:hidden;text-overflow:ellipsis'>"
                f"👤 <b style='color:var(--text-main)'>{_uname_short}</b>"
                f" {_tbadge}</span></div>",
                unsafe_allow_html=True,
            )

        with _lc:
            if st.button("⏻ 退出", key="logout_btn_hdr", help="Sign out",
                         use_container_width=True):
                logout()

# ─────────────────────────────────────────────────────────────
# Project bar — full-width, above column split
# ─────────────────────────────────────────────────────────────
render_project_bar()
_hr_bc = "#CBD5E0" if st.session_state.get("_light_mode", False) else "#1e2a3a"
st.markdown(f"<hr style='margin:4px 0 8px;border-color:{_hr_bc}'>",
            unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────
# Main layout (7 left content + 3 right params)
# Main layout: content (left 7) + params panel (right 3)
# ─────────────────────────────────────────────────────────────
col_content, col_params = st.columns([7, 3], gap="large")

# ══════════════════════════════════════════════════════════════
# RIGHT SIDE — Parameter Panel
# ══════════════════════════════════════════════════════════════
with col_params:
    # Native Streamlit scrollable container — fixed height, own scrollbar
    _scroll = st.container(height=900, border=False)

with _scroll:

    # ══════════════════════════════════════════════════════════
    # 1. Site Location
    # ══════════════════════════════════════════════════════════
    with st.expander("📍 Site Location", expanded=True):
        if MAP_AVAILABLE:
            m = folium.Map(
                location=[st.session_state.lat, st.session_state.lon],
                zoom_start=7,
                tiles="https://mt1.google.com/vt/lyrs=s&x={x}&y={y}&z={z}",
                attr="Google Satellite",
            )
            folium.Marker(
                location=[st.session_state.lat, st.session_state.lon],
                popup=f"📍 {st.session_state.lat:.4f}, {st.session_state.lon:.4f}",
                icon=folium.Icon(color="red", icon="map-pin", prefix="fa"),
            ).add_to(m)
            map_data = st_folium(m, width=None, height=200, key="site_map")

            # Anti-crash: use temp vars before committing to session state
            if map_data and map_data.get("last_clicked"):
                _tmp_lat = map_data["last_clicked"]["lat"]
                _tmp_lon = map_data["last_clicked"]["lng"]
                if (abs(_tmp_lat - st.session_state.lat) > 1e-5 or
                        abs(_tmp_lon - st.session_state.lon) > 1e-5):
                    st.session_state.lat = _tmp_lat
                    st.session_state.lon = _tmp_lon
                    # Sync widget-state keys so number_inputs display the clicked values
                    st.session_state["_lat_in"] = _tmp_lat
                    st.session_state["_lon_in"] = _tmp_lon
                    st.session_state.tilt = round(abs(_tmp_lat), 1)
                    # =180°=0° / SH north-facing=180°, NH south-facing=0°
                    st.session_state.azimuth = 180.0 if _tmp_lat < 0 else 0.0
                    st.rerun()
        else:
            st.info("Install folium + streamlit-folium to enable the map" )

        ca, cb = st.columns(2)
        with ca:
            new_lat = st.number_input("Latitude", value=st.session_state.lat,
                                      min_value=-90.0, max_value=90.0,
                                      format="%.4f", key="_lat_in")
        with cb:
            new_lon = st.number_input("Longitude", value=st.session_state.lon,
                                      min_value=-180.0, max_value=180.0,
                                      format="%.4f", key="_lon_in")

        if abs(new_lat - st.session_state.lat) > 1e-5:
            st.session_state.lat = new_lat
            st.session_state.tilt = round(abs(new_lat), 1)
            st.session_state.azimuth = 180.0 if new_lat < 0 else 0.0
        if abs(new_lon - st.session_state.lon) > 1e-5:
            st.session_state.lon = new_lon

        # PVGIS
        check_auto_pvgis()

        # PVGIS
        ok = "success-box" if "✓" in st.session_state.pvgis_status else "warning-box"
        st.markdown(f'<div class="{ok}" style="font-size:0.7rem">🌤 {st.session_state.pvgis_status}</div>',
                    unsafe_allow_html=True)
        if st.session_state.annual_pv_kwh:
            eq_h = st.session_state.annual_pv_kwh / max(st.session_state.pv_kwp, 1)
            st.markdown(
                f'<div class="derived-value">☀️ {_fmw(st.session_state.annual_pv_kwh,"kWh/yr")}'
                f' ({eq_h:.0f}h equiv)</div>',
                unsafe_allow_html=True
            )

    # ══════════════════════════════════════════════════════════
    # 2. PV System
    # ══════════════════════════════════════════════════════════
    with st.expander("☀️ PV System", expanded=False):
        pv_kwp = st.number_input("Capacity (kWp)", value=st.session_state.pv_kwp,
                                  min_value=0.0, step=50.0)
        st.session_state.pv_kwp = pv_kwp
        pv_dis = pv_kwp == 0.0

        pv_loss = st.number_input("System Loss (%)", value=st.session_state.pv_loss,
                                   min_value=0.0, max_value=50.0, step=0.5,
                                   disabled=pv_dis)
        _tilt_lbl = f"Tilt (°) [optimal≈{abs(st.session_state.lat):.0f}°]"
        tilt = st.number_input(
            _tilt_lbl,
            value=st.session_state.tilt,
            min_value=0.0, max_value=90.0, step=1.0, disabled=pv_dis
        )
        azimuth = st.number_input(
            "Azimuth (°)",
            value=st.session_state.azimuth,
            min_value=-180.0, max_value=180.0, step=5.0, disabled=pv_dis
        )
        pv_deg = st.number_input("Annual PV Degradation (%)",
                                  value=st.session_state.pv_degradation,
                                  min_value=0.0, max_value=5.0, step=0.1,
                                  disabled=pv_dis)
        if not pv_dis:
            st.session_state.pv_loss       = pv_loss
            st.session_state.tilt          = tilt
            st.session_state.azimuth       = azimuth
            st.session_state.pv_degradation = pv_deg

        # PV PVGIS PV
        # Re-run PVGIS if any PV parameter changed (lat/lon handled above; this covers the rest)
        check_auto_pvgis()
        _ok2 = "success-box" if "✓" in st.session_state.pvgis_status else "warning-box"
        st.markdown(
            f'<div class="{_ok2}" style="font-size:0.7rem">🌤 {st.session_state.pvgis_status}</div>',
            unsafe_allow_html=True,
        )
        if st.session_state.annual_pv_kwh and not pv_dis:
            _eq_h = st.session_state.annual_pv_kwh / max(st.session_state.pv_kwp, 1)
            st.markdown(
                f'<div class="derived-value">☀️ {_fmw(st.session_state.annual_pv_kwh,"kWh/yr")}'
                f' ({_eq_h:.0f} h equiv)</div>',
                unsafe_allow_html=True,
            )

    # ══════════════════════════════════════════════════════════
    # 3. BESS System
    # ══════════════════════════════════════════════════════════
    with st.expander("🔋 BESS System", expanded=False):
        bess_kwh = st.number_input("Capacity (kWh)", value=st.session_state.bess_kwh,
                                    min_value=0.0, step=50.0)
        st.session_state.bess_kwh = bess_kwh

        # C/ C-rate selector (reference only)
        c_rate_label = st.radio(
            "C-rate",
            options=list(C_RATE_OPTIONS.keys()),
            index=list(C_RATE_OPTIONS.keys()).index(st.session_state.c_rate_label),
            horizontal=True,
        )
        st.session_state.c_rate_label = c_rate_label
        c_rate_val = C_RATE_OPTIONS[c_rate_label]
        bess_kw_max = bess_kwh * c_rate_val # = C ×
        _bess_pw_lbl = ("Max Power" )
        st.markdown(
            f'<div class="derived-value">⚡ {_bess_pw_lbl}: <b>{_fmw(bess_kw_max,"kW")}</b> '
            f'({c_rate_label}) — dispatch engine auto-optimises output</div>',
            unsafe_allow_html=True
        )

        rte = st.number_input("Round-Trip Efficiency RTE (%)", value=st.session_state.rte,
                               min_value=50.0, max_value=100.0, step=0.5)
        dod = st.number_input("Depth of Discharge DoD (%)", value=st.session_state.dod,
                               min_value=50.0, max_value=100.0, step=1.0)  # max=100% allowed
        bess_cycles = st.number_input("Nominal Cycle Life", value=int(st.session_state.bess_cycles),
                                       min_value=1000, max_value=20000, step=500)

        st.session_state.rte         = rte
        st.session_state.dod         = dod
        st.session_state.bess_cycles = bess_cycles

    # ══════════════════════════════════════════════════════════
    # 4. Site Load Profile
    # ══════════════════════════════════════════════════════════
    with st.expander("🏭 Load Profile", expanded=False):

        _lp_mode = st.radio(
            "Load input mode",
            ["Manual (3-period average)", "Upload CSV / Excel (8760 h)"],
            key="load_profile_mode",
            horizontal=True,
            label_visibility="collapsed",
        )
        _csv_mode = _lp_mode.startswith("Upload")

        if _csv_mode:
            # ── CSV / Excel upload path ──────────────────────────────
            st.caption(
                "Upload a file with **8 760 rows** (1-h intervals) or **17 520 rows** "
                "(30-min intervals, auto-resampled). One numeric column in **kW**."
            )
            _uploaded = st.file_uploader(
                "Load profile file",
                type=["csv", "txt", "xlsx", "xls"],
                key="load_profile_uploader",
                label_visibility="collapsed",
                help="Comma / semicolon / tab delimited · optional header row · kW positive values",
            )

            if _uploaded is not None:
                _arr, _msg = parse_load_csv(_uploaded)
                if _arr is not None:
                    st.session_state["load_profile_8760"] = _arr
                    st.session_state["load_profile_name"] = _uploaded.name
                    # ── Auto-derive 3-period averages from the real profile ──
                    _pk_mask  = np.array([(g % 24 in range(7, 10)) or (g % 24 in range(17, 20))
                                          for g in range(8760)])
                    _op_mask  = np.array([g % 24 < 6 or g % 24 >= 22 for g in range(8760)])
                    _std_mask = ~_pk_mask & ~_op_mask
                    st.session_state.load_peak_kw    = float(np.mean(_arr[_pk_mask]))
                    st.session_state.load_std_kw     = float(np.mean(_arr[_std_mask]))
                    st.session_state.load_offpeak_kw = float(np.mean(_arr[_op_mask]))
                    st.success(_msg)
                else:
                    st.error(_msg)

            _active_arr = st.session_state.get("load_profile_8760")
            if _active_arr is not None:
                _fname = st.session_state.get("load_profile_name", "profile")
                st.markdown(
                    f'<div class="derived-value">📋 Active: <b>{_fname}</b> &nbsp;·&nbsp; '
                    f'Annual {_active_arr.sum()/1000:.1f} MWh &nbsp;·&nbsp; '
                    f'Peak {_active_arr.max():.0f} kW &nbsp;·&nbsp; '
                    f'Avg {_active_arr.mean():.0f} kW</div>',
                    unsafe_allow_html=True,
                )
                # ── First-week preview chart ─────────────────────────
                _prev = _active_arr[:168].tolist()
                _pfig = go.Figure(go.Scatter(
                    x=list(range(168)), y=_prev,
                    fill="tozeroy", fillcolor="rgba(0,229,160,0.07)",
                    line=dict(color="#00E5A0", width=1.2),
                ))
                _pb, _plb, _fc, _gc, _lb = _plt()
                _pfig.update_layout(
                    height=130, margin=dict(l=0, r=0, t=4, b=24),
                    paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                    font=dict(color=_fc, size=9), showlegend=False,
                    xaxis=dict(gridcolor=_gc, title="Hour (first 7 days)",
                               title_font_size=9, showgrid=True, tickfont_size=8),
                    yaxis=dict(gridcolor=_gc, title="kW",
                               title_font_size=9, showgrid=True, tickfont_size=8),
                )
                st.plotly_chart(_pfig, use_container_width=True,
                                config={"displayModeBar": False})
                # Period-average summary (auto-derived)
                st.markdown(
                    f'<div class="param-label">Auto-derived period averages '
                    f'(used for BESS charging strategy):</div>',
                    unsafe_allow_html=True,
                )
                _pa1, _pa2, _pa3 = st.columns(3)
                _pa1.metric("Peak avg",    _fmw(st.session_state.load_peak_kw,    "kW"))
                _pa2.metric("Std avg",     _fmw(st.session_state.load_std_kw,     "kW"))
                _pa3.metric("Off-pk avg",  _fmw(st.session_state.load_offpeak_kw, "kW"))
                if st.button("🗑️ Clear profile — switch to manual",
                             key="clear_load_profile_btn"):
                    st.session_state["load_profile_8760"] = None
                    st.session_state["load_profile_name"] = ""
                    st.rerun()
            else:
                st.info("📂 No profile loaded yet — upload a file above.")

        else:
            # ── Manual 3-period path (existing logic) ────────────────
            # Clear any stored CSV profile so dispatch uses manual values
            st.session_state["load_profile_8760"] = None

            st.markdown('<div class="param-label">Peak hours (07–09 h · 17–20 h)</div>',
                        unsafe_allow_html=True)
            load_peak = st.number_input("Peak Load (kW)",
                                        value=st.session_state.load_peak_kw,
                                        min_value=0.0, step=10.0,
                                        label_visibility="collapsed")
            st.markdown('<div class="param-label">Standard hours (06–07 · 10–17 · 20–22 h)</div>',
                        unsafe_allow_html=True)
            load_std = st.number_input("Standard Load (kW)",
                                       value=st.session_state.load_std_kw,
                                       min_value=0.0, step=10.0,
                                       label_visibility="collapsed")
            st.markdown('<div class="param-label">Off-peak hours (22–06 h)</div>',
                        unsafe_allow_html=True)
            load_offpeak = st.number_input("Off-Peak Load (kW)",
                                           value=st.session_state.load_offpeak_kw,
                                           min_value=0.0, step=5.0,
                                           label_visibility="collapsed")

            st.session_state.load_peak_kw    = load_peak
            st.session_state.load_std_kw     = load_std
            st.session_state.load_offpeak_kw = load_offpeak

            # Estimated daily energy: peak 5 h, std 11 h, off-peak 8 h
            est_daily = load_peak * 5 + load_std * 11 + load_offpeak * 8
            st.markdown(
                f'<div class="derived-value">📊 Est. daily load ≈ {est_daily:.0f} kWh/day'
                f' · {est_daily*365/1000:.0f} MWh/yr</div>',
                unsafe_allow_html=True,
            )

    # ══════════════════════════════════════════════════════════
    # 5. Eskom 2025/26 Tariffs
    # ══════════════════════════════════════════════════════════
    with st.expander("💡 Tariffs (ZAR/kWh)", expanded=False):

        # Tariff mode selector (auto-populates rates from TARIFF_DB)
        all_modes = list(TARIFF_DB.keys())
        cur_mode  = st.session_state.get("tariff_mode", "Megaflex ≤300km <500V")
        if cur_mode not in all_modes:
            cur_mode = all_modes[0]
        tariff_mode = st.selectbox(
            "Tariff Mode",
            all_modes,
            index=all_modes.index(cur_mode),
            key="tariff_mode_sel",
        )

        # Auto-populate rates when mode changes; do NOT overwrite on same mode
        prev_mode = st.session_state.get("_prev_tariff_mode")
        if tariff_mode != prev_mode:
            st.session_state.tariff_mode = tariff_mode
            if tariff_mode != "Custom (manual)" and TARIFF_DB.get(tariff_mode):
                w_pk, w_std, w_op, s_pk, s_std, s_op = TARIFF_DB[tariff_mode]
                st.session_state.w_morning_peak = w_pk
                st.session_state.w_evening_peak = w_pk
                st.session_state.w_standard     = w_std
                st.session_state.w_off_peak     = w_op
                st.session_state.s_morning_peak = s_pk
                st.session_state.s_evening_peak = s_pk
                st.session_state.s_standard     = s_std
                st.session_state.s_off_peak     = s_op
            st.session_state._prev_tariff_mode = tariff_mode
            st.rerun()

        if tariff_mode == "PPA Custom (flat rate)":
            st.caption("PPA: enter one flat rate applied to all periods")
            ppa_rate = st.number_input(
                "PPA Rate (ZAR/kWh)",
                value=st.session_state.get("ppa_rate", 1.20),
                min_value=0.10, max_value=5.00, step=0.05, format="%.2f",
            )
            st.session_state.ppa_rate = ppa_rate
            # PPA
            for key in ("w_morning_peak","w_evening_peak","w_standard","w_off_peak",
                        "s_morning_peak","s_evening_peak","s_standard","s_off_peak"):
                st.session_state[key] = ppa_rate
        elif tariff_mode == "Custom (manual)":
            st.caption("Custom: edit all rates below")
        else:
            st.caption("Select mode to auto-fill; manual edits override")

        st.markdown(f'<div class="tariff-winter"><b style="color:#4ECDC4">❄️ High Season (Jun-Aug)</b></div>',
                    unsafe_allow_html=True)
        cw1, cw2 = st.columns(2)
        with cw1:
            st.session_state.w_morning_peak = st.number_input(
                "Morning Peak (W)", value=st.session_state.w_morning_peak,
                min_value=0.1, step=0.05, format="%.2f")
            st.session_state.w_standard = st.number_input(
                "Standard (W)", value=st.session_state.w_standard,
                min_value=0.1, step=0.05, format="%.2f")
        with cw2:
            st.session_state.w_evening_peak = st.number_input(
                "Evening Peak (W)", value=st.session_state.w_evening_peak,
                min_value=0.1, step=0.05, format="%.2f")
            st.session_state.w_off_peak = st.number_input(
                "Off-Peak (W)", value=st.session_state.w_off_peak,
                min_value=0.05, step=0.05, format="%.2f")

        st.markdown(f'<div class="tariff-summer"><b style="color:#F6C90E">☀️ Low Season (Sep-May)</b></div>',
                    unsafe_allow_html=True)
        cs1, cs2 = st.columns(2)
        with cs1:
            st.session_state.s_morning_peak = st.number_input(
                "Morn Peak (S)", value=st.session_state.s_morning_peak,
                min_value=0.1, step=0.05, format="%.2f")
            st.session_state.s_standard = st.number_input(
                "Standard (S)", value=st.session_state.s_standard,
                min_value=0.1, step=0.05, format="%.2f")
        with cs2:
            st.session_state.s_evening_peak = st.number_input(
                "Eve Peak (S)", value=st.session_state.s_evening_peak,
                min_value=0.1, step=0.05, format="%.2f")
            st.session_state.s_off_peak = st.number_input(
                "Off-Peak (S)", value=st.session_state.s_off_peak,
                min_value=0.05, step=0.05, format="%.2f")

        # Peak-valley spread display
        w_spread = st.session_state.w_morning_peak - st.session_state.w_off_peak
        s_spread = st.session_state.s_morning_peak - st.session_state.s_off_peak
        _spread_lbl = "Peak-Valley Spread"
        st.markdown(
            f'<div class="derived-value">'
            f'❄️ {_spread_lbl} {w_spread:.2f} | ☀️ {_spread_lbl} {s_spread:.2f} ZAR/kWh'
            f'</div>',
            unsafe_allow_html=True
        )

    # ══════════════════════════════════════════════════════════
    # 6. Financial Parameters
    # ══════════════════════════════════════════════════════════
    with st.expander("💰 Financial Parameters", expanded=False):

        # Forex — currency exchange rate
        fx_col1, fx_col2 = st.columns([3, 1])
        with fx_col1:
            forex = st.number_input("USD/ZAR Exchange Rate", value=st.session_state.forex_usd_zar,
                                     min_value=1.0, step=0.1, format="%.2f")
            st.session_state.forex_usd_zar = forex
        with fx_col2:
            if st.button("🔄", help=("Refresh live exchange rate"),
                          use_container_width=True):
                live = fetch_forex_rate()
                st.session_state.forex_usd_zar = live
                st.toast(f"{'Exchange rate updated'}: 1 USD = {live:.2f} ZAR")
                st.rerun()

        # PV & BESS USD unit costs
        pv_usd = st.number_input("PV Unit Cost (USD/W)", value=st.session_state.pv_usd_per_w,
                                  min_value=0.1, step=0.05, format="%.3f",
                                  disabled=pv_kwp == 0)
        bess_usd = st.number_input("BESS Unit Cost (USD/Wh)", value=st.session_state.bess_usd_per_wh,
                                    min_value=0.05, step=0.01, format="%.3f")
        st.session_state.pv_usd_per_w    = pv_usd
        st.session_state.bess_usd_per_wh = bess_usd

        # Show ZAR equivalents
        pv_zar_kwp, bess_zar_kwh = get_capex_zar()
        st.markdown(
            f'<div class="derived-value">'
            f'PV: {pv_zar_kwp:,.0f} ZAR/kWp | BESS: {bess_zar_kwh:,.0f} ZAR/kWh'
            f'</div>',
            unsafe_allow_html=True
        )

        # O&M inputs
        pv_opex = st.number_input("PV O&M (ZAR/kWp/yr)",
                                   value=st.session_state.pv_opex_per_kwp,
                                   min_value=0.0, step=5.0, disabled=pv_kwp == 0)
        bess_opex = st.number_input("BESS O&M (ZAR/kWh/yr)",
                                     value=st.session_state.bess_opex_per_kwh,
                                     min_value=0.0, step=1.0)
        st.session_state.pv_opex_per_kwp   = pv_opex
        st.session_state.bess_opex_per_kwh = bess_opex

        # Financial assumptions
        st.session_state.tariff_escalation = st.number_input(
            "Tariff Escalation (%/yr)", value=st.session_state.tariff_escalation,
            min_value=0.0, max_value=30.0, step=0.5)
        st.session_state.discount_rate = st.number_input(
            "Discount Rate (%)", value=st.session_state.discount_rate,
            min_value=0.0, max_value=50.0, step=0.5)
        st.session_state.tax_rate = st.number_input(
            "Corporate Tax Rate (%)", value=st.session_state.tax_rate,
            min_value=0.0, max_value=50.0, step=1.0)

        # ── Project Timeline ───────────────────────────────────
        st.markdown("---")
        st.markdown("**🗓️ Project Timeline**")

        import datetime as _dt_ui
        _tl_c1, _tl_c2 = st.columns(2)
        with _tl_c1:
            st.session_state.bess_lead_months = st.number_input(
                "BESS Delivery Lead (months)",
                value=int(st.session_state.get("bess_lead_months", 6)),
                min_value=1, max_value=36, step=1,
                help="Months from PO issuance to BESS commissioning",
            )
        with _tl_c2:
            st.session_state.pv_lead_months = st.number_input(
                "PV Delivery Lead (months)",
                value=int(st.session_state.get("pv_lead_months", 12)),
                min_value=1, max_value=36, step=1,
                help="Months from PO issuance to PV commissioning",
            )

        # Derived timeline display
        _po_today = _dt_ui.date.today()
        try:
            from dateutil.relativedelta import relativedelta as _rdelta_ui
            _bess_live_dt = _po_today + _rdelta_ui(months=int(st.session_state.bess_lead_months))
            _pv_live_dt   = _po_today + _rdelta_ui(months=int(st.session_state.pv_lead_months))
            _end_dt       = _pv_live_dt + _rdelta_ui(years=ANALYSIS_YEARS)
            _precomm_mo   = max(0, int(st.session_state.pv_lead_months) - int(st.session_state.bess_lead_months))
        except Exception:
            _bess_live_dt = _po_today
            _pv_live_dt   = _po_today
            _end_dt       = _po_today
            _precomm_mo   = 0

        _tl_lines = [
            f"📅 **PO Date:** {_po_today.strftime('%Y-%m-%d')}",
            f"⚡ **BESS Go-Live:** {_bess_live_dt.strftime('%Y-%m-%d')}  (+{st.session_state.bess_lead_months} mo)",
            f"☀️ **PV Go-Live:** {_pv_live_dt.strftime('%Y-%m-%d')}  (+{st.session_state.pv_lead_months} mo)",
            f"🏁 **Model End:** {_end_dt.strftime('%Y-%m-%d')}  (PV+{ANALYSIS_YEARS}yr)",
        ]
        if _precomm_mo > 0:
            _tl_lines.append(
                f"💡 **BESS pre-commission:** {_precomm_mo} months of BESS-only revenue included in Year-0"
            )

        st.markdown("\n\n".join(_tl_lines))


# ══════════════════════════════════════════════════════════════
# LEFT — Content Area (Tabs)
# ══════════════════════════════════════════════════════════════
with col_content:

    _tab_labels = [
        "📊 Run & Results",
        "📋 20-Year Financial Model",
        "⏱️ Hourly Dispatch",
        "🔍 Capacity Optimisation",
    ]
    if is_admin():
        _tab_labels.append("🔴 Admin")
    _all_tabs = st.tabs(_tab_labels)
    tab1, tab2, tab3, tab4 = _all_tabs[:4]
    _admin_tab = _all_tabs[4] if len(_all_tabs) > 4 else None

    # English-only: no column renaming needed; dict kept for safe .get() calls
    _COL_MAP: dict = {}

    # ──────────────────────────────────────────────────────────
    # Tab 1: Run & Results
    # ──────────────────────────────────────────────────────────
    with tab1:
        st.markdown("### 🚀 Run Physical Simulation (8,760-Hour Dispatch)")

        run_calc = st.button("▶ Run Simulation", type="primary",
                              use_container_width=True)

        if run_calc:
            if st.session_state.pv_kwp == 0 and st.session_state.bess_kwh == 0:
                st.error("❌ PV and BESS are both 0 — please configure at least one")
            else:
                # PVGIS
                if "pvgis_data" not in st.session_state or st.session_state.pv_kwp == 0:
                    pvgis_data = get_pvgis_data(
                        st.session_state.lat, st.session_state.lon,
                        max(st.session_state.pv_kwp, 1),
                        st.session_state.pv_loss, st.session_state.tilt, st.session_state.azimuth,
                    )
                    if st.session_state.pv_kwp == 0:
                        pvgis_data.update({"annual_kwh": 0, "monthly_kwh": [0] * 12,
                                           "winter_daily_kwh": 0, "summer_daily_kwh": 0})
                else:
                    pvgis_data = st.session_state.pvgis_data

                # Derive BESS power from C-rate selection
                bess_kw_use = st.session_state.bess_kwh * C_RATE_OPTIONS[st.session_state.c_rate_label]

                with st.spinner("⚙️ Running 8,760-hour dispatch engine…"):
                    dispatch_yr1 = run_8760_dispatch(
                        pv_kwp=st.session_state.pv_kwp,
                        bess_kwh=st.session_state.bess_kwh,
                        bess_kw=bess_kw_use,
                        load_peak_kw=st.session_state.load_peak_kw,
                        load_std_kw=st.session_state.load_std_kw,
                        load_offpeak_kw=st.session_state.load_offpeak_kw,
                        rte=st.session_state.rte,
                        dod=st.session_state.dod,
                        pvgis_data=pvgis_data,
                        load_profile_8760=st.session_state.get("load_profile_8760"),
                    )

                eol_years, annual_cycles = compute_bess_eol(
                    dispatch_yr1["tot_throughput_kWh"],
                    st.session_state.bess_kwh,
                    int(st.session_state.bess_cycles),
                    st.session_state.dod,
                )

                pv_zar_kwp, bess_zar_kwh = get_capex_zar()
                params = {
                    "pv_capex_per_kwp":   pv_zar_kwp,
                    "bess_capex_per_kwh": bess_zar_kwh,
                    "pv_opex_per_kwp":    st.session_state.pv_opex_per_kwp,
                    "bess_opex_per_kwh":  st.session_state.bess_opex_per_kwh,
                    "tariff_escalation":  st.session_state.tariff_escalation,
                    "discount_rate":      st.session_state.discount_rate,
                    "tax_rate":           st.session_state.tax_rate,
                    "pv_degradation":     st.session_state.pv_degradation,
                }

                _c_rate_val = C_RATE_OPTIONS[st.session_state.c_rate_label]

                # ── Commissioning timeline ────────────────────────────────
                import datetime as _dt_sim
                _bess_lead = int(st.session_state.get("bess_lead_months", 6))
                _pv_lead   = int(st.session_state.get("pv_lead_months",  12))
                _po_date   = _dt_sim.date.today()
                # Pre-commissioning BESS-only months: gap between BESS live and PV live
                # Only applicable when both PV and BESS are installed
                _both = (st.session_state.pv_kwp > 0 and st.session_state.bess_kwh > 0)
                _precomm_months = max(0, _pv_lead - _bess_lead) if _both else 0

                with st.spinner("📊 Building 20-year financial model…"):
                    fin_df, _precomm_ncf = run_20yr_financial_model(
                        dispatch_yr1=dispatch_yr1,
                        pv_kwp=st.session_state.pv_kwp,
                        bess_kwh=st.session_state.bess_kwh,
                        eol_years=eol_years,
                        params=params,
                        annual_cycles=annual_cycles,
                        c_rate=_c_rate_val,
                        precomm_bess_months=_precomm_months,
                    )

                total_capex = (st.session_state.pv_kwp * pv_zar_kwp +
                               st.session_state.bess_kwh * bess_zar_kwh)
                npv, irr = compute_npv_irr(
                    fin_df, total_capex, params["discount_rate"],
                    year0_extra_cf=_precomm_ncf,
                )

                lcoe_result = compute_lcoe(
                    pv_kwp=st.session_state.pv_kwp,
                    bess_kwh=st.session_state.bess_kwh,
                    params=params,
                    dispatch_yr1=dispatch_yr1,
                    annual_cycles=annual_cycles,
                    c_rate=_c_rate_val,
                )

                cum = fin_df["Cumulative CF (ZAR)"].tolist()
                # Linear interpolation for fractional payback year
                payback = None
                for _i, _v in enumerate(cum):
                    if _v >= 0:
                        if _i == 0:
                            payback = float(fin_df["Year"].iloc[0])
                        else:
                            _v0 = cum[_i - 1]          # negative
                            _y0 = float(fin_df["Year"].iloc[_i - 1])
                            payback = _y0 + (-_v0) / (_v - _v0)   # interpolate
                        break

                # Derived timeline dates
                from dateutil.relativedelta import relativedelta as _rdelta
                _bess_golive = _po_date + _rdelta(months=_bess_lead)
                _pv_golive   = _po_date + _rdelta(months=_pv_lead)
                _model_end   = _pv_golive + _rdelta(years=ANALYSIS_YEARS)

                st.session_state.results = {
                    "dispatch_yr1": dispatch_yr1, "eol_years": min(20.0, eol_years),
                    "annual_cycles": annual_cycles, "pvgis_data": pvgis_data,
                    "npv": npv, "irr": irr, "payback": payback,
                    "total_capex": total_capex, "bess_kw": bess_kw_use,
                    "c_rate": _c_rate_val,
                    "lcoe": lcoe_result.get("lcoe"),   # None if no PV
                    "lcos": lcoe_result.get("lcos"),   # None if no BESS
                    # timeline
                    "po_date":          _po_date.isoformat(),
                    "bess_lead_months": _bess_lead,
                    "pv_lead_months":   _pv_lead,
                    "bess_golive":      _bess_golive.isoformat(),
                    "pv_golive":        _pv_golive.isoformat(),
                    "model_end":        _model_end.isoformat(),
                    "precomm_months":   _precomm_months,
                    "precomm_ncf":      float(_precomm_ncf),
                }
                st.session_state.hourly_df = dispatch_yr1["hourly_df"]
                st.session_state.fin_df = fin_df
                st.success("✅ Calculation complete!")

        # ── Display simulation results ──
        if st.session_state.results:
            res = st.session_state.results

            st.markdown('<div class="section-header">📈 Key Financial Metrics</div>',
                        unsafe_allow_html=True)

            # ── Row 1: core financials (always 4 columns) ──────────────────
            m1, m2, m3, m4 = st.columns(4)

            with m1:
                st.markdown(f"""<div class="metric-card">
                    <div class="metric-label">Total CAPEX</div>
                    <div class="metric-value">R{res['total_capex']/1e6:.2f}M</div>
                    <div class="metric-unit">ZAR ({st.session_state.forex_usd_zar:.1f} USD/ZAR)</div>
                </div>""", unsafe_allow_html=True)

            with m2:
                _npv_c = "var(--primary)" if res['npv'] > 0 else "var(--danger)"
                st.markdown(f"""<div class="metric-card">
                    <div class="metric-label">NPV (20-Year)</div>
                    <div class="metric-value" style="color:{_npv_c}">R{res['npv']/1e6:.2f}M</div>
                    <div class="metric-unit">@ {st.session_state.discount_rate:.1f}% disc</div>
                </div>""", unsafe_allow_html=True)

            with m3:
                st.markdown(f"""<div class="metric-card">
                    <div class="metric-label">Project IRR</div>
                    <div class="metric-value">{res['irr']:.1f}%</div>
                    <div class="metric-unit">20-Year Project IRR</div>
                </div>""", unsafe_allow_html=True)

            with m4:
                pb = f"{res['payback']:.2f}yr" if res['payback'] else "20yr+"
                st.markdown(f"""<div class="metric-card">
                    <div class="metric-label">Simple Payback</div>
                    <div class="metric-value">{pb}</div>
                    <div class="metric-unit">Cumulative CF = 0</div>
                </div>""", unsafe_allow_html=True)

            # ── Row 2: LCOE / LCOS / BESS EoL (scenario-adaptive) ──────────
            # LCOE: PV cost only / PV gen kWh  — shown only when PV present
            # LCOS: BESS cost only / discharge kWh — shown only when BESS present
            # BESS EoL: always shown when BESS present
            _lcoe_d = res.get("lcoe")   # dict or None
            _lcos_d = res.get("lcos")   # dict or None
            _r2_items = (
                (["lcoe"] if _lcoe_d else [])
                + (["lcos"] if _lcos_d else [])
                + (["eol"]  if _lcos_d else [])   # BESS EoL when BESS present
            )
            if _r2_items:
                _r2_cols = st.columns(len(_r2_items))
                for _ri, _rt in enumerate(_r2_items):
                    with _r2_cols[_ri]:
                        if _rt == "lcoe":
                            _lv  = _lcoe_d["lcoe_zar_kwh"]
                            _lc  = "var(--primary)" if _lv > 0 else "var(--text-dim)"
                            _lmwh = _lcoe_d["total_avoided_mwh"]
                            st.markdown(f"""<div class="metric-card">
                                <div class="metric-label">1st yr LCOE · PV only</div>
                                <div class="metric-value" style="color:{_lc}">R{_lv:.2f}</div>
                                <div class="metric-unit">/kWh · {_lmwh:,.0f} MWh generated</div>
                            </div>""", unsafe_allow_html=True)
                        elif _rt == "lcos":
                            _sv   = _lcos_d["lcos_zar_kwh"]
                            _sc   = "var(--primary)" if _sv > 0 else "var(--text-dim)"
                            _smwh = _lcos_d["total_discharge_mwh"]
                            st.markdown(f"""<div class="metric-card">
                                <div class="metric-label">1st yr LCOS · BESS only</div>
                                <div class="metric-value" style="color:{_sc}">R{_sv:.2f}</div>
                                <div class="metric-unit">/kWh · {_smwh:,.0f} MWh discharged</div>
                            </div>""", unsafe_allow_html=True)
                        elif _rt == "eol":
                            st.markdown(f"""<div class="metric-card">
                                <div class="metric-label">BESS EoL</div>
                                <div class="metric-value" style="color:var(--secondary)">Yr {min(20, int(res['eol_years']))}</div>
                                <div class="metric-unit">{res['annual_cycles']:.1f} cycles/yr</div>
                            </div>""", unsafe_allow_html=True)

            # ── Project Timeline banner ───────────────────────────
            _tl = {k: res.get(k) for k in ("po_date","bess_golive","pv_golive","model_end","precomm_months","precomm_ncf")}
            if _tl.get("po_date"):
                _pc_mo  = _tl["precomm_months"] or 0
                _pc_ncf = _tl["precomm_ncf"] or 0.0
                _tl_html = (
                    f"<b>Project Timeline:</b>&nbsp;&nbsp;"
                    f"📅 PO: <b>{_tl['po_date']}</b>"
                    f" &rarr; ⚡ BESS: <b>{_tl['bess_golive']}</b>"
                    f" &rarr; ☀️ PV: <b>{_tl['pv_golive']}</b>"
                    f" &rarr; 🏁 End: <b>{_tl['model_end']}</b>"
                    + (f"&nbsp;&nbsp;|&nbsp;&nbsp;💡 BESS pre-commission: <b>{_pc_mo} months</b>"
                       f" → Year-0 BESS net saving: <b>R {_pc_ncf/1000:.1f}K</b>" if _pc_mo > 0 else "")
                )
                st.markdown(
                    f'<div class="info-box" style="font-size:0.82rem;padding:6px 12px;">{_tl_html}</div>',
                    unsafe_allow_html=True
                )

            st.markdown('<div class="section-header">⚡ Year 1 Operations</div>',
                        unsafe_allow_html=True)
            d = res['dispatch_yr1']
            o1, o2, o3, o4 = st.columns(4)

            with o1:
                st.markdown(f"""<div class="metric-card">
                    <div class="metric-label">Yr1 Total Saving</div>
                    <div class="metric-value">R{d['annual_saving_ZAR']/1000:.1f}K</div>
                    <div class="metric-unit">PV:{d['annual_pv_saving_ZAR']/1000:.1f}K + BESS:{d['annual_bess_saving_ZAR']/1000:.1f}K</div>
                </div>""", unsafe_allow_html=True)

            with o2:
                st.markdown(f"""<div class="metric-card">
                    <div class="metric-label">Annual PV Generation</div>
                    <div class="metric-value">{d['annual_pv_gen_kWh']/1000:.1f}</div>
                    <div class="metric-unit">MWh/yr</div>
                </div>""", unsafe_allow_html=True)

            with o3:
                st.markdown(f"""<div class="metric-card">
                    <div class="metric-label">BESS Throughput</div>
                    <div class="metric-value">{d['tot_throughput_kWh']/1000:.1f}</div>
                    <div class="metric-unit">MWh (charge×RTE + discharge)</div>
                </div>""", unsafe_allow_html=True)

            with o4:
                self_pct = min(d['annual_pv_gen_kWh'] / max(d['annual_load_kWh'], 1) * 100, 100)
                st.markdown(f"""<div class="metric-card">
                    <div class="metric-label">PV Self-Consumption</div>
                    <div class="metric-value">{self_pct:.1f}%</div>
                    <div class="metric-unit">PV vs Annual Load</div>
                </div>""", unsafe_allow_html=True)

            st.markdown('<div class="section-header">📊 20-Year Cash Flow</div>',
                        unsafe_allow_html=True)

            if st.session_state.fin_df is not None:
                import plotly.graph_objects as go
                from plotly.subplots import make_subplots

                df = st.session_state.fin_df
                eol_yr = min(20, int(res['eol_years']))

                fig = make_subplots(rows=1, cols=2,
                                    subplot_titles=("Annual Net Cash Flow (ZAR)",
                                                    "Cumulative Cash Flow (ZAR)"))

                colors = ["#00E5A0" if v >= 0 else "#FF4444" for v in df["Net Cash Flow NCF (ZAR)"]]
                fig.add_trace(go.Bar(x=df["Year"], y=df["Net Cash Flow NCF (ZAR)"],
                                     marker_color=colors, name="NCF",
                                     text=[f"R{v/1000:.0f}K" for v in df["Net Cash Flow NCF (ZAR)"]],
                                     textposition="outside", textfont=dict(size=7)),
                              row=1, col=1)

                fig.add_vline(x=eol_yr, line_dash="dash", line_color="#FF6B35", line_width=2,
                              annotation_text=f"BESS EoL Yr{eol_yr}",
                              annotation_font_color="#FF6B35", row=1, col=1)

                fig.add_trace(go.Scatter(x=df["Year"], y=df["Cumulative CF (ZAR)"],
                                         mode="lines+markers",
                                         line=dict(color="#4ECDC4", width=2),
                                         marker=dict(size=4), fill="tozeroy",
                                         fillcolor="rgba(78,205,196,0.08)",
                                         name="CF"),
                              row=1, col=2)

                fig.add_hline(y=0, line_dash="dot", line_color="white", line_width=1, row=1, col=2)
                fig.add_vline(x=eol_yr, line_dash="dash", line_color="#FF6B35", line_width=2,
                              annotation_text="BESS EoL", annotation_font_color="#FF6B35",
                              row=1, col=2)

                _pb, _plb, _fc, _gc, _lb = _plt()
                fig.update_layout(paper_bgcolor=_pb, plot_bgcolor=_plb,
                                  font=dict(color=_fc, family="IBM Plex Mono"),
                                  showlegend=False, height=380,
                                  margin=dict(l=10, r=10, t=50, b=10))
                fig.update_xaxes(gridcolor=_gc, title_text="Year")
                fig.update_yaxes(gridcolor=_gc)
                st.plotly_chart(fig, use_container_width=True)

            st.markdown('<div class="section-header">📥 Data Export</div>', unsafe_allow_html=True)
            dl1, dl2 = st.columns(2)

            with dl1:
                if st.session_state.hourly_df is not None:
                    buf = io.StringIO()
                    st.session_state.hourly_df.to_csv(buf, index=False, encoding="utf-8-sig")
                    st.download_button("⬇ 8760_Hourly_Dispatch.csv",
                                       data=buf.getvalue().encode("utf-8-sig"),
                                       file_name="8760_Hourly_Dispatch.csv",
                                       mime="text/csv", use_container_width=True)

            with dl2:
                if st.session_state.fin_df is not None:
                    _mk = "Metric"
                    _vk = "Value"
                    xbuf = io.BytesIO()
                    with pd.ExcelWriter(xbuf, engine="openpyxl") as writer:
                        st.session_state.fin_df.rename(columns=_COL_MAP).to_excel(
                            writer, sheet_name="20yr_Financial_Model", index=False)
                        _lcoe_x = res.get("lcoe") or {}
                        _lcos_x = res.get("lcos") or {}
                        _xrows  = [
                            {_mk: "Annual Saving (ZAR)",        _vk: res['dispatch_yr1']['annual_saving_ZAR']},
                            {_mk: "PV Saving (ZAR)",            _vk: res['dispatch_yr1']['annual_pv_saving_ZAR']},
                            {_mk: "BESS Net Saving (ZAR)",      _vk: res['dispatch_yr1']['annual_bess_saving_ZAR']},
                            {_mk: "Annual PV Gen (kWh)",        _vk: res['dispatch_yr1']['annual_pv_gen_kWh']},
                            {_mk: "Annual Avoided Grid (kWh)",  _vk: res['dispatch_yr1']['annual_load_kWh'] - res['dispatch_yr1']['annual_grid_buy_kWh']},
                            {_mk: "Throughput (kWh)",           _vk: res['dispatch_yr1']['tot_throughput_kWh']},
                            {_mk: "Annual Cycles",              _vk: res['annual_cycles']},
                            {_mk: "BESS EoL (yr)",              _vk: min(20.0, res['eol_years'])},
                            {_mk: "NPV (ZAR)",                  _vk: res['npv']},
                            {_mk: "IRR (%)",                    _vk: res['irr']},
                            {_mk: "USD/ZAR Rate",               _vk: st.session_state.forex_usd_zar},
                        ]
                        # LCOE — PV only, shown only if PV present
                        if _lcoe_x:
                            _xrows += [
                                {_mk: "LCOE · PV only (ZAR/kWh)",    _vk: _lcoe_x.get("lcoe_zar_kwh", "")},
                                {_mk: "  Year-1 PV Generation (kWh)", _vk: _lcoe_x.get("avoided_kwh_yr1", "")},
                                {_mk: "  PV Annualised Cost (ZAR/yr)",_vk: _lcoe_x.get("annual_cost_zar", "")},
                                {_mk: "  PV Lifetime Gen (MWh)",      _vk: _lcoe_x.get("total_avoided_mwh", "")},
                            ]
                        # LCOS — BESS only, shown only if BESS present
                        if _lcos_x:
                            _xrows += [
                                {_mk: "LCOS · BESS only (ZAR/kWh)",       _vk: _lcos_x.get("lcos_zar_kwh", "")},
                                {_mk: "  Year-1 BESS Discharge (kWh)",     _vk: _lcos_x.get("discharge_kwh_yr1", "")},
                                {_mk: "  BESS Annualised Cost (ZAR/yr)",   _vk: _lcos_x.get("annual_cost_zar", "")},
                                {_mk: "  BESS Lifetime Discharge (MWh)",   _vk: _lcos_x.get("total_discharge_mwh", "")},
                            ]
                        ops = pd.DataFrame(_xrows)
                        ops.to_excel(writer, sheet_name="Summary", index=False)
                    st.download_button("⬇ BTM_Pure_Data_Export.xlsx",
                                       data=xbuf.getvalue(),
                                       file_name="BTM_Pure_Data_Export.xlsx",
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                       use_container_width=True)

            # ── Excel / Professional Excel Report ──
            st.markdown("---")
            st.markdown(
                '<div class="section-header">📊 Professional Excel Report</div>',
                unsafe_allow_html=True,
            )
            st.markdown("""<div class="info-box">
                6-sheet report: Cover · Parameters (with Excel formulas) · 20Y Financial Model · Monthly Summary · Charts · 8760h Raw Data
                &nbsp;|&nbsp; White cells in Parameters sheet are editable; Sheet3 recalculates automatically
                &nbsp;|&nbsp; 🔵 <b>Pro / Admin feature</b>
            </div>""", unsafe_allow_html=True)

            if not is_pro():
                # Free users see a locked message
                st.markdown("""<div class="warning-box">
                    🔒 <b>Pro / Pro Feature</b> — Excel Pro Admin <br>
                    <br>
                    Excel export is available for 🔵 Pro and 🔴 Admin accounts only.
                    Contact admin to upgrade.
                </div>""", unsafe_allow_html=True)
            else:
                _xbtn_col, _xdl_col = st.columns([2, 3])
                with _xbtn_col:
                    gen_excel_btn = st.button(
                        "📊 Export Financial Report",
                        key="gen_excel_report_btn",
                        use_container_width=True,
                    )

                if gen_excel_btn:
                    with st.spinner("📊 Generating 6-sheet Excel report, please wait..."):
                        try:
                            _xlsx_bytes = generate_excel_report()
                            st.session_state["_excel_rpt_bytes"] = _xlsx_bytes
                            import datetime as _edt
                            _xproj = (st.session_state.get("_active_snap_name") or "BTM_Project")
                            _xproj_safe = re.sub(r'[^\w\-]', '_', _xproj).strip('_')
                            _xirr = ((st.session_state.results or {}).get("irr") or 0)
                            st.session_state["_excel_rpt_fname"] = (
                                f"{_xproj_safe}_IRR{_xirr:.1f}pct_{_edt.datetime.now():%Y%m%d}.xlsx"
                            )
                            st.success("✅ Report ready — click below to download")
                        except Exception as _e:
                            st.error(f"❌ Generation failed: {_e}")

                if "_excel_rpt_bytes" in st.session_state:
                    with _xdl_col:
                        st.markdown("<br>", unsafe_allow_html=True)
                        st.download_button(
                            "⬇ Download Excel Report",
                            data=st.session_state["_excel_rpt_bytes"],
                            file_name=st.session_state.get(
                                "_excel_rpt_fname", "BTM_Report.xlsx"),
                            mime=("application/vnd.openxmlformats-"
                                  "officedocument.spreadsheetml.sheet"),
                            use_container_width=True,
                        )

            # ── PPTX / PowerPoint Presentation Report ──────────────
            st.markdown("---")
            st.markdown(
                '<div class="section-header">📊 PowerPoint Presentation Report</div>',
                unsafe_allow_html=True,
            )
            st.markdown("""<div class="info-box">
                8-slide executive deck: Cover · Investment Thesis · System Overview ·
                Financial Returns · Energy Analysis · Tariff Opportunity ·
                Implementation Roadmap · Assumptions
                &nbsp;|&nbsp; White professional template · McKinsey-style narrative per slide
            </div>""", unsafe_allow_html=True)

            # Report branding inputs
            _rpt_col1, _rpt_col2 = st.columns(2)
            with _rpt_col1:
                st.text_input(
                    "EPC / Consultant",
                    value=st.session_state.get("_pptx_consultant", ""),
                    placeholder="e.g. Lanxi Engineering",
                    key="_pptx_consultant",
                )
            with _rpt_col2:
                st.text_input(
                    "Client Name",
                    value=st.session_state.get("_pptx_client_name", ""),
                    placeholder="e.g. Lanxi Mining Company",
                    help='Appears on cover: "Confidential · Only For: {Client Name}"',
                    key="_pptx_client_name",
                )
            # OEM is hardcoded per Huawei partner requirement
            st.caption("🏭 OEM: **Huawei Technologies SA PTY LTD** (fixed — Huawei partner requirement)")

            _pp_btn_col, _pp_dl_col = st.columns([2, 3])
            with _pp_btn_col:
                _gen_pptx_btn = st.button(
                    "📊 Export PowerPoint Report",
                    key="gen_pptx_btn",
                    use_container_width=True,
                )

            if _gen_pptx_btn:
                with st.spinner("📊 Building executive PPTX report…"):
                    try:
                        from report_pptx import generate_pptx as _gen_pptx
                        import datetime as _edt

                        # Collect all params from session state
                        _pptx_params = {
                            k: st.session_state.get(k)
                            for k in [
                                "pv_kwp", "bess_kwh", "c_rate_label", "c_rate",
                                "pv_loss", "tilt", "azimuth", "pv_degradation",
                                "rte", "dod", "bess_cycles",
                                "tariff_mode",
                                "w_morning_peak", "w_evening_peak", "w_standard", "w_off_peak",
                                "s_morning_peak", "s_evening_peak", "s_standard", "s_off_peak",
                                "tariff_escalation", "discount_rate", "tax_rate",
                                "forex_usd_zar", "pv_usd_per_w", "bess_usd_per_wh",
                                "pv_opex_per_kwp", "bess_opex_per_kwh",
                                "lat", "lon",
                            ]
                        }
                        # Merge results-derived capex + tax_rate into params
                        _res = st.session_state.results or {}
                        _pptx_params["tax_rate"] = _pptx_params.get("tax_rate") or 27
                        _pptx_params["c_rate"] = C_RATE_OPTIONS.get(
                            str(st.session_state.get("c_rate_label", "0.25C")), 0.25)

                        _pptx_bytes = _gen_pptx(
                            params           = _pptx_params,
                            results          = _res,
                            fin_df           = st.session_state.fin_df,
                            pvgis_data       = _res.get("pvgis_data") or st.session_state.get("pvgis_data") or {},
                            project_name     = st.session_state.get("_active_snap_name", ""),
                            client_name      = st.session_state.get("_pptx_client_name", ""),
                            consultant_name  = st.session_state.get("_pptx_consultant", ""),
                        )
                        st.session_state["_pptx_bytes"] = _pptx_bytes
                        _pproj = (st.session_state.get("_active_snap_name") or "BTM_Project")
                        _pproj_safe = re.sub(r'[^\w\-]', '_', _pproj).strip('_')
                        _ppirr = (_res.get("irr") or 0)
                        st.session_state["_pptx_fname"] = (
                            f"{_pproj_safe}_IRR{_ppirr:.1f}pct_{_edt.datetime.now():%Y%m%d}.pptx"
                        )
                        st.success("✅ PPTX ready — click below to download")
                    except Exception as _pe:
                        st.error(f"❌ PPTX generation failed: {_pe}")

            if "_pptx_bytes" in st.session_state:
                with _pp_dl_col:
                    st.markdown("<br>", unsafe_allow_html=True)
                    st.download_button(
                        "⬇ Download PowerPoint Report",
                        data      = st.session_state["_pptx_bytes"],
                        file_name = st.session_state.get("_pptx_fname", "BTM_Report.pptx"),
                        mime      = ("application/vnd.openxmlformats-officedocument"
                                     ".presentationml.presentation"),
                        use_container_width=True,
                    )

            st.markdown(
                "**Solution Info:** "
                "[info.support.huawei.com/Energy](https://info.support.huawei.com/Energy/info/en_US/all/index)",
                unsafe_allow_html=False,
            )

    # ──────────────────────────────────────────────────────────
    # Tab 2: 20-Year Financial Model
    # ──────────────────────────────────────────────────────────
    with tab2:
        st.markdown("### 📋 20-Year Annual Financial Statement")

        if st.session_state.fin_df is not None:
            st.markdown("""<div class="info-box">
                📌 <b>Section 12B</b> accelerated depreciation (PV, &gt;1MW): Yr1 50% · Yr2 30% · Yr3 20% · Pure BESS: 20%×5yr straight-line (SA Income Tax Act) &nbsp;|&nbsp;
                SA assessed loss carry-forward applied — early-year losses reduce tax in future profitable years &nbsp;|&nbsp;
                PV &amp; BESS savings shown separately &nbsp;|&nbsp; BESS revenue &amp; OPEX zeroed after EoL (SOH &lt; 60%)
            </div>""", unsafe_allow_html=True)

            eol_yr = min(20, int(st.session_state.results["eol_years"])) if st.session_state.results else 999

            _display_fin = st.session_state.fin_df.rename(columns=_COL_MAP)
            _yr_col   = _COL_MAP.get("Year", "Year")
            _soh_col  = _COL_MAP.get("SOH%", "SOH%")

            def highlight_eol(row):
                if row[_yr_col] > eol_yr:
                    return ["background-color: rgba(255,107,53,0.08)"] * len(row)
                elif row[_yr_col] <= 3:
                    return ["background-color: rgba(0,229,160,0.05)"] * len(row)
                return [""] * len(row)

            num_cols = {c: "{:,.0f}" for c in _display_fin.columns
                        if _display_fin[c].dtype in [np.float64, np.int64]
                        and c not in [_yr_col, _soh_col]}

            styled = _display_fin.style.apply(highlight_eol, axis=1).format(num_cols)
            st.dataframe(styled, use_container_width=True, height=580)

        else:
            st.markdown('<div class="warning-box">⚠️ Please run the simulation first (Run & Results tab)</div>',
                        unsafe_allow_html=True)

    # ──────────────────────────────────────────────────────────
    # Tab 3: Hourly Dispatch (8760h)
    # ──────────────────────────────────────────────────────────
    with tab3:
        st.markdown("### ⏱️ Hourly Dispatch Log")

        if st.session_state.hourly_df is not None:
            import plotly.graph_objects as go

            df_h = st.session_state.hourly_df

            st.markdown('<div class="section-header">📅 Monthly Summary</div>',
                        unsafe_allow_html=True)
            mth_agg = df_h.groupby("month").agg({
                "pv_gen_kWh": "sum", "charge_grid_kWh": "sum",
                "discharge_kWh": "sum", "grid_buy_kWh": "sum",
                "net_saving_ZAR": "sum",
            }).reset_index()

            mnames = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

            fig_m = go.Figure()
            fig_m.add_trace(go.Bar(x=mnames, y=mth_agg["pv_gen_kWh"],
                                    name="PV Gen", marker_color="#F6C90E"))
            fig_m.add_trace(go.Bar(x=mnames, y=mth_agg["discharge_kWh"],
                                    name="BESS Discharge", marker_color="#00E5A0"))
            fig_m.add_trace(go.Bar(x=mnames, y=mth_agg["charge_grid_kWh"],
                                    name="Grid Charge", marker_color="#4ECDC4"))
            fig_m.add_trace(go.Scatter(x=mnames, y=mth_agg["net_saving_ZAR"],
                                        name="Net Saving (ZAR)", yaxis="y2",
                                        line=dict(color="#FF6B35", width=2),
                                        marker=dict(size=5)))
            # Mark winter months
            for mi, mn in enumerate(mnames):
                if (mi + 1) in WINTER_MONTHS:
                    fig_m.add_vrect(x0=mn, x1=mn, fillcolor="#4ECDC4",
                                    opacity=0.08, line_width=0)

            _pb, _plb, _fc, _gc, _lb = _plt()
            fig_m.update_layout(
                barmode="stack", paper_bgcolor=_pb, plot_bgcolor=_plb,
                font=dict(color=_fc, family="IBM Plex Mono"),
                yaxis=dict(title="Energy (kWh)", gridcolor=_gc),
                yaxis2=dict(title="Saving (ZAR)", overlaying="y", side="right"),
                legend=dict(bgcolor=_lb, bordercolor=_gc, orientation="h"),
                height=320, margin=dict(l=10, r=10, t=20, b=10),
                annotations=[dict(text="❄️ High Season Jun–Aug", x="Jun", y=1.05,
                                  xref="x", yref="paper", showarrow=False,
                                  font=dict(color="#4ECDC4", size=9))]
            )
            st.plotly_chart(fig_m, use_container_width=True)

            # + — Pro+ only
            if not is_pro():
                st.markdown("""<div class="warning-box">
                    🔒 <b>Pro / Pro Feature</b> — Pro Admin <br>
                    Typical day profile &amp; hourly log available for 🔵 Pro / 🔴 Admin accounts only.
                </div>""", unsafe_allow_html=True)
            else:
                st.markdown('<div class="section-header">📈 Typical Daily Dispatch Profile</div>',
                            unsafe_allow_html=True)
                sel_month = st.selectbox("Select Month", range(1, 13),
                                         format_func=lambda x: mnames[x-1], index=6)

                mdata = df_h[df_h["month"] == sel_month]
                if len(mdata) > 0:
                    tgt = int(mdata["day"].median())
                    ddata = mdata[mdata["day"] == tgt]
                    if len(ddata) == 24:
                        fig_d = go.Figure()
                        fig_d.add_trace(go.Scatter(x=ddata["hour_of_day"], y=ddata["load_kWh"],
                                                   name="Load", line=dict(color="white", width=2, dash="dot")))
                        fig_d.add_trace(go.Bar(x=ddata["hour_of_day"], y=ddata["pv_gen_kWh"],
                                               name="PV Gen", marker_color="#F6C90E", opacity=0.8))
                        fig_d.add_trace(go.Bar(x=ddata["hour_of_day"], y=ddata["discharge_kWh"],
                                               name="BESS Discharge", marker_color="#00E5A0", opacity=0.8))
                        fig_d.add_trace(go.Bar(x=ddata["hour_of_day"],
                                               y=[-v for v in ddata["charge_grid_kWh"]],
                                               name="Grid Charge", marker_color="#4ECDC4", opacity=0.7))
                        fig_d.add_trace(go.Scatter(x=ddata["hour_of_day"], y=ddata["soc_kWh"],
                                                   name="SOC (kWh)", yaxis="y2",
                                                   line=dict(color="#FF6B35", width=2)))
                        fig_d.add_trace(go.Scatter(x=ddata["hour_of_day"],
                                                   y=ddata["tariff_ZAR_kWh"] * 50,
                                                   name="Tariff×50", yaxis="y2",
                                                   line=dict(color="#E040FB", width=1.5, dash="dash")))
                        _pb, _plb, _fc, _gc, _lb = _plt()
                        fig_d.update_layout(
                            paper_bgcolor=_pb, plot_bgcolor=_plb,
                            font=dict(color=_fc, family="IBM Plex Mono"),
                            barmode="overlay",
                            yaxis=dict(title="kWh/h", gridcolor=_gc),
                            yaxis2=dict(title="SOC/Tariff", overlaying="y", side="right"),
                            legend=dict(bgcolor=_lb, orientation="h"),
                            height=360, margin=dict(l=10, r=10, t=20, b=10),
                            xaxis=dict(title="Hour", tickvals=list(range(24)), gridcolor=_gc),
                        )
                        st.plotly_chart(fig_d, use_container_width=True)

                st.markdown('<div class="section-header">📋 Hourly Log (First 100 Hours)</div>',
                            unsafe_allow_html=True)
                st.dataframe(df_h.head(100), use_container_width=True, height=380)

        else:
            st.markdown(f'<div class="warning-box">⚠️ Run simulation first (Tab 1)</div>', unsafe_allow_html=True)

    # ──────────────────────────────────────────────────────────
    # Tab 4: Capacity Optimisation
    # ──────────────────────────────────────────────────────────
    with tab4:
        st.markdown("### 🔍 Global Capacity Optimisation (Max NPV)")

        st.markdown("""<div class="info-box">
            🤖 Sweeps all PV × BESS combinations within 60%–150% of current sizing to find the NPV-maximising configuration.
            BESS step fixed at 5 MWh; PV step is user-adjustable.
        </div>""", unsafe_allow_html=True)

        # ── Sweep 60%~150% of current sizing ──────────────────────────
        _BESS_STEP = 5000.0 # 5 MWh
        cur_pv   = float(st.session_state.pv_kwp)
        cur_bess = float(st.session_state.bess_kwh)

        def _snap_bess(v: float) -> float:
            """ 5 MWh """
            return max(_BESS_STEP, round(v / _BESS_STEP) * _BESS_STEP)

        auto_pv_min   = float(max(100.0, round(cur_pv  * 0.6 / 100) * 100))
        auto_pv_max   = float(max(100.0, round(cur_pv  * 1.5 / 100) * 100))
        auto_bess_min = float(_snap_bess(cur_bess * 0.6))
        auto_bess_max = float(_snap_bess(cur_bess * 1.5))

        oc1, oc2 = st.columns(2)
        with oc1:
            st.markdown(f"**PV Range** (current {_fmw(cur_pv,'kWp')} × 60–150%)")
            pv_min = st.number_input("PV Min (kWp)", value=auto_pv_min, min_value=100.0, step=100.0)
            pv_max = st.number_input("PV Max (kWp)", value=auto_pv_max, min_value=100.0, step=100.0)
            pv_stp = st.number_input("PV Step (kWp)", value=200.0, min_value=100.0, step=100.0)
        with oc2:
            st.markdown(f"**BESS Range** (current {_fmw(cur_bess,'kWh')} × 60–150%, fixed 5 MWh step)")
            bess_min = st.number_input("BESS Min (kWh)", value=auto_bess_min,
                                        min_value=_BESS_STEP, step=_BESS_STEP)
            bess_max = st.number_input("BESS Max (kWh)", value=auto_bess_max,
                                        min_value=_BESS_STEP, step=_BESS_STEP)
            # BESS step 5 MWh
            st.markdown(f'<div class="derived-value">BESS Step fixed at 5 MWh (5,000 kWh)</div>',
                        unsafe_allow_html=True)
            bess_stp = _BESS_STEP

        # Snap inputs to 5 MWh multiples
        bess_min = _snap_bess(bess_min)
        bess_max = _snap_bess(bess_max)

        pv_range   = np.arange(pv_min,   pv_max   + pv_stp,   pv_stp)
        bess_range = np.arange(bess_min, bess_max + bess_stp, bess_stp)
        n_combos   = len(pv_range) * len(bess_range)

        st.markdown(
            f"Combinations: **{n_combos}** "
            f"(PV {len(pv_range)} × BESS {len(bess_range)})"
        )
        if n_combos > 200:
            st.markdown(f'<div class="warning-box">⚠️ >200 combinations — may be slow. Increase PV step size to reduce.</div>',
                        unsafe_allow_html=True)

        if st.button("🚀 Start Optimisation", type="primary", use_container_width=True):
            if n_combos > 500:
                st.error("❌ Exceeds 500-combination limit — please increase PV step size")
            else:
                opt_pvgis = get_pvgis_data(
                    st.session_state.lat, st.session_state.lon,
                    max(float(pv_max), 1.0),
                    st.session_state.pv_loss, st.session_state.tilt, st.session_state.azimuth,
                )
                pv_zar_k, bess_zar_k = get_capex_zar()
                params_opt = {
                    "pv_capex_per_kwp":   pv_zar_k,
                    "bess_capex_per_kwh": bess_zar_k,
                    "pv_opex_per_kwp":    st.session_state.pv_opex_per_kwp,
                    "bess_opex_per_kwh":  st.session_state.bess_opex_per_kwh,
                    "tariff_escalation":  st.session_state.tariff_escalation,
                    "discount_rate":      st.session_state.discount_rate,
                    "tax_rate":           st.session_state.tax_rate,
                    "pv_degradation":     st.session_state.pv_degradation,
                }
                # C
                c_actual = C_RATE_OPTIONS[st.session_state.c_rate_label]

                prog = st.progress(0)
                stat = st.empty()
                results_opt = []
                done = 0

                for pv_o in pv_range:
                    for bess_o in bess_range:
                        scale = pv_o / max(float(pv_max), 1.0)
                        pvgis_s = {
                            **opt_pvgis,
                            "monthly_kwh": [v * scale for v in opt_pvgis["monthly_kwh"]],
                            "annual_kwh":        opt_pvgis["annual_kwh"] * scale,
                            "winter_daily_kwh":  opt_pvgis["winter_daily_kwh"] * scale,
                            "summer_daily_kwh":  opt_pvgis["summer_daily_kwh"] * scale,
                        }
                        try:
                            d_o = run_8760_dispatch(
                                pv_kwp=pv_o, bess_kwh=bess_o,
                                bess_kw=bess_o * c_actual,
                                load_peak_kw=st.session_state.load_peak_kw,
                                load_std_kw=st.session_state.load_std_kw,
                                load_offpeak_kw=st.session_state.load_offpeak_kw,
                                rte=st.session_state.rte,
                                dod=st.session_state.dod,
                                pvgis_data=pvgis_s,
                                load_profile_8760=st.session_state.get("load_profile_8760"),
                            )
                            eol_o, ac_o = compute_bess_eol(
                                d_o["tot_throughput_kWh"], bess_o,
                                int(st.session_state.bess_cycles), st.session_state.dod,
                            )
                            _precomm_m_o = max(0, int(st.session_state.get("pv_lead_months", 12))
                                               - int(st.session_state.get("bess_lead_months", 6))
                                               ) if (pv_o > 0 and bess_o > 0) else 0
                            fin_o, _precomm_ncf_o = run_20yr_financial_model(
                                dispatch_yr1=d_o, pv_kwp=pv_o, bess_kwh=bess_o,
                                eol_years=eol_o, params=params_opt,
                                annual_cycles=ac_o, c_rate=c_actual,
                                precomm_bess_months=_precomm_m_o,
                            )
                            cap_o = pv_o * pv_zar_k + bess_o * bess_zar_k
                            npv_o, irr_o = compute_npv_irr(
                                fin_o, cap_o, params_opt["discount_rate"],
                                year0_extra_cf=_precomm_ncf_o,
                            )
                            results_opt.append({
                                "PV (kWp)": pv_o, "BESS (kWh)": bess_o,
                                "CAPEX (ZAR)": cap_o, "NPV (ZAR)": npv_o,
                                "IRR (%)": irr_o, "EoL (yr)": min(20.0, eol_o),
                                "Annual Cycles": ac_o,
                                "Yr1 Saving (ZAR)": d_o["annual_saving_ZAR"],
                            })
                        except Exception:
                            pass
                        done += 1
                        prog.progress(done / n_combos)
                        if done % 5 == 0:
                            stat.text(f"⚙️ {done}/{n_combos}")

                prog.progress(1.0)
                stat.text("✅ Optimisation complete!")

                if results_opt:
                    import plotly.graph_objects as go
                    opt_df = pd.DataFrame(results_opt)
                    best = opt_df.loc[opt_df["NPV (ZAR)"].idxmax()]

                    st.markdown("#### 🏆 Optimal Configuration")
                    bc1, bc2, bc3, bc4 = st.columns(4)
                    with bc1:
                        _opv = best['PV (kWp)']
                        st.markdown(f"""<div class="metric-card">
                            <div class="metric-label">Optimal PV</div>
                            <div class="metric-value">{_fmw(_opv,'kWp')}</div>
                            <div class="metric-unit">Solar PV capacity</div></div>""", unsafe_allow_html=True)
                    with bc2:
                        _obs = best['BESS (kWh)']
                        st.markdown(f"""<div class="metric-card">
                            <div class="metric-label">Optimal BESS</div>
                            <div class="metric-value">{_fmw(_obs,'kWh')}</div>
                            <div class="metric-unit">{_fmw(_obs*c_actual,'kW')} max power</div></div>""",
                            unsafe_allow_html=True)
                    with bc3:
                        st.markdown(f"""<div class="metric-card">
                            <div class="metric-label">Best NPV (ZAR)</div>
                            <div class="metric-value" style="color:var(--primary)">R{best['NPV (ZAR)']/1e6:.2f}M</div>
                            <div class="metric-unit">ZAR</div></div>""", unsafe_allow_html=True)
                    with bc4:
                        st.markdown(f"""<div class="metric-card">
                            <div class="metric-label">Best IRR (%)</div>
                            <div class="metric-value">{best['IRR (%)']:.1f}%</div>
                            <div class="metric-unit">Project IRR</div></div>""", unsafe_allow_html=True)

                    # NPV
                    st.markdown(f"#### 🗺️ NPV Heatmap")
                    pivot = opt_df.pivot_table(index="BESS (kWh)", columns="PV (kWp)", values="NPV (ZAR)")
                    _pb, _plb, _fc, _gc, _lb = _plt()
                    _cs_mid = "#FFFFFF" if st.session_state.get("_light_mode", False) else "#111827"
                    fig_h = go.Figure(data=go.Heatmap(
                        z=pivot.values / 1e6, x=pivot.columns, y=pivot.index,
                        colorscale=[[0,"#FF4444"],[0.5,_cs_mid],[1,"#00E5A0"]],
                        colorbar=dict(title="NPV (M ZAR)"),
                        text=[[f"R{v:.1f}M" for v in row] for row in pivot.values / 1e6],
                        texttemplate="%{text}", textfont={"size": 8},
                    ))
                    fig_h.add_trace(go.Scatter(
                        x=[best["PV (kWp)"]], y=[best["BESS (kWh)"]],
                        mode="markers",
                        marker=dict(symbol="star", size=14, color="white",
                                    line=dict(color="#FF6B35", width=2)),
                        name="★ Optimal",
                    ))
                    fig_h.update_layout(
                        paper_bgcolor=_pb, plot_bgcolor=_plb,
                        font=dict(color=_fc, family="IBM Plex Mono"),
                        xaxis=dict(title="PV (kWp)", gridcolor=_gc),
                        yaxis=dict(title="BESS (kWh)", gridcolor=_gc),
                        height=480, margin=dict(l=10, r=10, t=20, b=10),
                    )
                    st.plotly_chart(fig_h, use_container_width=True)

                    st.markdown(f"#### 📊 Full Optimisation Results")
                    st.dataframe(opt_df.sort_values("NPV (ZAR)", ascending=False),
                                 use_container_width=True, height=380)

# ── Admin Tab (only visible to admins) ──────────────────────
if _admin_tab is not None:
    with _admin_tab:
        render_admin_panel()

# ─────────────────────────────────────────────────────────────
# Footer
# ─────────────────────────────────────────────────────────────
st.markdown("---")
st.markdown("""
<div style="text-align:center; color:var(--text-dim); font-family:'IBM Plex Mono',monospace;
            font-size:0.68rem; padding:0.8rem 0;">
    Powered by Huawei SA Digital Power — BD Kevin Yi &nbsp;|&nbsp;
    <a href="https://wa.me/27834976899?text=Hi%20Kevin%2C%20I%20would%20like%20more%20info%20on%20BTM%20PV%2BBESS%20solutions"
       target="_blank"
       style="color:#25D366; text-decoration:none; font-weight:600;">
       💬 WhatsApp: 083 497 6899
    </a>
    &nbsp;|&nbsp; SA Megaflex · PVGIS · Section 12B · 8760H Dispatch · Weekday/Weekend TOU
</div>
""", unsafe_allow_html=True)
