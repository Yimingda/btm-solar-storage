"""
ppa_wheeling.py — Scenario 2: PPA / Wheeling financial model
BTM PV+BESS Financial Modelling System

Reuses the BTM engine (PVGIS generation, tariff & escalation logic,
20-year cash-flow framework) injected from app.py via an `eng` dict:

    eng = {
        "get_pvgis_data":      PVGIS fetcher        (lat, lon, kwp, loss, tilt, az),
        "get_tariff_for_hour": Megaflex TOU lookup  (hour, month, day_type),
        "get_capex_zar":       USD→ZAR unit costs,
        "SECTION_12B":         {1:.5, 2:.3, 3:.2} SA accelerated depreciation,
        "ANALYSIS_YEARS":      20,
        "fmw":                 kW→MW auto-scale formatter,
    }

Two linked perspectives sharing one PPA / Wheeling term sheet:
  • Developer / IPP   — revenue = PPA price × billed energy;
                        costs = CAPEX, OPEX, insurance, (wheeling if borne);
                        outputs: 20-yr cash flow, IRR, NPV, payback.
  • End-user / Offtaker — savings = avoided grid cost − (PPA cost +
                        wheeling-if-borne + loss share); 20-yr cumulative model.

Both Excel (openpyxl) and PPTX (report_pptx.generate_ppa_pptx) reports are
generated per perspective.
"""
from __future__ import annotations

import io
import re
import datetime as _dt

import numpy as np
import pandas as pd
import streamlit as st

DEV = "Developer / IPP"
USR = "End-user / Offtaker"

# ─────────────────────────────────────────────────────────────
# Session-state defaults (whl_* namespace, seeded from BTM params)
# ─────────────────────────────────────────────────────────────

def _init_state(eng: dict) -> None:
    if st.session_state.get("_whl_initialized"):
        return
    ss = st.session_state
    pv_capex_zar, _bess_capex_zar = eng["get_capex_zar"]()
    _defaults = {
        # PV plant
        "whl_pv_kwp":          float(ss.get("pv_kwp", 4000.0) or 4000.0),
        "whl_spec_yield":      1800.0,            # kWh/kWp/yr (PVGIS can refresh)
        "whl_pv_degradation":  float(ss.get("pv_degradation", 0.5)),
        "whl_capex_per_kwp":   round(pv_capex_zar, 0),
        "whl_opex_per_kwp":    float(ss.get("pv_opex_per_kwp", 125.0)),
        "whl_insurance_pct":   0.5,               # % of CAPEX / yr
        # PPA terms (shared)
        "whl_ppa_price":       1.20,              # ZAR/kWh — single-price mode
        "whl_ppa_escalation":  6.0,               # %/yr
        "whl_contract_years":  20,
        # PPA pricing structure — three independent dimension toggles.
        # Each active combination gets its own price slider on the balancer
        # (keys created lazily as whl_p_pv_<season>_<period> / whl_p_bess_<season>).
        "whl_split_season":    False,             # distinguish winter / summer
        "whl_split_tou":       False,             # distinguish peak/std/off-peak
        "whl_split_asset":     False,             # distinguish PV / battery price
        # Battery leg (only used when the PV/battery asset dimension is on)
        "whl_bess_kwh":          float(ss.get("bess_kwh", 0.0) or 0.0),
        "whl_bess_cycles":       365.0,           # equivalent full cycles / yr
        "whl_bess_rte":          90.0,            # round-trip efficiency %
        "whl_bess_dod":          90.0,            # depth of discharge %
        "whl_bess_capex_per_kwh": round(_bess_capex_zar, 0),
        # Wheeling terms (shared)
        "whl_fee_mode":        "Per kWh (ZAR/kWh)",
        "whl_fee_kwh":         0.25,              # ZAR/kWh use-of-system
        "whl_fee_month":       50000.0,           # ZAR/month fixed
        "whl_loss_pct":        3.0,               # network loss %
        "whl_fee_borne_by":    "Developer",
        "whl_loss_borne_by":   "End-user",
        # Finance
        "whl_discount_rate":   float(ss.get("discount_rate", 10.0)),
        "whl_tax_rate":        float(ss.get("tax_rate", 27.0)),
        "whl_cost_escalation": 6.0,               # OPEX / insurance / wheeling CPI
        "whl_grid_escalation": float(ss.get("tariff_escalation", 8.76)),
        "whl_grid_tariff":     0.0,               # 0 → auto-fill from tariff table
    }
    for k, v in _defaults.items():
        ss.setdefault(k, v)
    # Auto-fill blended grid tariff from the BTM Megaflex tariff engine
    if not ss["whl_grid_tariff"]:
        ss["whl_grid_tariff"] = _blended_grid_tariff(eng)
    ss["_whl_initialized"] = True


def _blended_grid_tariff(eng: dict) -> float:
    """Blended daylight (07-17h) weekday grid rate from the BTM TOU tariff
    engine — the rate PPA energy displaces. Same tariff logic as BTM."""
    try:
        days = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
        tot, wsum = 0.0, 0.0
        for m in range(1, 13):
            for h in range(7, 18):
                rate, _ = eng["get_tariff_for_hour"](h, m, "weekday")
                tot  += rate * days[m - 1]
                wsum += days[m - 1]
        return round(tot / max(wsum, 1), 4)
    except Exception:
        return 2.50


_WINTER_MONTHS = (6, 7, 8)        # Eskom high season Jun–Aug = "winter"


def energy_split(eng: dict) -> dict:
    """
    PV generation fractions over (season, TOU-period) buckets, plus the
    battery season split. Season = Eskom high season (Jun–Aug = winter).
    Builds an 8760-h Gaussian daylight PV profile and classifies each weekday
    hour via the BTM tariff engine, so the split honours the active
    tariff_mode's peak windows. PV is daylight only → off-peak share ≈ 0.

    Returns {
      "pv":          {("win"|"sum", "peak"|"std"|"off"): fraction},  # Σ = 1.0
      "bess_season": {"win": f, "sum": f},                           # Σ = 1.0
    }
    """
    days = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    try:
        import numpy as _np
        hours = _np.arange(24)
        w = _np.exp(-0.5 * ((hours - 12) / 2.5) ** 2)   # daylight bell (as BTM)
        w[(hours < 6) | (hours > 18)] = 0.0
        if w.sum() <= 0:
            raise ValueError
        w /= w.sum()
        pv = {}
        for m in range(1, 13):
            season = "win" if m in _WINTER_MONTHS else "sum"
            for h in range(24):
                _, period = eng["get_tariff_for_hour"](h, m, "weekday")
                pk = ("peak" if "peak" in str(period) and "off" not in str(period)
                      else "off" if "off" in str(period) else "std")
                pv[(season, pk)] = pv.get((season, pk), 0.0) + w[h] * days[m - 1]
        tot = sum(pv.values()) or 1.0
        pv = {k: v / tot for k, v in pv.items()}
    except Exception:
        # Fallback: daylight PV split, no off-peak, ~25% winter (Jun–Aug)
        pv = {("win", "std"): 0.205, ("win", "peak"): 0.045,
              ("sum", "std"): 0.615, ("sum", "peak"): 0.135}
    win_days = sum(days[m - 1] for m in _WINTER_MONTHS)
    return {"pv": pv,
            "bess_season": {"win": win_days / 365.0,
                            "sum": 1.0 - win_days / 365.0}}


def price_buckets(split_season: bool, split_tou: bool, split_asset: bool):
    """
    Enumerate the active price axes for the given dimension toggles.
    Pure (no streamlit) — used by the param panel, the balancer sliders and
    the term-sheet builder so all three agree on keys.

    Returns (pv_axes, bess_axes):
      pv_axes  : list of (state_key, season_code, period_code, label)
      bess_axes: list of (state_key, season_code, label)   [empty if no asset]
    Battery is NOT split by TOU (it dispatches into peak); it only splits by
    season when the season toggle is on.
    """
    seasons = ([("win", "Winter"), ("sum", "Summer")] if split_season
               else [("all", "All-year")])
    periods = ([("peak", "Peak"), ("std", "Standard"), ("off", "Off-peak")]
               if split_tou else [("all", "All-day")])
    pv_axes = [
        (f"whl_p_pv_{s}_{pd}", s, pd,
         "PV" + (f" · {sl}" if s != "all" else "")
              + (f" · {pl}" if pd != "all" else ""))
        for (s, sl) in seasons for (pd, pl) in periods
    ]
    bess_axes = []
    if split_asset:
        bess_axes = [
            (f"whl_p_bess_{s}", s,
             "Battery" + (f" · {sl}" if s != "all" else " · dispatch"))
            for (s, sl) in seasons
        ]
    return pv_axes, bess_axes


def _default_price(base: float, season: str, period: str,
                   is_bess: bool = False) -> float:
    """Sensible seed price for a bucket: peak×1.4 / off×0.6, winter×1.15,
    battery dispatch premium ×1.8."""
    f = 1.0
    if season == "win":
        f *= 1.15
    if period == "peak":
        f *= 1.40
    elif period == "off":
        f *= 0.60
    if is_bess:
        f *= 1.80
    return round(base * f, 2)


# ─────────────────────────────────────────────────────────────
# Core model — 20-yr PPA/Wheeling cash flows, both perspectives
# (pure function: no streamlit access — unit-testable)
# ─────────────────────────────────────────────────────────────

def _npv_irr(cfs: list[float], disc: float) -> tuple[float, float]:
    """NPV + IRR (bisection) — mirrors BTM compute_npv_irr framework."""
    npv = sum(cf / (1 + disc) ** t for t, cf in enumerate(cfs))
    irr = 0.0
    try:
        lo, hi = -0.5, 5.0
        mid = 0.0
        for _ in range(120):
            mid = (lo + hi) / 2
            pv = sum(cf / (1 + mid) ** t for t, cf in enumerate(cfs))
            if abs(pv) < 0.5:
                break
            if pv > 0:
                lo = mid
            else:
                hi = mid
        irr = mid * 100
    except Exception:
        irr = 0.0
    return round(npv, 0), round(irr, 2)


def run_ppa_models(p: dict) -> dict:
    """
    Run both linked 20-yr models off one shared term sheet.

    p keys: pv_kwp, spec_yield, pv_degradation, capex_per_kwp, opex_per_kwp,
            insurance_pct, ppa_price, ppa_escalation, contract_years,
            fee_mode, fee_kwh, fee_month, loss_pct, fee_borne_by,
            loss_borne_by, discount_rate, tax_rate, cost_escalation,
            grid_escalation, grid_tariff, section_12b, service_fraction

    Pricing structure (optional — defaults to single flat price, identical
    to before).  Three independent dimension toggles combine into a price
    matrix; PV price axes = (2 if season)·(3 if TOU), battery = (2 if season):
      split_season / split_tou / split_asset : bool dimension toggles
      pv_prices   : {(season,period): ZAR/kWh}  season∈{win,sum,all} period∈{peak,std,off,all}
      bess_prices : {season: ZAR/kWh}           (asset only)
      energy_split: {"pv":{(season,period):frac}, "bess_season":{win,sum}}
      bess_kwh, bess_cycles, bess_rte, bess_dod, bess_capex_per_kwh : battery leg
    """
    yrs   = int(p["contract_years"])
    gen1  = p["pv_kwp"] * p["spec_yield"]
    loss  = p["loss_pct"] / 100.0
    deg   = p["pv_degradation"] / 100.0
    p_esc = p["ppa_escalation"] / 100.0
    g_esc = p["grid_escalation"] / 100.0
    c_esc = p["cost_escalation"] / 100.0
    disc  = p["discount_rate"] / 100.0
    tax   = p["tax_rate"] / 100.0

    # ── Pricing structure: toggles → blended year-1 PV / battery price ──
    split_season = bool(p.get("split_season", False))
    split_tou    = bool(p.get("split_tou", False))
    split_asset  = bool(p.get("split_asset", False))
    es           = p.get("energy_split") or {
        "pv": {("sum", "std"): 0.75, ("win", "std"): 0.25},
        "bess_season": {"win": 0.25, "sum": 0.75}}
    pv_fine   = es["pv"]
    pv_prices = p.get("pv_prices") or {}
    flat      = float(p.get("ppa_price", 1.20))

    skeys = ["win", "sum"] if split_season else ["all"]
    pkeys = ["peak", "std", "off"] if split_tou else ["all"]

    def _pv_frac(s, pd):
        seasons = ["win", "sum"] if s == "all" else [s]
        periods = ["peak", "std", "off"] if pd == "all" else [pd]
        return sum(pv_fine.get((ss, pp), 0.0) for ss in seasons for pp in periods)

    pv_price1 = sum((pv_prices.get((s, pd), flat)) * _pv_frac(s, pd)
                    for s in skeys for pd in pkeys)   # blended ZAR/kWh, Σfrac=1

    # ── Battery leg (asset dimension only) ─────────────────────────────
    if split_asset:
        bess_kwh   = float(p.get("bess_kwh", 0.0) or 0.0)
        bess_e1    = (bess_kwh * (p.get("bess_dod", 90) / 100.0)
                      * float(p.get("bess_cycles", 365))
                      * (p.get("bess_rte", 90) / 100.0))    # delivered kWh/yr
        bess_capex = bess_kwh * float(p.get("bess_capex_per_kwh", 0.0) or 0.0)
        bess_prices = p.get("bess_prices") or {}
        bseason     = es.get("bess_season", {"win": 0.25, "sum": 0.75})
        _bdef       = float(p.get("ppa_price_bess", flat * 1.8))
        bess_price1 = sum((bess_prices.get(s, _bdef))
                          * (1.0 if s == "all" else bseason.get(s, 0.5))
                          for s in skeys)
    else:
        bess_e1, bess_capex, bess_price1 = 0.0, 0.0, 0.0

    capex     = p["pv_kwp"] * p["capex_per_kwp"] + bess_capex
    svc_frac  = p.get("service_fraction", 0.40)
    dep_basis = capex * (1.0 - svc_frac)          # equipment portion only (12B)
    sec12b    = p.get("section_12b") or {1: 0.50, 2: 0.30, 3: 0.20}

    dev_rows, usr_rows = [], []
    dev_cfs, usr_cfs   = [-capex], [0.0]
    cum_dev, cum_usr   = -capex, 0.0
    payback = None

    for y in range(1, yrs + 1):
        pv_gen    = gen1 * (1 - deg) ** (y - 1)
        pv_deliv  = pv_gen * (1 - loss)
        # Billing point: if end-user bears losses they pay for generated
        # energy (loss share); if developer bears, billing = delivered.
        pv_billed = pv_gen if p["loss_borne_by"] == "End-user" else pv_deliv
        # Battery energy (no network loss share applied; billed = delivered)
        bess_deliv = bess_e1                         # 0 unless asset-split mode
        gen        = pv_gen + bess_deliv
        delivered  = pv_deliv + bess_deliv
        billed     = pv_billed + bess_deliv
        # Revenue per pricing model (escalation uniform across components)
        _esc_f = (1 + p_esc) ** (y - 1)
        revenue = (pv_price1 * pv_billed + bess_price1 * bess_deliv) * _esc_f
        price   = revenue / max(billed, 1e-9)        # blended display price

        if str(p["fee_mode"]).startswith("Per kWh"):
            wheel = delivered * p["fee_kwh"] * (1 + c_esc) ** (y - 1)
        else:
            wheel = p["fee_month"] * 12.0 * (1 + c_esc) ** (y - 1)
        wheel_dev = wheel if p["fee_borne_by"] == "Developer" else 0.0
        wheel_usr = wheel - wheel_dev

        opex = p["pv_kwp"] * p["opex_per_kwp"] * (1 + c_esc) ** (y - 1)
        ins  = capex * p["insurance_pct"] / 100.0 * (1 + c_esc) ** (y - 1)

        # ── Developer / IPP cash flow (BTM 20-yr framework) ─────────────
        ebitda   = revenue - opex - ins - wheel_dev
        dep      = dep_basis * sec12b.get(y, 0.0)
        cash_tax = (ebitda - dep) * tax     # negative in 12B years → tax shield
        ncf      = ebitda - cash_tax
        prev_cum = cum_dev
        cum_dev += ncf
        if payback is None and prev_cum < 0 <= cum_dev:
            payback = round((y - 1) + (-prev_cum) / max(ncf, 1e-9), 2)
        dev_cfs.append(ncf)
        dev_rows.append({
            "Year": y,
            "Generation (kWh)":     round(gen, 0),
            "Delivered (kWh)":      round(delivered, 0),
            "PPA Price (ZAR/kWh)":  round(price, 4),
            "Revenue (ZAR)":        round(revenue, 0),
            "O&M (ZAR)":            round(opex, 0),
            "Insurance (ZAR)":      round(ins, 0),
            "Wheeling (ZAR)":       round(wheel_dev, 0),
            "EBITDA (ZAR)":         round(ebitda, 0),
            "Depreciation (ZAR)":   round(dep, 0),
            "Tax (ZAR)":            round(cash_tax, 0),
            "Net Cash Flow (ZAR)":  round(ncf, 0),
            "Discounted CF (ZAR)":  round(ncf / (1 + disc) ** y, 0),
            "Cumulative CF (ZAR)":  round(cum_dev, 0),
        })

        # ── End-user / Offtaker savings model ───────────────────────────
        grid_rate  = p["grid_tariff"] * (1 + g_esc) ** (y - 1)
        baseline   = delivered * grid_rate          # grid bill avoided (BTM logic)
        loss_share = price * (billed - delivered)   # paid-but-not-received energy
        ppa_energy = revenue - loss_share           # price × delivered
        saving     = baseline - revenue - wheel_usr
        cum_usr   += saving
        usr_cfs.append(saving)
        usr_rows.append({
            "Year": y,
            "Delivered (kWh)":        round(delivered, 0),
            "Grid Rate (ZAR/kWh)":    round(grid_rate, 4),
            "Avoided Grid Cost (ZAR)": round(baseline, 0),
            "PPA Price (ZAR/kWh)":    round(price, 4),
            "PPA Energy Cost (ZAR)":  round(ppa_energy, 0),
            "Loss Share (ZAR)":       round(loss_share, 0),
            "Wheeling (ZAR)":         round(wheel_usr, 0),
            "Net Saving (ZAR)":       round(saving, 0),
            "Discounted Saving (ZAR)": round(saving / (1 + disc) ** y, 0),
            "Cumulative Saving (ZAR)": round(cum_usr, 0),
        })

    dev_df = pd.DataFrame(dev_rows)
    usr_df = pd.DataFrame(usr_rows)
    dev_npv, dev_irr = _npv_irr(dev_cfs, disc)
    usr_npv, _       = _npv_irr(usr_cfs, disc)

    y1 = dev_rows[0]
    u1 = usr_rows[0]
    base1 = max(u1["Avoided Grid Cost (ZAR)"], 1e-9)
    return {
        "dev_df": dev_df,
        "usr_df": usr_df,
        "capex": round(capex, 0),
        "dev_npv": dev_npv,
        "dev_irr": dev_irr,
        "dev_payback": payback,
        "dev_rev_yr1": y1["Revenue (ZAR)"],
        "usr_npv": usr_npv,
        "usr_saving_yr1": u1["Net Saving (ZAR)"],
        "usr_cum_saving": round(cum_usr, 0),
        "usr_discount_pct": round(100.0 * u1["Net Saving (ZAR)"] / base1, 1),
        "gen_yr1": y1["Generation (kWh)"],
        "delivered_yr1": y1["Delivered (kWh)"],
    }


# ─────────────────────────────────────────────────────────────
# Excel report — per perspective (pure function, no streamlit)
# ─────────────────────────────────────────────────────────────

def generate_ppa_excel(perspective: str, p: dict, model: dict,
                       project_name: str = "", client_name: str = "",
                       consultant_name: str = "") -> bytes:
    """3-sheet report (Cover · Cash-flow/Savings · Parameters) styled to
    match the BTM professional Excel report."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import LineChart, BarChart, Reference

    is_dev = perspective == DEV
    df     = model["dev_df"] if is_dev else model["usr_df"]

    C_NAVY, C_DARK, C_LTBLUE, C_ALT = "1F3864", "2E4057", "D6E4F0", "F2F6FC"
    C_GREEN, C_RED = "196F3D", "922B21"
    _thin = Side(style="thin", color="BDBDBD")

    def _fill(c): return PatternFill("solid", fgColor=c)
    def _font(bold=False, sz=10, color="1A1A2E", italic=False):
        return Font(name="Calibri", size=sz, bold=bold, color=color, italic=italic)
    def _align(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    def _bdr(): return Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

    wb  = Workbook()

    # ── Sheet 1: Cover ───────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Cover"
    ws1.sheet_view.showGridLines = False
    for ci, w in enumerate([2, 28, 22, 22, 22, 22, 2], 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w

    ws1.merge_cells("B2:G2")
    c = ws1["B2"]
    c.value = (f"PPA / Wheeling Financial Report — "
               f"{'Developer (IPP)' if is_dev else 'End-user (Offtaker)'} Perspective")
    c.font = Font("Calibri", size=15, bold=True, color="FFFFFF")
    c.fill = _fill(C_NAVY); c.alignment = _align("center")
    ws1.row_dimensions[2].height = 40

    ws1.merge_cells("B3:G3")
    c = ws1["B3"]
    c.value = "Powered by Huawei SA Digital Power · FusionSolar · Grid-Wheeled PPA Model"
    c.font = Font("Calibri", size=9, color="D6E4F0")
    c.fill = _fill(C_NAVY); c.alignment = _align("center")

    ws1.merge_cells("B5:G5")
    c = ws1["B5"]
    c.value = " Project Overview"
    c.font = Font("Calibri", size=10, bold=True, color="FFFFFF")
    c.fill = _fill(C_DARK)
    ws1.row_dimensions[5].height = 20

    _fee_str = (f"{p['fee_kwh']:.4f} ZAR/kWh"
                if str(p["fee_mode"]).startswith("Per kWh")
                else f"{p['fee_month']:,.0f} ZAR/month")
    info_rows = [
        ("Project",            project_name or "PPA Wheeling Project"),
        ("Client Name",        client_name or "—"),
        ("EPC / Consultant",   consultant_name or "—"),
        ("PV Capacity",        f"{p['pv_kwp']:,.0f} kWp"),
        ("Yr-1 Generation",    f"{model['gen_yr1']:,.0f} kWh  "
                               f"(delivered {model['delivered_yr1']:,.0f} kWh)"),
        ("PPA Price",          f"{p['ppa_price']:.4f} ZAR/kWh  "
                               f"(+{p['ppa_escalation']:.1f}%/yr)"),
        ("Contract Term",      f"{int(p['contract_years'])} years"),
        ("Wheeling Fee",       f"{_fee_str}  · borne by {p['fee_borne_by']}"),
        ("Wheeling Loss",      f"{p['loss_pct']:.1f}%  · borne by {p['loss_borne_by']}"),
        ("Grid Ref. Tariff",   f"{p['grid_tariff']:.4f} ZAR/kWh  "
                               f"(+{p['grid_escalation']:.2f}%/yr)"),
        ("Discount Rate",      f"{p['discount_rate']:.1f}%"),
        ("Export Time",        _dt.datetime.now().strftime("%Y-%m-%d  %H:%M")),
    ]
    for i, (lbl, val) in enumerate(info_rows):
        r = 6 + i
        ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        cl = ws1.cell(row=r, column=2, value=lbl)
        cl.font = _font(bold=True); cl.fill = _fill(C_LTBLUE)
        cl.alignment = _align(); cl.border = _bdr()
        ws1.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
        cv = ws1.cell(row=r, column=4, value=val)
        cv.font = _font(); cv.fill = _fill("FFFFFF")
        cv.alignment = _align(); cv.border = _bdr()
        ws1.row_dimensions[r].height = 18

    r0 = 6 + len(info_rows) + 1
    ws1.merge_cells(start_row=r0, start_column=2, end_row=r0, end_column=7)
    c = ws1.cell(row=r0, column=2, value=" Key Financial Metrics")
    c.font = Font("Calibri", size=10, bold=True, color="FFFFFF")
    c.fill = _fill(C_DARK)
    ws1.row_dimensions[r0].height = 20

    if is_dev:
        _pb = f"{model['dev_payback']:.2f} yr" if model["dev_payback"] else \
              f"{int(p['contract_years'])}yr+"
        kpis = [
            ("Total CAPEX",     f"R {model['capex']/1e6:.2f} M"),
            ("NPV",             f"R {model['dev_npv']/1e6:.2f} M"),
            ("Project IRR",     f"{model['dev_irr']:.2f}%"),
            ("Simple Payback",  _pb),
            ("Yr-1 Revenue",    f"R {model['dev_rev_yr1']/1e6:.2f} M"),
        ]
        _kpi_good = model["dev_npv"] > 0
    else:
        kpis = [
            ("Yr-1 Net Saving",       f"R {model['usr_saving_yr1']/1e6:.2f} M"),
            (f"{int(p['contract_years'])}-yr Cumulative Saving",
                                      f"R {model['usr_cum_saving']/1e6:.2f} M"),
            ("NPV of Savings",        f"R {model['usr_npv']/1e6:.2f} M"),
            ("Yr-1 Saving vs Grid",   f"{model['usr_discount_pct']:.1f}%"),
        ]
        _kpi_good = model["usr_cum_saving"] > 0
    for i, (lbl, val) in enumerate(kpis):
        r = r0 + 1 + i
        ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        cl = ws1.cell(row=r, column=2, value=lbl)
        cl.font = _font(bold=True); cl.fill = _fill(C_LTBLUE)
        cl.alignment = _align(); cl.border = _bdr()
        ws1.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
        cv = ws1.cell(row=r, column=4, value=val)
        cv.font = Font("Calibri", size=12, bold=True,
                       color=C_GREEN if _kpi_good else C_RED)
        cv.fill = _fill("FFFFFF"); cv.alignment = _align(); cv.border = _bdr()
        ws1.row_dimensions[r].height = 20

    # ── Sheet 2: 20-yr table + charts ────────────────────────
    ws2 = wb.create_sheet("Cash Flow" if is_dev else "Savings Model")
    ws2.freeze_panes = "A2"
    cols = list(df.columns)
    for ci, hdr in enumerate(cols, 1):
        hc = ws2.cell(row=1, column=ci, value=hdr)
        hc.font = Font("Calibri", size=9, bold=True, color="FFFFFF")
        hc.fill = _fill(C_NAVY); hc.alignment = _align("center", wrap=True)
        hc.border = _bdr()
        ws2.column_dimensions[get_column_letter(ci)].width = max(12, len(hdr) + 2)
    ws2.row_dimensions[1].height = 30
    for ri, rec in enumerate(df.itertuples(index=False), 2):
        for ci, v in enumerate(rec, 1):
            cell = ws2.cell(row=ri, column=ci, value=v)
            cell.font = _font(sz=9)
            cell.fill = _fill(C_ALT if ri % 2 == 0 else "FFFFFF")
            cell.border = _bdr()
            cell.number_format = "#,##0" if ci > 1 else "0"
            cell.alignment = _align("right")

    n = len(df)
    cum_col = cols.index("Cumulative CF (ZAR)" if is_dev
                         else "Cumulative Saving (ZAR)") + 1
    ncf_col = cols.index("Net Cash Flow (ZAR)" if is_dev
                         else "Net Saving (ZAR)") + 1
    cat = Reference(ws2, min_col=1, max_col=1, min_row=2, max_row=n + 1)
    ch1 = LineChart()
    ch1.title = ("Cumulative Cash Flow (ZAR)" if is_dev
                 else "Cumulative Savings (ZAR)")
    ch1.style = 10; ch1.width = 16; ch1.height = 9
    ch1.add_data(Reference(ws2, min_col=cum_col, max_col=cum_col,
                           min_row=1, max_row=n + 1), titles_from_data=True)
    ch1.set_categories(cat)
    ws2.add_chart(ch1, f"{get_column_letter(len(cols) + 2)}2")
    ch2 = BarChart()
    ch2.type = "col"
    ch2.title = ("Annual Net Cash Flow (ZAR)" if is_dev
                 else "Annual Net Saving (ZAR)")
    ch2.style = 10; ch2.width = 16; ch2.height = 9
    ch2.add_data(Reference(ws2, min_col=ncf_col, max_col=ncf_col,
                           min_row=1, max_row=n + 1), titles_from_data=True)
    ch2.set_categories(cat)
    ws2.add_chart(ch2, f"{get_column_letter(len(cols) + 2)}22")

    # ── Sheet 3: Parameters ──────────────────────────────────
    ws3 = wb.create_sheet("Parameters")
    ws3.sheet_view.showGridLines = False
    for ci, w in enumerate([3, 34, 18, 16], 1):
        ws3.column_dimensions[get_column_letter(ci)].width = w
    prm = [
        ("PV Capacity",              p["pv_kwp"],          "kWp"),
        ("Specific Yield (Yr-1)",    p["spec_yield"],      "kWh/kWp/yr"),
        ("PV Degradation",           p["pv_degradation"],  "%/yr"),
        ("CAPEX",                    p["capex_per_kwp"],   "ZAR/kWp"),
        ("O&M",                      p["opex_per_kwp"],    "ZAR/kWp/yr"),
        ("Insurance",                p["insurance_pct"],   "% CAPEX/yr"),
        ("PPA Price (Yr-1)",         p["ppa_price"],       "ZAR/kWh"),
        ("PPA Escalation",           p["ppa_escalation"],  "%/yr"),
        ("Contract Term",            p["contract_years"],  "years"),
        ("Wheeling Fee Mode",        p["fee_mode"],        ""),
        ("Wheeling Fee (per kWh)",   p["fee_kwh"],         "ZAR/kWh"),
        ("Wheeling Fee (fixed)",     p["fee_month"],       "ZAR/month"),
        ("Wheeling Loss",            p["loss_pct"],        "%"),
        ("Wheeling Fee Borne By",    p["fee_borne_by"],    ""),
        ("Loss Borne By",            p["loss_borne_by"],   ""),
        ("Grid Reference Tariff",    p["grid_tariff"],     "ZAR/kWh"),
        ("Grid Tariff Escalation",   p["grid_escalation"], "%/yr"),
        ("Cost Escalation (CPI)",    p["cost_escalation"], "%/yr"),
        ("Discount Rate",            p["discount_rate"],   "%"),
        ("Tax Rate",                 p["tax_rate"],        "%"),
    ]
    ws3.merge_cells("B1:D1")
    c = ws3.cell(row=1, column=2, value="PPA / Wheeling — Shared Term Sheet & Assumptions")
    c.font = Font("Calibri", size=12, bold=True, color="FFFFFF")
    c.fill = _fill(C_NAVY); c.alignment = _align("center")
    ws3.row_dimensions[1].height = 28
    for i, (lbl, val, unit) in enumerate(prm):
        r = 3 + i
        lb = ws3.cell(row=r, column=2, value=lbl)
        lb.font = _font(); lb.fill = _fill(C_ALT)
        lb.alignment = _align(); lb.border = _bdr()
        vc = ws3.cell(row=r, column=3, value=val)
        vc.alignment = _align("center"); vc.border = _bdr()
        un = ws3.cell(row=r, column=4, value=unit)
        un.font = _font(sz=9, italic=True, color="666666")
        un.fill = _fill(C_ALT); un.alignment = _align(); un.border = _bdr()
        ws3.row_dimensions[r].height = 17

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────
# UI
# ─────────────────────────────────────────────────────────────

def _kpi(label: str, value: str, sub: str = "") -> None:
    st.markdown(
        f'<div class="metric-card"><div class="metric-label">{label}</div>'
        f'<div class="metric-value">{value}</div>'
        f'<div class="metric-unit">{sub}</div></div>',
        unsafe_allow_html=True)


def _collect_bucket_prices(ss) -> tuple[dict, dict]:
    """
    Read the active price-bucket values from session state, seeding sensible
    defaults for any bucket not yet set. Returns (pv_prices, bess_prices)
    keyed by (season,period) / season for run_ppa_models.
    """
    base = float(ss.get("whl_ppa_price", 1.20) or 1.20)
    pv_axes, bess_axes = price_buckets(
        bool(ss.get("whl_split_season", False)),
        bool(ss.get("whl_split_tou", False)),
        bool(ss.get("whl_split_asset", False)))
    pv_prices, bess_prices = {}, {}
    for key, s, pd, _lbl in pv_axes:
        if key not in ss:
            ss[key] = _default_price(base, s, pd)
        pv_prices[(s, pd)] = float(ss[key])
    for key, s, _lbl in bess_axes:
        if key not in ss:
            ss[key] = _default_price(base, s, "peak", is_bess=True)
        bess_prices[s] = float(ss[key])
    return pv_prices, bess_prices


def _params_dict() -> dict:
    ss = st.session_state
    return {
        "pv_kwp":          ss.whl_pv_kwp,
        "spec_yield":      ss.whl_spec_yield,
        "pv_degradation":  ss.whl_pv_degradation,
        "capex_per_kwp":   ss.whl_capex_per_kwp,
        "opex_per_kwp":    ss.whl_opex_per_kwp,
        "insurance_pct":   ss.whl_insurance_pct,
        "ppa_price":       ss.whl_ppa_price,
        "ppa_escalation":  ss.whl_ppa_escalation,
        "contract_years":  ss.whl_contract_years,
        # Pricing structure — dimension toggles + per-bucket price matrix
        "split_season": bool(ss.get("whl_split_season", False)),
        "split_tou":    bool(ss.get("whl_split_tou", False)),
        "split_asset":  bool(ss.get("whl_split_asset", False)),
        "pv_prices":    _collect_bucket_prices(ss)[0],
        "bess_prices":  _collect_bucket_prices(ss)[1],
        # Battery leg (asset dimension)
        "bess_kwh":            ss.get("whl_bess_kwh", 0.0),
        "bess_cycles":         ss.get("whl_bess_cycles", 365.0),
        "bess_rte":            ss.get("whl_bess_rte", 90.0),
        "bess_dod":            ss.get("whl_bess_dod", 90.0),
        "bess_capex_per_kwh":  ss.get("whl_bess_capex_per_kwh", 0.0),
        "fee_mode":        ss.whl_fee_mode,
        "fee_kwh":         ss.whl_fee_kwh,
        "fee_month":       ss.whl_fee_month,
        "loss_pct":        ss.whl_loss_pct,
        "fee_borne_by":    ss.whl_fee_borne_by,
        "loss_borne_by":   ss.whl_loss_borne_by,
        "discount_rate":   ss.whl_discount_rate,
        "tax_rate":        ss.whl_tax_rate,
        "cost_escalation": ss.whl_cost_escalation,
        "grid_escalation": ss.whl_grid_escalation,
        "grid_tariff":     ss.whl_grid_tariff,
    }


def _render_balancer(p: dict, model: dict, eng: dict) -> None:
    """
    Interactive PPA-price balancer: one slider per price axis of the active
    pricing model. Dragging recomputes (on release) the developer return and
    the offtaker saving live, and plots the full price→outcome trade-off with
    a marker at the current price so the win-win band is visible.
    """
    import plotly.graph_objects as go
    ss = st.session_state

    split_season = bool(ss.get("whl_split_season", False))
    split_tou    = bool(ss.get("whl_split_tou", False))
    split_asset  = bool(ss.get("whl_split_asset", False))
    pv_axes, bess_axes = price_buckets(split_season, split_tou, split_asset)
    n_axes = len(pv_axes) + len(bess_axes)

    _dims = [d for d, on in (("season", split_season), ("TOU", split_tou),
                             ("PV/battery", split_asset)) if on]
    _dim_txt = " × ".join(_dims) if _dims else "single flat price"
    grid = float(p.get("grid_tariff", 1.5)) or 1.5
    st.markdown(
        f'<div class="section-header">⚖️ PPA Price Balancer — '
        f'{n_axes} price {"axis" if n_axes == 1 else "axes"} '
        f'({_dim_txt})</div>', unsafe_allow_html=True)
    st.caption(f"Grid reference tariff R {grid:.2f}/kWh · toggle dimensions in "
               "the right panel · drag a slider and release to recompute.")

    _smax = round(max(grid * 1.6, 3.0), 1)
    base  = float(ss.get("whl_ppa_price", 1.20) or 1.20)

    def _slider(key, lbl, seed):
        if key not in ss:
            ss[key] = seed
        ss[key] = min(max(float(ss[key]), 0.20), _smax)
        st.slider(f"{lbl} (ZAR/kWh)", 0.20, _smax, step=0.05, key=key)

    # ── PV price axes (grouped season × period) ──────────────────────
    st.markdown("**☀️ PV energy price**")
    for c, (key, s, pd, lbl) in zip(st.columns(len(pv_axes)), pv_axes):
        with c:
            _slider(key, lbl.replace("PV · ", "").replace("PV", "Flat"),
                    _default_price(base, s, pd))
    # ── Battery price axes (season only) ─────────────────────────────
    if bess_axes:
        st.markdown("**🔋 Battery dispatch price**")
        for c, (key, s, lbl) in zip(st.columns(len(bess_axes)), bess_axes):
            with c:
                _slider(key, lbl.replace("Battery · ", "").replace("Battery", "Dispatch"),
                        _default_price(base, s, "peak", is_bess=True))

    # ── Price-structure matrix (reflects the active dimensions) ──────
    es = p.get("energy_split") or {"pv": {("sum", "std"): 1.0},
                                    "bess_season": {"win": 0.25, "sum": 0.75}}
    _slbl = {"win": "Winter", "sum": "Summer", "all": "All-year"}
    _plbl = {"peak": "Peak", "std": "Standard", "off": "Off-peak", "all": "All-day"}

    def _pvf(s, pd):
        seas = ["win", "sum"] if s == "all" else [s]
        pers = ["peak", "std", "off"] if pd == "all" else [pd]
        return sum(es["pv"].get((ss_, pp), 0.0) for ss_ in seas for pp in pers)

    rows = [{"Axis": lbl, "Price (ZAR/kWh)": f"{float(ss.get(key, 0)):.2f}",
             "PV energy share": f"{_pvf(s, pd)*100:.0f}%"}
            for key, s, pd, lbl in pv_axes]
    for key, s, lbl in bess_axes:
        bf = 1.0 if s == "all" else es.get("bess_season", {}).get(s, 0.5)
        rows.append({"Axis": lbl, "Price (ZAR/kWh)": f"{float(ss.get(key, 0)):.2f}",
                     "PV energy share": f"(batt {bf*100:.0f}%)"})
    with st.expander("📋 Active price matrix", expanded=False):
        st.dataframe(__import__("pandas").DataFrame(rows),
                     use_container_width=True, hide_index=True)
        st.caption("One row per active price axis = "
                   f"{'season×' if split_season else ''}"
                   f"{'TOU×' if split_tou else ''}"
                   f"{'asset' if split_asset else 'flat'} structure. "
                   "Shares are the fraction of annual PV generation each axis prices.")

    # ── Live verdict KPIs: developer ↔ offtaker ──────────────────────
    dev_irr  = model["dev_irr"]
    dev_npv  = model["dev_npv"] / 1e6
    usr_disc = model["usr_discount_pct"]
    usr_cum  = model["usr_cum_saving"] / 1e6
    _pb = (f"{model['dev_payback']:.1f} yr" if model["dev_payback"]
           else f"{int(p['contract_years'])}yr+")
    cL, cR = st.columns(2)
    with cL:
        st.markdown('<div class="section-header">🏗 Developer / IPP</div>',
                    unsafe_allow_html=True)
        d1, d2, d3 = st.columns(3)
        with d1: _kpi("IRR", f"{dev_irr:.1f}%", "after-tax")
        with d2: _kpi("NPV", f"R {dev_npv:.1f}M", f"@{p['discount_rate']:.0f}%")
        with d3: _kpi("PAYBACK", _pb, "simple")
    with cR:
        st.markdown('<div class="section-header">🏭 End-user / Offtaker</div>',
                    unsafe_allow_html=True)
        u1, u2, u3 = st.columns(3)
        with u1: _kpi("YR-1 SAVING", f"{usr_disc:.1f}%", "vs grid bill")
        with u2: _kpi("CUM. SAVING", f"R {usr_cum:.1f}M", "nominal")
        with u3: _kpi("NPV", f"R {model['usr_npv']/1e6:.1f}M", "of savings")

    # Win-win verdict
    if dev_irr >= p["discount_rate"] and usr_disc > 0:
        _vc, _vt = "#00E5A0", "✅ WIN-WIN — developer clears hurdle rate and offtaker saves vs grid"
    elif dev_irr < p["discount_rate"] and usr_disc > 0:
        _vc, _vt = "#F6C90E", "⚠️ Offtaker-favoured — developer IRR below hurdle rate"
    elif dev_irr >= p["discount_rate"] and usr_disc <= 0:
        _vc, _vt = "#F6C90E", "⚠️ Developer-favoured — offtaker pays above grid parity"
    else:
        _vc, _vt = "#FF4444", "⛔ Lose-lose — price outside the viable band"
    st.markdown(
        f'<div style="border-left:4px solid {_vc};background:rgba(0,229,160,0.06);'
        f'padding:8px 14px;border-radius:4px;margin:6px 0 12px;color:var(--text-main);'
        f'font-family:IBM Plex Mono,monospace;font-size:0.85rem">{_vt}</div>',
        unsafe_allow_html=True)

    # ── Trade-off sweep: scale all PV bucket prices, x = blended PV price ──
    es = p.get("energy_split") or {"pv": {("sum", "std"): 1.0}}
    pv_fine = es["pv"]
    skeys = ["win", "sum"] if split_season else ["all"]
    pkeys = ["peak", "std", "off"] if split_tou else ["all"]

    def _frac(s, pd):
        seas = ["win", "sum"] if s == "all" else [s]
        pers = ["peak", "std", "off"] if pd == "all" else [pd]
        return sum(pv_fine.get((ss_, pp), 0.0) for ss_ in seas for pp in pers)

    pv_blend = sum(p["pv_prices"].get((s, pd), base) * _frac(s, pd)
                   for s in skeys for pd in pkeys) or base

    xs = [round(0.20 + i * (_smax - 0.20) / 40, 3) for i in range(41)]
    dev_y, usr_y = [], []
    for xv in xs:
        k = xv / max(pv_blend, 1e-6)
        q = dict(p)
        q["pv_prices"] = {b: pr * k for b, pr in p["pv_prices"].items()}
        m = run_ppa_models(q)
        dev_y.append(m["dev_irr"])
        usr_y.append(m["usr_cum_saving"] / 1e6)

    _dark = not ss.get("_light_mode", False)
    fig = go.Figure()
    fig.add_scatter(x=xs, y=dev_y, name="Developer IRR (%)",
                    line=dict(color="#00E5A0", width=2.4), yaxis="y")
    fig.add_scatter(x=xs, y=usr_y, name="Offtaker cum. saving (R M)",
                    line=dict(color="#4ECDC4", width=2.4, dash="dot"), yaxis="y2")
    fig.add_hline(y=p["discount_rate"], line=dict(color="#F6C90E", dash="dash"),
                  annotation_text=f"Hurdle {p['discount_rate']:.0f}%",
                  annotation_font_color="#F6C90E")
    fig.add_vline(x=pv_blend, line=dict(color="#FF6B35", width=2),
                  annotation_text=f"R {pv_blend:.2f}",
                  annotation_font_color="#FF6B35")
    fig.add_vline(x=grid, line=dict(color="#9CA3AF", dash="dot"),
                  annotation_text="grid", annotation_font_color="#9CA3AF")
    fig.update_layout(
        height=360, paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        font=dict(color="#E8ECF0" if _dark else "#1A202C",
                  family="IBM Plex Mono"),
        xaxis=dict(title="Blended PV PPA price (ZAR/kWh)"),
        yaxis=dict(title="Developer IRR (%)", side="left"),
        yaxis2=dict(title="Offtaker cum. saving (R M)", overlaying="y",
                    side="right"),
        legend=dict(orientation="h", y=-0.22),
        margin=dict(l=10, r=10, t=20, b=10))
    st.plotly_chart(fig, use_container_width=True)
    if n_axes > 1:
        st.caption("Sweep scales all PV price axes together (battery held "
                   "fixed); marker = current blended PV price. Adjust any "
                   "slider to move the curves.")


def render_ppa_wheeling(eng: dict) -> None:
    """Scenario-2 main page. `eng` = BTM engine functions injected by app.py."""
    _init_state(eng)
    ss  = st.session_state
    fmw = eng["fmw"]

    # ── Header ───────────────────────────────────────────────
    _b, _h, _x = st.columns([1.8, 6.2, 2], gap="small")
    with _b:
        if st.button("◀ Back to Scenarios", key="whl_back_btn2",
                     use_container_width=True):
            ss.pop("_scenario", None)
            st.rerun()
    with _h:
        st.markdown("""
<div class="main-header">
    <div style="font-size:1.8rem;line-height:1">🔄</div>
    <div>
        <div class="main-title">PPA / Wheeling Financial Model</div>
        <div class="sub-title">GRID-WHEELED PPA &nbsp;·&nbsp; DEVELOPER + OFFTAKER DUAL VIEW &nbsp;·&nbsp; BTM 20-YR ENGINE</div>
    </div>
</div>""", unsafe_allow_html=True)

    col_main, col_prm = st.columns([7, 3], gap="large")

    # ── Right: shared parameter panel ────────────────────────
    with col_prm:
        _scroll = st.container(height=860, border=False)
    with _scroll:
        # ── Site Location (interactive map, shared lat/lon with BTM) ──────
        with st.expander("📍 Site Location", expanded=True):
            ss.setdefault("lat", -26.1)
            ss.setdefault("lon", 28.0)
            ss.setdefault("tilt", 26.0)
            ss.setdefault("azimuth", 180.0)
            if eng.get("MAP_AVAILABLE"):
                try:
                    import folium
                    from streamlit_folium import st_folium
                    _m = folium.Map(
                        location=[ss.lat, ss.lon], zoom_start=7,
                        tiles="https://mt1.google.com/vt/lyrs=s&x={x}&y={y}&z={z}",
                        attr="Google Satellite")
                    folium.Marker(
                        [ss.lat, ss.lon],
                        popup=f"📍 {ss.lat:.4f}, {ss.lon:.4f}",
                        icon=folium.Icon(color="red", icon="map-pin",
                                         prefix="fa")).add_to(_m)
                    _md = st_folium(_m, width=None, height=200, key="whl_site_map")
                    if _md and _md.get("last_clicked"):
                        _la = _md["last_clicked"]["lat"]; _lo = _md["last_clicked"]["lng"]
                        if (abs(_la - ss.lat) > 1e-5 or abs(_lo - ss.lon) > 1e-5):
                            ss.lat, ss.lon = _la, _lo
                            ss["_whl_lat_in"] = _la; ss["_whl_lon_in"] = _lo
                            ss.tilt = round(abs(_la), 1)
                            ss.azimuth = 180.0 if _la < 0 else 0.0
                            st.rerun()
                except Exception as _e:
                    st.info(f"Map unavailable: {_e}")
            else:
                st.info("Install folium + streamlit-folium to enable the map")
            _lc, _oc = st.columns(2)
            with _lc:
                _nla = st.number_input("Latitude", -90.0, 90.0,
                                       value=float(ss.lat), format="%.4f",
                                       key="_whl_lat_in")
            with _oc:
                _nlo = st.number_input("Longitude", -180.0, 180.0,
                                       value=float(ss.lon), format="%.4f",
                                       key="_whl_lon_in")
            if abs(_nla - ss.lat) > 1e-5:
                ss.lat = _nla; ss.tilt = round(abs(_nla), 1)
                ss.azimuth = 180.0 if _nla < 0 else 0.0
            if abs(_nlo - ss.lon) > 1e-5:
                ss.lon = _nlo

        with st.expander("☀️ PV Plant & Generation", expanded=True):
            st.number_input("PV Capacity (kWp)", 0.0, 1e6,
                            key="whl_pv_kwp", step=100.0)
            # Generation source: PVGIS model vs uploaded 8760-h profile
            _src = st.radio(
                "Generation source",
                ["PVGIS model (lat/lon/tilt/azimuth)",
                 "Upload 8760-h profile"],
                key="whl_pv_src", label_visibility="collapsed")
            _csv = _src.startswith("Upload")
            if _csv and eng.get("parse_load_csv"):
                _up = st.file_uploader("PV generation 8760-h (kW)",
                                       type=["csv", "txt", "xlsx", "xls"],
                                       key="whl_pv_upload",
                                       label_visibility="collapsed")
                if _up is not None:
                    _arr, _msg = eng["parse_load_csv"](_up)
                    if _arr is not None and ss.whl_pv_kwp > 0:
                        ss.whl_spec_yield = round(float(_arr.sum()) / ss.whl_pv_kwp, 1)
                        st.success(f"Annual {_arr.sum()/1e3:,.0f} MWh → "
                                   f"yield {ss.whl_spec_yield:.0f} kWh/kWp")
                    elif _arr is not None:
                        st.warning("Set PV capacity first.")
                    else:
                        st.error(_msg)
            else:
                ss.setdefault("pv_loss", 14.0)
                st.number_input("System Loss (%)", 0.0, 50.0,
                                step=0.5, key="pv_loss")
                _t1, _t2 = st.columns(2)
                with _t1:
                    st.number_input("Tilt (°)", 0.0, 90.0,
                                    step=1.0, key="tilt")
                with _t2:
                    st.number_input("Azimuth (°)", -180.0, 180.0,
                                    step=5.0, key="azimuth")
                if st.button("↻ Fetch yield from PVGIS", key="whl_pvgis_btn",
                             use_container_width=True):
                    try:
                        pvg = eng["get_pvgis_data"](
                            ss.lat, ss.lon, ss.whl_pv_kwp,
                            ss.get("pv_loss", 14.0), ss.get("tilt", 26.0),
                            ss.get("azimuth", 180.0))
                        _akwh = float(pvg.get("annual_kwh", 0) or 0)
                        if _akwh > 0 and ss.whl_pv_kwp > 0:
                            ss.whl_spec_yield = round(_akwh / ss.whl_pv_kwp, 1)
                            st.rerun()
                        else:
                            st.warning("PVGIS returned no data — kept manual yield.")
                    except Exception as _e:
                        st.warning(f"PVGIS unavailable: {_e}")
            st.number_input("Specific Yield Yr-1 (kWh/kWp)", 500.0, 3000.0,
                            key="whl_spec_yield", step=10.0)
            st.number_input("PV Degradation (%/yr)", 0.0, 3.0,
                            key="whl_pv_degradation", step=0.1)
            st.markdown(
                f'<div class="derived-value">☀️ Yr-1 Gen: '
                f'{fmw(ss.whl_pv_kwp * ss.whl_spec_yield, "kWh/yr")}</div>',
                unsafe_allow_html=True)

        with st.expander("📜 PPA Terms (shared)", expanded=True):
            st.markdown("**PPA price structure** — combine any of:")
            st.checkbox("Distinguish winter / summer", key="whl_split_season")
            st.checkbox("Distinguish TOU (peak / standard / off-peak)",
                        key="whl_split_tou")
            st.checkbox("Distinguish PV / battery price", key="whl_split_asset")
            _ns = (2 if ss.get("whl_split_season") else 1)
            _nt = (3 if ss.get("whl_split_tou") else 1)
            _na = _ns * _nt + (_ns if ss.get("whl_split_asset") else 0)
            st.caption(f"⚖️ {_na} price "
                       f"{'axis' if _na == 1 else 'axes'} — drag on the "
                       "**Price Balancer** tab to set each price.")
            st.number_input("Base PPA Price (ZAR/kWh)", 0.0, 20.0,
                            key="whl_ppa_price", step=0.01, format="%.4f",
                            help="Seeds new price axes when you toggle a "
                                 "dimension; also the price when all toggles off.")
            st.number_input("PPA Escalation (%/yr)", 0.0, 25.0,
                            key="whl_ppa_escalation", step=0.25)
            st.number_input("Contract Term (years)", 1, 25,
                            key="whl_contract_years", step=1)
            if ss.get("whl_split_asset"):
                st.markdown("**🔋 Battery leg**")
                st.number_input("Battery Capacity (kWh)", 0.0, 1e6,
                                key="whl_bess_kwh", step=100.0)
                st.number_input("Equivalent Full Cycles / yr", 0.0, 730.0,
                                key="whl_bess_cycles", step=5.0)
                st.number_input("Round-trip Efficiency (%)", 50.0, 100.0,
                                key="whl_bess_rte", step=1.0)
                st.number_input("Depth of Discharge (%)", 10.0, 100.0,
                                key="whl_bess_dod", step=1.0)
                st.number_input("Battery CAPEX (ZAR/kWh)", 0.0, 1e5,
                                key="whl_bess_capex_per_kwh", step=100.0)

        with st.expander("🔌 Wheeling Terms (shared)", expanded=True):
            st.selectbox("Use-of-System Fee Basis",
                         ["Per kWh (ZAR/kWh)", "Fixed (ZAR/month)"],
                         key="whl_fee_mode")
            if str(ss.whl_fee_mode).startswith("Per kWh"):
                st.number_input("Wheeling Fee (ZAR/kWh)", 0.0, 5.0,
                                key="whl_fee_kwh", step=0.01, format="%.4f")
            else:
                st.number_input("Wheeling Fee (ZAR/month)", 0.0, 1e7,
                                key="whl_fee_month", step=1000.0)
            st.number_input("Wheeling Loss (%)", 0.0, 15.0,
                            key="whl_loss_pct", step=0.1)
            st.selectbox("Wheeling Fee Borne By",
                         ["Developer", "End-user"], key="whl_fee_borne_by")
            st.selectbox("Network Loss Borne By",
                         ["End-user", "Developer"], key="whl_loss_borne_by",
                         help="End-user → billed on generated energy (pays loss "
                              "share). Developer → billed on delivered energy.")

        with st.expander("💰 Developer Costs", expanded=False):
            st.number_input("CAPEX (ZAR/kWp)", 0.0, 1e5,
                            key="whl_capex_per_kwp", step=100.0)
            st.number_input("O&M (ZAR/kWp/yr)", 0.0, 2000.0,
                            key="whl_opex_per_kwp", step=5.0)
            st.number_input("Insurance (% CAPEX/yr)", 0.0, 5.0,
                            key="whl_insurance_pct", step=0.1)

        with st.expander("💡 Grid Tariff (baseline)", expanded=False):
            _tdb = eng.get("TARIFF_DB") or {}
            if _tdb:
                _modes = list(_tdb.keys())
                _cur = ss.get("tariff_mode", _modes[0])
                if _cur not in _modes:
                    _cur = _modes[0]
                _sel = st.selectbox("Eskom / Municipal Tariff Mode", _modes,
                                    index=_modes.index(_cur),
                                    key="whl_tariff_mode_sel",
                                    help="The grid tariff the wheeled PPA energy "
                                         "displaces — also drives the TOU split.")
                if _sel != ss.get("_whl_prev_tariff_mode"):
                    ss.tariff_mode = _sel
                    _rates = _tdb.get(_sel)
                    if _sel not in ("Custom (manual)",) and _rates:
                        w_pk, w_std, w_op, s_pk, s_std, s_op = _rates
                        ss.w_morning_peak = ss.w_evening_peak = w_pk
                        ss.w_standard = w_std; ss.w_off_peak = w_op
                        ss.s_morning_peak = ss.s_evening_peak = s_pk
                        ss.s_standard = s_std; ss.s_off_peak = s_op
                    ss._whl_prev_tariff_mode = _sel
                    ss.whl_grid_tariff = _blended_grid_tariff(eng)
                    st.rerun()
            st.number_input("Grid Reference Tariff (ZAR/kWh)", 0.0, 20.0,
                            key="whl_grid_tariff", step=0.05, format="%.4f",
                            help="Blended daylight grid rate the wheeled energy "
                                 "displaces — auto-seeded from the tariff table")
            if st.button("↻ Re-blend from tariff table", key="whl_blend_btn",
                         use_container_width=True):
                ss.whl_grid_tariff = _blended_grid_tariff(eng)
                st.rerun()

        with st.expander("🏦 Finance", expanded=False):
            st.number_input("Grid Tariff Escalation (%/yr)", 0.0, 25.0,
                            key="whl_grid_escalation", step=0.25)
            st.number_input("Cost Escalation CPI (%/yr)", 0.0, 25.0,
                            key="whl_cost_escalation", step=0.25)
            st.number_input("Discount Rate (%)", 0.0, 30.0,
                            key="whl_discount_rate", step=0.5)
            st.number_input("Tax Rate (%)", 0.0, 50.0,
                            key="whl_tax_rate", step=1.0)

    # ── Model run (cheap — every rerun) ──────────────────────
    p     = _params_dict()
    p["section_12b"]  = eng["SECTION_12B"]
    p["energy_split"] = energy_split(eng)
    model = run_ppa_models(p)
    ss["_whl_model"] = model

    # ── Left: dual-perspective tabs ──────────────────────────
    with col_main:
        tab_bal, tab_dev, tab_usr, tab_rpt = st.tabs(
            ["⚖️ Price Balancer", "🏗 Developer / IPP",
             "🏭 End-user / Offtaker", "📁 Reports"])

        import plotly.graph_objects as go

        # ── Price Balancer tab ───────────────────────────────
        with tab_bal:
            _render_balancer(p, model, eng)

        def _chart(df, bar_col, line_col, title):
            fig = go.Figure()
            fig.add_bar(x=df["Year"], y=df[bar_col], name=bar_col,
                        marker_color="#4ECDC4")
            fig.add_scatter(x=df["Year"], y=df[line_col], name=line_col,
                            mode="lines+markers", line=dict(color="#00E5A0"),
                            yaxis="y2")
            _dark = not st.session_state.get("_light_mode", False)
            fig.update_layout(
                title=title, height=380,
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                font=dict(color="#E8ECF0" if _dark else "#1A202C",
                          family="IBM Plex Mono"),
                yaxis=dict(title="ZAR/yr"),
                yaxis2=dict(title="Cumulative ZAR", overlaying="y", side="right"),
                legend=dict(orientation="h", y=-0.2),
                margin=dict(l=10, r=10, t=50, b=10))
            st.plotly_chart(fig, use_container_width=True)

        # ── Developer tab ────────────────────────────────────
        with tab_dev:
            _pb_str = (f"{model['dev_payback']:.2f} yr" if model["dev_payback"]
                       else f"{int(p['contract_years'])}yr+")
            k1, k2, k3, k4, k5 = st.columns(5)
            with k1: _kpi("TOTAL CAPEX",  f"R {model['capex']/1e6:.2f}M", "ZAR")
            with k2: _kpi("NPV",          f"R {model['dev_npv']/1e6:.2f}M",
                          f"@ {p['discount_rate']:.1f}%")
            with k3: _kpi("PROJECT IRR",  f"{model['dev_irr']:.2f}%", "after-tax")
            with k4: _kpi("PAYBACK",      _pb_str, "simple")
            with k5: _kpi("YR-1 REVENUE", f"R {model['dev_rev_yr1']/1e6:.2f}M",
                          "PPA sales")
            _chart(model["dev_df"], "Net Cash Flow (ZAR)",
                   "Cumulative CF (ZAR)",
                   f"Developer — {int(p['contract_years'])}-Year Cash Flow")
            st.dataframe(model["dev_df"], use_container_width=True,
                         hide_index=True, height=420)
            st.caption(
                f"Revenue = PPA price × "
                f"{'generated' if p['loss_borne_by'] == 'End-user' else 'delivered'} "
                f"energy · Costs = CAPEX / O&M / insurance"
                f"{' / wheeling fee' if p['fee_borne_by'] == 'Developer' else ''}"
                f" · Section 12B depreciation (50/30/20, equipment portion) · "
                f"Tax {p['tax_rate']:.0f}%")

        # ── End-user tab ─────────────────────────────────────
        with tab_usr:
            j1, j2, j3, j4 = st.columns(4)
            with j1: _kpi("YR-1 NET SAVING",
                          f"R {model['usr_saving_yr1']/1e6:.2f}M", "vs grid bill")
            with j2: _kpi(f"{int(p['contract_years'])}-YR CUM. SAVING",
                          f"R {model['usr_cum_saving']/1e6:.2f}M", "nominal")
            with j3: _kpi("NPV OF SAVINGS",
                          f"R {model['usr_npv']/1e6:.2f}M",
                          f"@ {p['discount_rate']:.1f}%")
            with j4: _kpi("YR-1 SAVING RATE",
                          f"{model['usr_discount_pct']:.1f}%", "of avoided cost")
            _chart(model["usr_df"], "Net Saving (ZAR)",
                   "Cumulative Saving (ZAR)",
                   f"End-user — {int(p['contract_years'])}-Year Cumulative Savings")
            st.dataframe(model["usr_df"], use_container_width=True,
                         hide_index=True, height=420)
            st.caption(
                "Saving = avoided grid cost (BTM tariff logic, "
                f"+{p['grid_escalation']:.2f}%/yr) − PPA energy cost − "
                f"{'loss share − ' if p['loss_borne_by'] == 'End-user' else ''}"
                f"{'wheeling fee' if p['fee_borne_by'] == 'End-user' else 'wheeling (borne by developer)'}")

        # ── Reports tab ──────────────────────────────────────
        with tab_rpt:
            st.markdown('<div class="section-header">📊 Financial Reports — per perspective</div>',
                        unsafe_allow_html=True)
            rc1, rc2 = st.columns(2)
            with rc1:
                st.text_input("EPC / Consultant",
                              value=ss.get("_pptx_consultant", ""),
                              placeholder="e.g. Lanxi Engineering",
                              key="_pptx_consultant")
            with rc2:
                st.text_input("Client Name",
                              value=ss.get("_pptx_client_name", ""),
                              placeholder="e.g. Lanxi Mining Company",
                              key="_pptx_client_name")
            _proj = ss.get("_active_snap_name") or "PPA_Wheeling_Project"
            _proj_safe = re.sub(r"[^\w\-]", "_", _proj).strip("_")
            _today = _dt.datetime.now().strftime("%Y%m%d")

            for _persp, _tag in ((DEV, "Developer"), (USR, "EndUser")):
                st.markdown(f"##### {'🏗' if _persp == DEV else '🏭'} {_persp}")
                b1, b2 = st.columns(2)
                with b1:
                    try:
                        _xb = generate_ppa_excel(
                            _persp, p, model,
                            project_name=_proj,
                            client_name=ss.get("_pptx_client_name", ""),
                            consultant_name=ss.get("_pptx_consultant", ""))
                        st.download_button(
                            f"⬇ Excel — {_tag} Report", data=_xb,
                            file_name=f"{_proj_safe}_PPA_{_tag}_{_today}.xlsx",
                            mime=("application/vnd.openxmlformats-officedocument"
                                  ".spreadsheetml.sheet"),
                            use_container_width=True,
                            key=f"whl_xl_{_tag}")
                    except Exception as _e:
                        st.error(f"Excel failed: {_e}")
                with b2:
                    if st.button(f"📊 Build PPTX — {_tag} Report",
                                 key=f"whl_pp_btn_{_tag}",
                                 use_container_width=True):
                        with st.spinner("Building PPTX…"):
                            try:
                                from report_pptx import generate_ppa_pptx
                                ss[f"_whl_pptx_{_tag}"] = generate_ppa_pptx(
                                    perspective=_persp, p=p, model=model,
                                    project_name=_proj,
                                    client_name=ss.get("_pptx_client_name", ""),
                                    consultant_name=ss.get("_pptx_consultant", ""))
                            except Exception as _e:
                                st.error(f"PPTX failed: {_e}")
                    if ss.get(f"_whl_pptx_{_tag}"):
                        st.download_button(
                            f"⬇ Download PPTX — {_tag}",
                            data=ss[f"_whl_pptx_{_tag}"],
                            file_name=f"{_proj_safe}_PPA_{_tag}_{_today}.pptx",
                            mime=("application/vnd.openxmlformats-officedocument"
                                  ".presentationml.presentation"),
                            use_container_width=True,
                            key=f"whl_pp_dl_{_tag}")
            st.caption("Both reports share one PPA / Wheeling term sheet — "
                       "developer revenue and end-user PPA cost are two sides "
                       "of the same contracted energy flows.")
